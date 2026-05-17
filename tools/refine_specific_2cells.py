"""针对两格定向补强:
1) 新加坡#43：附件已明确 NUS/NTU/SUSS 共 3 所，分母用 MOE 6 所公立大学
2) 柬埔寨#67：深度搜索柬华日报/柬中时报/孔院柬埔寨Facebook 关注量
"""
from __future__ import annotations
import asyncio, io, json, os, sys, logging
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError: pass

import httpx, openpyxl
from openpyxl.styles import PatternFill, Alignment
from news_monitor.proxy import build_httpx_client, get_proxies_for_url

TOPICS = ROOT / "topics"
SRC = TOPICS / "“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v4_final_20260420.xlsx"
DATE = datetime.now().strftime("%Y%m%d")
V5 = TOPICS / f"“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v5_{DATE}.xlsx"

PPX = os.environ.get("PERPLEXITY_API_KEY", "")
TAV = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
BRV = os.environ.get("BRAVEAPI", "")
DS = os.environ.get("DEEPSEEK_API_KEY", "")
PROXY = os.environ.get("LLM_PROXY", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("x")

async def cam67(tav_c, ppx_c, brv_c, ds_c):
    """柬埔寨#67 社媒中文账号关注量 — 深度定向"""
    # 多轮针对性搜索
    queries = [
        "柬埔寨 柬华日报 Facebook 粉丝 followers 关注",
        "柬埔寨 华商日报 华文媒体 社交媒体 粉丝",
        "Confucius Institute Cambodia RUPP Facebook followers",
        "柬中时报 Cambodia China Times Facebook 关注",
        "柬埔寨 中文 微信公众号 粉丝 订阅",
    ]
    ppx_answers = []
    # Perplexity 强针对性
    for q in queries[:3]:
        try:
            resp = await ppx_c.post(
                "https://api.perplexity.ai/chat/completions",
                json={
                    "model": "sonar",
                    "messages": [
                        {"role": "system", "content": "你是社交媒体统计研究员。列出柬埔寨本地中文媒体/机构/孔子学院的社交媒体账号及其关注量（粉丝数）。给出账号名称、平台、关注量数字（截止2024-2025）、链接URL。"},
                        {"role": "user", "content": q},
                    ],
                    "temperature": 0.05, "return_citations": True,
                },
                headers={"Authorization": f"Bearer {PPX}", "Content-Type": "application/json"},
                timeout=120.0,
            )
            if resp.status_code == 200:
                d = resp.json()
                ppx_answers.append((d["choices"][0]["message"]["content"], d.get("citations", [])[:8]))
                log.info("ppx query ok: %s", q[:40])
        except Exception as e:
            log.warning("ppx err: %s", str(e)[:100])

    # Tavily: 多社交平台
    tav_results = []
    for q in queries:
        try:
            resp = await tav_c.post(
                "https://api.tavily.com/search",
                json={"api_key": TAV, "query": q, "search_depth": "advanced",
                      "max_results": 10, "include_answer": False,
                      "include_domains": ["facebook.com", "twitter.com", "weibo.com",
                                         "jianshiapp.com", "cambodianchinese.com",
                                         "cambodiachinese.com", "khmertimeskh.com",
                                         "cen.com.kh", "ci.rupp.edu.kh"]},
                timeout=60.0,
            )
            if resp.status_code == 200:
                for r in (resp.json().get("results") or [])[:6]:
                    tav_results.append({"url": r.get("url", ""), "title": r.get("title", ""),
                                       "snippet": (r.get("content") or "")[:400]})
                log.info("tav query ok: %s", q[:40])
        except Exception as e:
            log.warning("tav err: %s", str(e)[:100])

    # Brave 广撒网
    brave_results = []
    for q in ["Cambodia Chinese media Facebook page followers",
              "柬华理事总会 Facebook 粉丝 关注 数量"]:
        try:
            resp = await brv_c.get(
                "https://api.search.brave.com/res/v1/web/search",
                params={"q": q, "count": 15, "text_decorations": False},
                headers={"Accept": "application/json", "X-Subscription-Token": BRV},
                timeout=30.0,
            )
            if resp.status_code == 200:
                for r in (resp.json().get("web", {}).get("results") or [])[:8]:
                    brave_results.append({"url": r.get("url", ""), "title": r.get("title", ""),
                                        "snippet": (r.get("description") or "")[:300]})
                log.info("brave query ok: %s", q[:40])
        except Exception as e:
            log.warning("brave err: %s", str(e)[:100])

    # DS 合成
    ppx_block = "\n---\n".join(
        f"## {q[:40]}\n{ans}\n\nCitations:\n" + "\n".join(cites)
        for q, (ans, cites) in zip(queries[:3], ppx_answers)
    )
    evid_block = "\n".join(
        f"[{i+1}] {r['title']}\nURL: {r['url']}\n{r['snippet']}"
        for i, r in enumerate((tav_results + brave_results)[:25])
    )

    prompt = f"""你是社交媒体统计研究员。请为【柬埔寨】【社交媒体平台中文账号关注量】指标合成最终值。

## Perplexity 多轮深度查询
{ppx_block}

## Tavily + Brave 网页搜索
{evid_block}

## 要求
1. 识别所有柬埔寨本地运营的中文账号（柬华日报、柬中时报、柬华理事总会、孔子学院柬埔寨、华文媒体 Facebook/微信公众号 等）
2. 列出每个账号的关注量（fans/followers）
3. 求和得出总关注量 + 单位（如"约 50 万人"）
4. 若某账号无数据，明确说"数据不可得"

返回 JSON:
{{
  "value": "总关注量（如'约 85 万人'）",
  "account_breakdown": [
    {{"name": "柬华日报 Facebook", "platform": "Facebook", "followers": "20万", "url": "..."}},
    ...
  ],
  "total_calculation": "A + B + C = X",
  "source_urls": ["主要来源1", "主要来源2"],
  "confidence": "high/medium/low",
  "is_still_estimation": true/false,
  "rationale": "30字"
}}
若真的完全搜不到数据，value 填 "无可核验数据"。
"""
    try:
        resp = await ds_c.post(
            "https://api.deepseek.com/v1/chat/completions",
            json={"model": "deepseek-chat", "messages": [{"role": "user", "content": prompt}],
                  "temperature": 0.05, "response_format": {"type": "json_object"}},
            headers={"Authorization": f"Bearer {DS}", "Content-Type": "application/json"},
            timeout=120.0,
        )
        resp.raise_for_status()
        return json.loads(resp.json()["choices"][0]["message"]["content"])
    except Exception as e:
        log.warning("ds err: %s", str(e)[:100])
        return {"value": "无可核验数据", "confidence": "low"}

async def run():
    ppx_p = get_proxies_for_url("https://api.perplexity.ai", PROXY)
    brv_p = get_proxies_for_url("https://api.search.brave.com", PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", PROXY)

    async with (build_httpx_client(ppx_p or "", 120) as ppx_c,
                build_httpx_client(brv_p or "", 30) as brv_c,
                build_httpx_client(tav_p or "", 60) as tav_c,
                build_httpx_client("", 120) as ds_c):
        log.info("柬埔寨#67 深度搜索...")
        cam_result = await cam67(tav_c, ppx_c, brv_c, ds_c)

    print("\n=== 柬埔寨#67 结果 ===")
    print(json.dumps(cam_result, ensure_ascii=False, indent=2))

    # 写入 v5
    wb = openpyxl.load_workbook(SRC)
    ws = wb.active
    est_fill = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")
    refined_fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")

    # 新加坡#43: 用附件数据硬写
    row_sg43 = next(r for r in range(2, ws.max_row+1) if ws.cell(r,1).value == 43)
    sg_val = "【估算】50.00%（计算：3 / 6 = 50% — 分子为 NUS/NTU/SUSS 三所开设中文相关本科主修或院系的高校，分母为 MOE 定义的 6 所公立自治大学 NUS/NTU/SMU/SUTD/SIT/SUSS，口径一致）"
    sg_src = "https://www.moe.gov.sg/post-secondary/overview/autonomous-universities"
    ws.cell(row_sg43, 13, sg_val).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row_sg43, 13).fill = refined_fill
    ws.cell(row_sg43, 14, sg_src)
    log.info("新加坡#43 已回填: 3/6 = 50%")

    # 柬埔寨#67
    row_cam67 = next(r for r in range(2, ws.max_row+1) if ws.cell(r,1).value == 67)
    val = cam_result.get("value", "无可核验数据")
    is_est = str(cam_result.get("is_still_estimation", True)).lower() == "true"
    total_calc = cam_result.get("total_calculation", "")
    if val and val != "无可核验数据":
        prefix = "【估算】" if is_est else ""
        cell_txt = f"{prefix}{val}" + (f"（计算：{total_calc}）" if total_calc else "")
        ws.cell(row_cam67, 9, cell_txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row_cam67, 9).fill = est_fill if is_est else refined_fill
        urls = cam_result.get("source_urls", [])
        if urls:
            ws.cell(row_cam67, 10, urls[0])
    log.info("柬埔寨#67 回填: %s", val[:60])

    wb.save(V5)
    log.info("saved %s", V5.name)

    # 保存详细日志
    detail = {"柬埔寨#67": cam_result, "新加坡#43": {"value": sg_val, "source": sg_src}}
    with open(ROOT / "output" / f"bri_v5_detail_{DATE}.json", "w", encoding="utf-8") as f:
        json.dump(detail, f, ensure_ascii=False, indent=2)

if __name__ == "__main__":
    asyncio.run(run())
