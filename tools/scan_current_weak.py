"""对 bri_64country_merged_{DATE}.xlsx 做低质量审计：

识别仍然模糊/0/无法获取的单元格，按严重程度分级：
  - hard_zero:   实际值为 0 且非分类字段
  - hard_wu:     实际值为 无/没有（非分类字段）
  - soft_fuzzy:  实际值以 少量/若干/有限 开头
  - explicit_fail: 实际值含 无可核验/无可靠数据/未公开/未披露 等
  - partial_range:  仅给区间而无精确值（如 5-10、约 X）— 可选标记
  - numeric_ok:  含具体数字
  - qualitative_ok: 扎实定性描述

输出：
  topics/bri_current_weak_audit_{DATE}.md
  output/bri_current_weak_targets_{DATE}.json
"""
from __future__ import annotations
import io, json, re, sys
from collections import defaultdict, Counter
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
TOPICS = ROOT / "topics"
OUTPUT = ROOT / "output"
DATE = datetime.now().strftime("%Y%m%d")
MERGED = TOPICS / f"bri_64country_merged_{DATE}.xlsx"
FIELD_DICT = TOPICS / "bri_field_dict.json"
OUT_MD = TOPICS / f"bri_current_weak_audit_{DATE}.md"
OUT_JSON = OUTPUT / f"bri_current_weak_targets_{DATE}.json"

import openpyxl

QUALITATIVE_IDS = {2, 3, 17, 51, 52, 64}

FAIL_TAGS = [
    "无可核验", "无可靠数据", "无确切", "无具体数据", "无官方", "无公开数据",
    "无统计数据", "无法核实", "无法确定", "无法计算", "无从查证",
    "未发布", "未公布", "未披露", "未见", "未找到", "未统计", "未知",
    "暂无", "尚无", "不公开", "不详", "待核", "N/A",
    "缺乏", "无直接", "无明确", "无系统",
]
SOFT_TERMS = ["少量", "若干", "有限", "极少", "零星"]


def classify(v_raw, iid):
    v = str(v_raw or "").strip()
    if not v:
        return "empty"
    # 剥所有前置【…】标签
    clean = re.sub(r"^(【[^】]+】\s*)+", "", v).strip()
    if not clean:
        return "explicit_fail"

    # 检查剥离的标签中是否含 FAIL_TAGS（如 【未公开】【无可核验】）
    prefix_text = v[: len(v) - len(clean)]
    if any(f in prefix_text for f in FAIL_TAGS):
        return "explicit_fail"

    first_seg = re.split(r"[（\(\n]", clean, maxsplit=1)[0].strip()
    first = re.split(r"[，,；;：:]", first_seg, maxsplit=1)[0].strip()

    # hard_zero
    if re.match(r"^0(\.0+)?(%|人|所|项|个|次|篇|家|门|部|册|台)?$", first) or first == "0":
        if iid in QUALITATIVE_IDS:
            return "qualitative_ok"
        return "hard_zero"

    # hard_wu
    if first in ("无", "没有", "无数据", "无明确数据", "无相关数据"):
        if iid in QUALITATIVE_IDS:
            return "qualitative_ok"
        if len(clean) > 80:  # 长推导的 "无" 视为有推理过程的定性
            return "qualitative_ok"
        return "hard_wu"

    # soft_fuzzy
    for s in SOFT_TERMS:
        if first.startswith(s):
            return "soft_fuzzy"

    # explicit_fail (value 开头就是失败标识)
    for f in FAIL_TAGS:
        if first.startswith(f) or first == f:
            return "explicit_fail"

    # partial_range: 仅有区间无精确值（X-Y 或 约 X）
    if re.match(r"^\d+\-\d+(%|人|所|项|个|次|篇)?$", first):
        return "partial_range"

    if any(c.isdigit() for c in first):
        return "numeric_ok"

    return "qualitative_ok"


def main():
    with open(FIELD_DICT, encoding="utf-8") as f:
        fields = [x for x in json.load(f) if x.get("指标ID")]
    fld_by_id = {int(x["指标ID"]): x for x in fields}

    wb = openpyxl.load_workbook(MERGED, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    country_cols = {}
    for c_idx, h in enumerate(headers, 1):
        if h and not str(h).endswith("_出处") and h not in (
            "序号", "搜集字段", "字段类型", "字段说明", "分类", "优先级",
        ):
            country_cols[h] = c_idx

    indicator_rows = {}
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, 1).value
        if iid is None:
            continue
        try:
            iid = int(iid)
        except (TypeError, ValueError):
            continue
        indicator_rows[iid] = r

    # 扫描
    weak_cells = []   # 低质量（hard_zero/hard_wu/soft_fuzzy/explicit_fail/empty）
    partial_cells = []
    all_stats = Counter()
    by_country = defaultdict(lambda: Counter())
    by_indicator = defaultdict(lambda: Counter())

    for country, vcol in country_cols.items():
        for iid, r in indicator_rows.items():
            v = ws.cell(r, vcol).value
            kind = classify(v, iid)
            all_stats[kind] += 1
            by_country[country][kind] += 1
            by_indicator[iid][kind] += 1
            if kind in ("hard_zero", "hard_wu", "soft_fuzzy", "explicit_fail", "empty"):
                fld = fld_by_id.get(iid, {})
                weak_cells.append({
                    "country": country,
                    "id": iid,
                    "name": fld.get("字段名", ""),
                    "type": fld.get("字段类型", ""),
                    "category": fld.get("分类", ""),
                    "priority": fld.get("优先级", ""),
                    "current": str(v) if v else "",
                    "kind": kind,
                })
            elif kind == "partial_range":
                partial_cells.append({
                    "country": country,
                    "id": iid,
                    "name": fld_by_id[iid].get("字段名", ""),
                    "current": str(v) if v else "",
                })

    print(f"总格数: {sum(all_stats.values())}")
    print(f"分类: {dict(all_stats)}")

    OUT_JSON.write_text(
        json.dumps({
            "weak_cells": weak_cells,
            "partial_cells": partial_cells,
            "stats": dict(all_stats),
        }, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"JSON: {OUT_JSON}")

    # ── MD 报告 ──
    tot = sum(all_stats.values())
    lines = []
    lines.append(f"# BRI 64 国数据质量审计 · 低质量点位清单")
    lines.append("")
    lines.append(f"**生成时间**：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"**底本**：`bri_64country_merged_{DATE}.xlsx`")
    lines.append(f"**总格数**：{tot}（64 国 × 68 指标）")
    lines.append("")
    lines.append(f"## 一、总体分布")
    lines.append("")
    lines.append(f"| 分级 | 数量 | 占比 | 说明 |")
    lines.append(f"|---|---|---|---|")
    legend = {
        "numeric_ok":   ("数字值", "含具体数字或区间（扎实）"),
        "qualitative_ok": ("定性OK", "扎实定性描述 / 分类字段的合法『无』 / 长推导"),
        "partial_range":("区间值", "仅有 X-Y 区间无精确数"),
        "soft_fuzzy":   ("软模糊", "以 少量/若干/有限/极少/零星 开头"),
        "hard_zero":    ("硬 0", "值为 0 且非分类字段"),
        "hard_wu":      ("硬 无", "值为 无/没有 且非分类字段"),
        "explicit_fail":("失败标记", "无可核验/未公开/未披露 等"),
        "empty":        ("空", "单元格为空"),
    }
    order = ["numeric_ok", "qualitative_ok", "partial_range",
             "soft_fuzzy", "hard_zero", "hard_wu", "explicit_fail", "empty"]
    for k in order:
        n = all_stats.get(k, 0)
        label, desc = legend[k]
        pct = n * 100 / tot if tot else 0
        lines.append(f"| {label} | {n} | {pct:.1f}% | {desc} |")
    weak_total = sum(all_stats.get(k, 0) for k in ["soft_fuzzy","hard_zero","hard_wu","explicit_fail","empty"])
    lines.append("")
    lines.append(f"**低质量合计（需关注）**：{weak_total} 格（{weak_total*100/tot:.1f}%）")
    lines.append("")

    # 按国家排名
    lines.append(f"## 二、按国家 · 低质量排名")
    lines.append("")
    lines.append(f"| 排名 | 国家 | 低质量数 | soft | hard_0 | hard_无 | 失败 | 空 |")
    lines.append(f"|---|---|---|---|---|---|---|---|")
    ranked = sorted(by_country.items(),
                    key=lambda x: -sum(x[1][k] for k in ("hard_zero","hard_wu","soft_fuzzy","explicit_fail","empty")))
    for i, (c, s) in enumerate(ranked, 1):
        weak_n = sum(s[k] for k in ("hard_zero","hard_wu","soft_fuzzy","explicit_fail","empty"))
        lines.append(f"| {i} | {c} | {weak_n} | {s['soft_fuzzy']} | {s['hard_zero']} | {s['hard_wu']} | {s['explicit_fail']} | {s['empty']} |")
    lines.append("")

    # 按指标排名
    lines.append(f"## 三、按指标 · 最难填 Top 20")
    lines.append("")
    lines.append(f"| 排名 | ID | 指标 | 分类 | 优先级 | 低质量国数 | soft | hard_0 | hard_无 | 失败 |")
    lines.append(f"|---|---|---|---|---|---|---|---|---|---|")
    ind_ranked = sorted(by_indicator.items(),
                        key=lambda x: -sum(x[1][k] for k in ("hard_zero","hard_wu","soft_fuzzy","explicit_fail","empty")))
    for i, (iid, s) in enumerate(ind_ranked[:20], 1):
        weak_n = sum(s[k] for k in ("hard_zero","hard_wu","soft_fuzzy","explicit_fail","empty"))
        fld = fld_by_id[iid]
        lines.append(f"| {i} | {iid} | {fld['字段名']} | {fld.get('分类','')} | {fld.get('优先级','')} | {weak_n} | {s['soft_fuzzy']} | {s['hard_zero']} | {s['hard_wu']} | {s['explicit_fail']} |")
    lines.append("")

    # 详细清单
    lines.append(f"## 四、详细清单（按国家 → 指标）")
    lines.append("")
    weak_cells.sort(key=lambda x: (x["country"], x["id"]))
    cur_country = None
    for c in weak_cells:
        if c["country"] != cur_country:
            cur_country = c["country"]
            n = by_country[cur_country]
            weak_n = sum(n[k] for k in ("hard_zero","hard_wu","soft_fuzzy","explicit_fail","empty"))
            lines.append(f"### {cur_country}（{weak_n} 格低质量）")
            lines.append("")
            lines.append(f"| ID | 字段名 | 分类 | 优先级 | 类型 | 当前值 | 分级 |")
            lines.append(f"|---|---|---|---|---|---|---|")
        cur_v = c["current"].replace("|", "\\|").replace("\n", " ")
        if len(cur_v) > 80:
            cur_v = cur_v[:80] + "…"
        kind_label = legend.get(c["kind"], (c["kind"], ""))[0]
        lines.append(f"| {c['id']} | {c['name']} | {c['category']} | {c['priority']} | {c['type']} | {cur_v} | {kind_label} |")

    lines.append("")
    lines.append(f"## 五、区间值（partial_range）提示")
    lines.append("")
    lines.append(f"这些格子给出了区间但无精确值（如 5-10），属于『有数据但需进一步精确化』：")
    lines.append("")
    if partial_cells:
        lines.append(f"| 国家 | ID | 字段名 | 当前值 |")
        lines.append(f"|---|---|---|---|")
        partial_cells.sort(key=lambda x: (x["country"], x["id"]))
        for c in partial_cells[:150]:
            cur_v = c["current"].replace("|", "\\|").replace("\n", " ")[:60]
            lines.append(f"| {c['country']} | {c['id']} | {c['name']} | {cur_v} |")
        if len(partial_cells) > 150:
            lines.append("")
            lines.append(f"...（共 {len(partial_cells)} 条，仅列前 150）")
    lines.append("")

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"MD: {OUT_MD}")


if __name__ == "__main__":
    main()
