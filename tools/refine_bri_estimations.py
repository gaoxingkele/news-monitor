"""对 19 个估算格做第二轮强化研究，目标：
- 每个数字带独立出处（分子、分母、计算式）
- 交叉验证 3 个来源（官方/学术/国际组织）
- 给出不确定区间或多版本对齐
- 如仍无法避免估算，将置信度升级为可论证的估算

产出: ..._v3_{date}.xlsx 覆盖这 19 格，追加 estimation_notes_refined_{date}.md
"""
from __future__ import annotations
import asyncio, io, json, os, sys, time, logging
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))

try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass

import httpx, openpyxl
from openpyxl.styles import PatternFill, Alignment
from news_monitor.proxy import build_httpx_client, get_proxies_for_url

TOPICS = ROOT / "topics"
DATE_TAG = datetime.now().strftime("%Y%m%d")
SRC = TOPICS / "“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v2_20260420.xlsx"
V3_OUT = TOPICS / f"“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v3_{DATE_TAG}.xlsx"
REFINED_MD = TOPICS / f"bri_5country_estimation_refined_{DATE_TAG}.md"
CKPT = ROOT / "output" / "bri_refine_checkpoint.json"
FIELD_DICT = TOPICS / "bri_field_dict.json"
LOG_FILE = ROOT / "output" / f"bri_refine_{DATE_TAG}.log"

PERPLEXITY_API_KEY = os.environ.get("PERPLEXITY_API_KEY", "")
BRAVE_API_KEY = os.environ.get("BRAVEAPI", "")
TAVILY_API_KEY = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
OVERSEAS_PROXY = os.environ.get("LLM_PROXY", "") or os.environ.get("OVERSEAS_PROXY", "")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.FileHandler(LOG_FILE, encoding='utf-8'), logging.StreamHandler()],
)
log = logging.getLogger("refine")

# 19 估算格 + 国家 col 映射
COUNTRY_COL = {"蒙古": (5, 6), "越南": (7, 8), "柬埔寨": (9, 10),
               "菲律宾": (11, 12), "新加坡": (13, 14)}
COUNTRY_EN = {"蒙古": "Mongolia", "越南": "Vietnam", "柬埔寨": "Cambodia",
              "菲律宾": "Philippines", "新加坡": "Singapore"}

TARGETS = [
    ("蒙古", 10), ("蒙古", 43), ("蒙古", 47), ("蒙古", 60),
    ("越南", 30), ("越南", 43),
    ("柬埔寨", 30), ("柬埔寨", 32), ("柬埔寨", 57), ("柬埔寨", 59), ("柬埔寨", 67),
    ("菲律宾", 12), ("菲律宾", 32), ("菲律宾", 41),
    ("新加坡", 12), ("新加坡", 27), ("新加坡", 31), ("新加坡", 35), ("新加坡", 43),
]

# 学术/国际组织 + 官方权威域名
ACADEMIC_DOMAINS = [
    "cnki.com.cn", "hanspub.org", "scholar.google.com",
    "sciencedirect.com", "jstor.org", "researchgate.net", "academia.edu",
    "taylorfrancis.com", "springer.com", "cambridge.org", "oup.com",
]
INTL_ORG_DOMAINS = [
    "worldbank.org", "unesco.org", "oecd.org", "un.org", "uis.unesco.org",
    "imf.org", "weforum.org", "adb.org",
]
HANBAN_DOMAINS = [
    "ci.cn", "chinese.cn", "hanban.org", "confucius.edu.cn",
    "bridge.chinese.cn", "clef.org.cn",
]

@dataclass
class Evidence:
    source: str
    url: str = ""
    title: str = ""
    snippet: str = ""

# ── API (enhanced, multi-round) ─────────────────────────────────────────────
async def ppx_deep(client: httpx.AsyncClient, country: str, ind_name: str,
                   ind_type: str, ind_desc: str) -> tuple[str, list[str]]:
    """深度 Perplexity 查询：要求分子/分母独立出处。"""
    system = (
        "You are an education-policy statistician. For every numeric answer, "
        "output the numerator value + its source URL, the denominator value + its "
        "source URL, the computation formula, and an uncertainty range if applicable. "
        "Prefer: official government statistics, UNESCO/World Bank, Hanban/CLEC, "
        "peer-reviewed papers. If multiple years differ, list each year separately. "
        "If the true value is unknown but proxies exist, explicitly label them as proxies."
    )
    user = (
        f"国家：{country}\n"
        f"指标：{ind_name}\n"
        f"取值口径：{ind_type}\n"
        f"字段说明：{ind_desc}\n\n"
        f"请提供：\n"
        f"1. 最权威的数值答案（按取值口径）。若需比率，分子与分母必须分别给出独立来源 URL。\n"
        f"2. 计算式：A / B = C（完整代入）。\n"
        f"3. 不确定区间：若不同口径/年份差异显著，给出 [low, high]。\n"
        f"4. 若无直接数据，找 3 种以上代理指标替代，并说明每种代理的边界。\n"
        f"5. 所有来源 URL 必须可直接访问。\n"
        f"用中文回答。"
    )
    body = {
        "model": "sonar",
        "messages": [{"role": "system", "content": system}, {"role": "user", "content": user}],
        "temperature": 0.05,
        "return_citations": True,
    }
    try:
        resp = await client.post(
            "https://api.perplexity.ai/chat/completions",
            json=body,
            headers={"Authorization": f"Bearer {PERPLEXITY_API_KEY}", "Content-Type": "application/json"},
            timeout=90.0,
        )
        resp.raise_for_status()
        d = resp.json()
        return d["choices"][0]["message"]["content"], d.get("citations", [])[:12]
    except Exception as exc:
        log.warning("ppx deep err: %s", str(exc)[:120])
        return "", []

async def tavily_domains(client: httpx.AsyncClient, query: str, domains: list[str]) -> list[Evidence]:
    if not TAVILY_API_KEY: return []
    try:
        resp = await client.post(
            "https://api.tavily.com/search",
            json={
                "api_key": TAVILY_API_KEY, "query": query, "search_depth": "advanced",
                "max_results": 10, "include_domains": domains[:20],
                "include_answer": False, "include_raw_content": False,
            },
            timeout=60.0,
        )
        resp.raise_for_status()
        return [Evidence(source="tavily", url=r.get("url",""), title=r.get("title",""),
                         snippet=(r.get("content") or "")[:500])
                for r in (resp.json().get("results") or [])[:8]]
    except Exception as exc:
        log.warning("tav err: %s", str(exc)[:120])
        return []

async def brave_search(client: httpx.AsyncClient, query: str) -> list[Evidence]:
    if not BRAVE_API_KEY: return []
    try:
        resp = await client.get(
            "https://api.search.brave.com/res/v1/web/search",
            params={"q": query, "count": 15, "text_decorations": False},
            headers={"Accept": "application/json", "X-Subscription-Token": BRAVE_API_KEY},
            timeout=30.0,
        )
        resp.raise_for_status()
        return [Evidence(source="brave", url=r.get("url",""), title=r.get("title",""),
                         snippet=(r.get("description") or "")[:400])
                for r in (resp.json().get("web",{}).get("results") or [])[:10]]
    except Exception as exc:
        log.warning("brave err: %s", str(exc)[:120])
        return []

async def ds_synthesize(client: httpx.AsyncClient, country: str, ind_name: str,
                        ind_type: str, ind_desc: str, ppx_answer: str,
                        evidence: list[Evidence]) -> dict:
    """用 DeepSeek 合成最终结果，要求输出结构化 {numerator, denominator, computation, value, uncertainty_range, confidence, sources}."""
    evid_block = "\n\n".join(
        f"[{i+1}] ({e.source}) {e.title}\nURL: {e.url}\n摘要: {e.snippet}"
        for i, e in enumerate(evidence[:15])
    )
    prompt = f"""你是严谨的统计研究员。为【{country}】在指标【{ind_name}】上做最终数值合成。

## 指标信息
- 取值口径：{ind_type}
- 字段说明：{ind_desc}

## Perplexity 初查
{ppx_answer or '(空)'}

## 扩证证据
{evid_block or '(无)'}

## 输出要求
严格返回 JSON（无多余文本）：
{{
  "value": "最终回答（格式贴合取值口径。若为比值给到小数点后2位；若有不确定性则给区间如'1.2%-1.8%'）",
  "numerator": "分子数值 + 口径说明（含年份）",
  "numerator_url": "分子出处 URL",
  "denominator": "分母数值 + 口径说明（含年份），若无则写 N/A",
  "denominator_url": "分母出处 URL 或 N/A",
  "computation": "计算式，形如'87447 / 16520000 = 0.529%' 或 'N/A'",
  "uncertainty_range": "[low, high] 或 'N/A'",
  "confidence": "high / medium / low",
  "is_still_estimation": "true/false —— 是否仍然依赖估算而非原始单一官方数据",
  "rationale": "50字内说明为什么是这个值、如何处理冲突源",
  "alt_sources": ["备选出处URL1", "备选出处URL2"]
}}

注意：
- 即便是估算，也必须给出可复现的计算式。
- 若多个年份/口径冲突，取"最新官方 > 最新学术 > 国际组织"优先级。
- 若真的无法找到任何数据，value 写 "无可核验数据"，confidence='low'。
"""
    try:
        resp = await client.post(
            "https://api.deepseek.com/v1/chat/completions",
            json={
                "model": "deepseek-chat",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.05,
                "response_format": {"type": "json_object"},
            },
            headers={"Authorization": f"Bearer {DEEPSEEK_API_KEY}", "Content-Type": "application/json"},
            timeout=90.0,
        )
        resp.raise_for_status()
        return json.loads(resp.json()["choices"][0]["message"]["content"])
    except Exception as exc:
        log.warning("ds err: %s", str(exc)[:120])
        return {"value": (ppx_answer or "")[:300], "confidence": "low",
                "is_still_estimation": True, "rationale": "LLM 失败"}

# ── pipeline per target ─────────────────────────────────────────────────────
async def refine_one(sem, country, iid, ind, ppx_c, brave_c, tav_c, ds_c, ckpt):
    key = f"{country}::{iid}"
    if key in ckpt:
        log.info("cached %s", key)
        return ckpt[key]

    async with sem:
        ind_name, ind_type, ind_desc = ind["字段名"], ind["字段类型"], ind["字段说明"]
        en_country = COUNTRY_EN[country]

        # Perplexity deep
        ppx_ans, ppx_cites = await ppx_deep(ppx_c, country, ind_name, ind_type, ind_desc)

        # Tavily: 3 轮（学术 / 国际组织 / 官方汉办）
        q_academic = f"{en_country} Chinese language education {ind_name} statistics"
        q_intl = f"{en_country} population education statistics {ind_name}"
        q_hanban = f"{country} {ind_name} 孔子学院 统计"
        tav_acad, tav_intl, tav_hb = await asyncio.gather(
            tavily_domains(tav_c, q_academic, ACADEMIC_DOMAINS),
            tavily_domains(tav_c, q_intl, INTL_ORG_DOMAINS),
            tavily_domains(tav_c, q_hanban, HANBAN_DOMAINS),
            return_exceptions=True,
        )
        # Brave：广撒网
        brave_r = await brave_search(brave_c, f"{country} {ind_name} 数据 2024 2025")

        evidence = []
        for url in ppx_cites:
            evidence.append(Evidence(source="perplexity", url=url, title="", snippet=""))
        for r in (tav_acad, tav_intl, tav_hb):
            if isinstance(r, list):
                evidence.extend(r)
        evidence.extend(brave_r[:6])

        # DeepSeek 合成
        result = await ds_synthesize(ds_c, country, ind_name, ind_type, ind_desc,
                                      ppx_ans, evidence)

        result["evidence_count"] = len(evidence)
        result["evidence_sample"] = [asdict(e) for e in evidence[:10]]
        ckpt[key] = result
        with open(CKPT, "w", encoding="utf-8") as f:
            json.dump(ckpt, f, ensure_ascii=False, indent=1)

        is_est = result.get("is_still_estimation", True)
        if isinstance(is_est, str): is_est = is_est.lower() == "true"
        marker = "✓" if not is_est else "[仍估算]"
        log.info("%s %s#%d %s conf=%s value=%s",
                 marker, country, iid, ind_name[:20],
                 result.get("confidence","?"), str(result.get("value",""))[:50])
        return result

async def run():
    with open(FIELD_DICT, encoding='utf-8') as f:
        fields = {int(x["指标ID"]): x for x in json.load(f)}

    ckpt = {}
    if CKPT.exists():
        try: ckpt = json.loads(CKPT.read_text(encoding='utf-8'))
        except Exception: ckpt = {}
    log.info("ckpt: %d cached", len(ckpt))

    ppx_p = get_proxies_for_url("https://api.perplexity.ai", OVERSEAS_PROXY)
    brv_p = get_proxies_for_url("https://api.search.brave.com", OVERSEAS_PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", OVERSEAS_PROXY)

    async with (build_httpx_client(ppx_p or "", 90) as ppx_c,
                build_httpx_client(brv_p or "", 30) as brv_c,
                build_httpx_client(tav_p or "", 60) as tav_c,
                build_httpx_client("", 90) as ds_c):
        sem = asyncio.Semaphore(4)
        t0 = time.monotonic()
        tasks = [
            refine_one(sem, c, i, fields[i], ppx_c, brv_c, tav_c, ds_c, ckpt)
            for c, i in TARGETS
        ]
        results = await asyncio.gather(*tasks)
        log.info("refine done in %.1fs", time.monotonic()-t0)

    # 写 v3 xlsx
    wb = openpyxl.load_workbook(SRC)
    ws = wb.active
    refined_fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")
    est_fill = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")

    refined_md = [f"# 估算指标补强报告 (v3 refinement)",
                  f"生成时间：{datetime.now():%Y-%m-%d %H:%M}",
                  f"对 {len(TARGETS)} 个估算格做二次深度研究，每格给出分子/分母/计算式/不确定区间。\n"]

    for (c, iid), r in zip(TARGETS, results):
        row_num = next(i for i in range(2, ws.max_row+1) if ws.cell(i,1).value == iid)
        vcol, scol = COUNTRY_COL[c]
        fname = fields[iid]["字段名"]
        val = str(r.get("value", "")).strip()
        if not val:
            continue
        is_est = r.get("is_still_estimation", True)
        if isinstance(is_est, str): is_est = is_est.lower() == "true"
        prefix = "【估算】" if is_est else ""
        computation = r.get("computation", "")
        if computation and computation.upper() != "N/A":
            cell_v = f"{prefix}{val}（计算：{computation}）"
        else:
            cell_v = f"{prefix}{val}"
        ws.cell(row_num, vcol, cell_v).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row_num, vcol).fill = est_fill if is_est else refined_fill
        # 选最佳出处
        best_url = r.get("numerator_url") or (r.get("alt_sources") or [""])[0] or ""
        if best_url and best_url != "N/A":
            ws.cell(row_num, scol, best_url)

        refined_md.append(f"\n## {c} #{iid} {fname}")
        refined_md.append(f"- **值**：{val}{'（仍估算）' if is_est else '（已落实）'}")
        if r.get("numerator") and r.get("numerator") != "N/A":
            refined_md.append(f"- **分子**：{r['numerator']}  \n  URL：{r.get('numerator_url','')}")
        if r.get("denominator") and r.get("denominator") != "N/A":
            refined_md.append(f"- **分母**：{r['denominator']}  \n  URL：{r.get('denominator_url','')}")
        if computation and computation != "N/A":
            refined_md.append(f"- **计算式**：{computation}")
        if r.get("uncertainty_range") and r.get("uncertainty_range") != "N/A":
            refined_md.append(f"- **不确定区间**：{r['uncertainty_range']}")
        refined_md.append(f"- **置信度**：{r.get('confidence','low')}")
        refined_md.append(f"- **理由**：{r.get('rationale','')}")
        alt = r.get("alt_sources") or []
        if alt:
            refined_md.append(f"- **备选出处**：")
            for a in alt[:3]:
                if a: refined_md.append(f"  - {a}")

    wb.save(V3_OUT)
    REFINED_MD.write_text("\n".join(refined_md), encoding="utf-8")
    log.info("saved %s", V3_OUT.name)
    log.info("saved %s", REFINED_MD.name)

    # stats
    still_est = sum(1 for r in results
                    if (str(r.get("is_still_estimation", True)).lower() == "true"))
    log.info("DONE: %d refined, %d still estimation (down from 19)",
             len(results), still_est)

if __name__ == "__main__":
    asyncio.run(run())
