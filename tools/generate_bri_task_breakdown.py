from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


SOURCE_XLS = Path(r"D:\BaiduSyncdisk\aicoding\news-monitor\topics\“一带一路”沿线国家中文教育发展指标体系数据总表.xls")
OUTPUT_XLSX = Path(r"D:\BaiduSyncdisk\aicoding\news-monitor\topics\一带一路_数据采集任务清单_拆解版.xlsx")
OUTPUT_MD = Path(r"D:\BaiduSyncdisk\aicoding\news-monitor\topics\一带一路_任务拆解说明.md")


@dataclass
class Metric:
    metric_id: int
    row_num: int
    name: str
    field_type: str
    desc: str


@dataclass
class Country:
    index: int
    name: str
    value_col_idx: int
    source_col_idx: int
    source_header: str
    region: str


def excel_col_name(zero_based_index: int) -> str:
    n = zero_based_index + 1
    out = []
    while n:
        n, rem = divmod(n - 1, 26)
        out.append(chr(65 + rem))
    return "".join(reversed(out))


def get_region(country: str) -> str:
    mapping = {
        "蒙古": "东北亚",
        "越南": "东南亚",
        "柬埔寨": "东南亚",
        "菲律宾": "东南亚",
        "新加坡": "东南亚",
        "文莱": "东南亚",
        "马来西亚": "东南亚",
        "印度尼西亚": "东南亚",
        "东帝汶": "东南亚",
        "老挝": "东南亚",
        "泰国": "东南亚",
        "缅甸": "东南亚",
        "不丹": "南亚",
        "孟加拉国": "南亚",
        "尼泊尔": "南亚",
        "斯里兰卡": "南亚",
        "印度": "南亚",
        "巴基斯坦": "南亚",
        "阿富汗": "南亚",
        "马尔代夫": "南亚",
        "土库曼斯坦": "中亚",
        "乌兹别克斯坦": "中亚",
        "哈萨克斯坦": "中亚",
        "塔吉克斯坦": "中亚",
        "吉尔吉斯斯坦": "中亚",
        "阿曼": "西亚中东",
        "伊朗": "西亚中东",
        "阿联酋": "西亚中东",
        "卡塔尔": "西亚中东",
        "巴林": "西亚中东",
        "沙特阿拉伯": "西亚中东",
        "科威特": "西亚中东",
        "也门": "西亚中东",
        "土耳其": "西亚中东",
        "伊拉克": "西亚中东",
        "叙利亚": "西亚中东",
        "约旦": "西亚中东",
        "黎巴嫩": "西亚中东",
        "以色列": "西亚中东",
        "格鲁吉亚": "高加索",
        "亚美尼亚": "高加索",
        "阿塞拜疆": "高加索",
        "埃及": "北非",
        "爱沙尼亚": "中东欧/欧洲",
        "拉脱维亚": "中东欧/欧洲",
        "立陶宛": "中东欧/欧洲",
        "德国": "欧洲",
        "波兰": "中东欧/欧洲",
        "捷克": "中东欧/欧洲",
        "斯洛文尼亚": "中东欧/欧洲",
        "克罗地亚": "中东欧/欧洲",
        "塞尔维亚": "中东欧/欧洲",
        "波黑": "中东欧/欧洲",
        "黑山": "中东欧/欧洲",
        "阿尔巴尼亚": "中东欧/欧洲",
        "匈牙利": "中东欧/欧洲",
        "斯洛伐克": "中东欧/欧洲",
        "北马其顿": "中东欧/欧洲",
        "罗马尼亚": "中东欧/欧洲",
        "保加利亚": "中东欧/欧洲",
        "摩尔多瓦": "中东欧/欧洲",
        "乌克兰": "中东欧/欧洲",
        "白俄罗斯": "中东欧/欧洲",
        "俄罗斯": "欧洲/欧亚",
    }
    return mapping.get(country, "待定")


def classify_metric(metric: Metric) -> tuple[str, str, str, str, str]:
    mid = metric.metric_id
    name = metric.name
    if 1 <= mid <= 11:
        category = "A.政策环境与制度基础"
    elif 12 <= mid <= 21:
        category = "B.学术研究与标准建设"
    elif 22 <= mid <= 29:
        category = "C.教材、资源与数字化"
    elif 30 <= mid <= 35:
        category = "D.学习者规模与考试"
    elif 36 <= mid <= 52:
        category = "E.师资、机构与教学供给"
    elif 53 <= mid <= 65:
        category = "F.产业合作、交流与外部支撑"
    else:
        category = "G.媒体传播与舆情"

    if metric.field_type in {"有/无", "官方语言/第二语言/关键语言/其他语言", "资本主义/社会主义/政教合一"}:
        method = "定性判定"
        verification = "需找到明确表述的官方/权威来源，并给出原文关键词。"
    elif metric.field_type in {"数值", "比例", "比值", "拟合值"}:
        method = "定量采集"
        verification = "需记录统计口径、时间点、单位/分母，并保留原始来源链接。"
    else:
        method = "混合采集"
        verification = "需明确最终值的分类口径，并附来源。"

    if mid in {1, 2, 3, 10, 11, 17, 30, 35, 42, 45, 48, 53, 55, 69}:
        priority = "P1"
    elif mid in {4, 5, 7, 8, 9, 23, 24, 26, 27, 31, 32, 37, 39, 40, 43, 44, 46, 47, 51, 56, 57, 58, 60, 61, 66, 67, 68}:
        priority = "P2"
    else:
        priority = "P3"

    if category == "A.政策环境与制度基础":
        source1, source2, source3 = "官方政府/教育部", "驻外使馆/双边声明", "权威研究报告"
    elif category == "B.学术研究与标准建设":
        source1, source2, source3 = "CNKI / WoS / Scopus", "高校/研究机构官网", "学术会议官网"
    elif category == "C.教材、资源与数字化":
        source1, source2, source3 = "教材出版社/教育部", "学校/平台官网", "应用商店/平台产品页"
    elif category == "D.学习者规模与考试":
        source1, source2, source3 = "教育部/统计局", "考试机构", "学术论文/年度报告"
    elif category == "E.师资、机构与教学供给":
        source1, source2, source3 = "教育部/学校清单", "教师协会/机构官网", "政府答复/招聘简章"
    elif category == "F.产业合作、交流与外部支撑":
        source1, source2, source3 = "商务部/投资统计", "校际合作协议/新闻", "企业/基金会官网"
    else:
        source1, source2, source3 = "主流媒体", "社交平台公开账号", "舆情检索工具/报告"

    return category, method, priority, verification, " | ".join([source1, source2, source3])


def search_hint(country: str, metric: Metric) -> tuple[str, str]:
    cn = f"{country} {metric.name}"
    en = f"{country} Chinese education {metric.name}"
    return cn, en


def load_template() -> tuple[list[Metric], list[Country]]:
    book = xlrd.open_workbook(str(SOURCE_XLS))
    sh = book.sheet_by_index(0)

    metrics: list[Metric] = []
    for r in range(1, sh.nrows):
        metric_id_raw = sh.cell_value(r, 0)
        name = str(sh.cell_value(r, 1)).strip()
        field_type = str(sh.cell_value(r, 2)).strip()
        desc = str(sh.cell_value(r, 3)).strip()
        if not metric_id_raw:
            continue
        metric_id = int(float(metric_id_raw))
        if not name:
            name = f"未命名字段_{metric_id}"
        metrics.append(Metric(metric_id=metric_id, row_num=r + 1, name=name, field_type=field_type, desc=desc))

    headers = sh.row_values(0)
    countries: list[Country] = []
    idx = 1
    for col in range(4, len(headers) - 1, 2):
        country = str(headers[col]).strip()
        source_header = str(headers[col + 1]).strip()
        if not country:
            continue
        countries.append(
            Country(
                index=idx,
                name=country,
                value_col_idx=col,
                source_col_idx=col + 1,
                source_header=source_header,
                region=get_region(country),
            )
        )
        idx += 1
    return metrics, countries


def style_header(ws, row_idx: int = 1) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[row_idx]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_workbook(metrics: list[Metric], countries: list[Country]) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "字段字典"
    ws1.append([
        "指标ID", "字段名", "字段类型", "分类", "采集方式", "优先级", "字段说明",
        "推荐来源优先级", "核验规则", "中文检索提示", "英文检索提示"
    ])
    style_header(ws1)
    for metric in metrics:
        category, method, priority, verification, source_pri = classify_metric(metric)
        cn, en = search_hint("国家名", metric)
        ws1.append([
            metric.metric_id, metric.name, metric.field_type, category, method, priority, metric.desc,
            source_pri, verification, cn, en
        ])
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws1, {"A": 8, "B": 28, "C": 22, "D": 22, "E": 14, "F": 8, "G": 54, "H": 32, "I": 30, "J": 28, "K": 28})

    ws2 = wb.create_sheet("国家列表")
    ws2.append(["序号", "国家", "区域", "原表值列", "原表出处列", "原表值单元格示例", "原表出处单元格示例"])
    style_header(ws2)
    for country in countries:
        ws2.append([
            country.index,
            country.name,
            country.region,
            excel_col_name(country.value_col_idx),
            excel_col_name(country.source_col_idx),
            f"{excel_col_name(country.value_col_idx)}2",
            f"{excel_col_name(country.source_col_idx)}2",
        ])
    autosize(ws2, {"A": 8, "B": 14, "C": 16, "D": 12, "E": 12, "F": 16, "G": 18})

    ws3 = wb.create_sheet("任务清单")
    ws3.append([
        "任务ID", "国家", "区域", "指标ID", "字段名", "分类", "字段类型", "优先级", "采集方式",
        "状态", "原表值单元格", "原表出处单元格", "推荐来源优先级", "中文检索提示", "英文检索提示", "核验规则"
    ])
    style_header(ws3)
    task_counter = 1
    for country in countries:
        for metric in metrics:
            category, method, priority, verification, source_pri = classify_metric(metric)
            cn, en = search_hint(country.name, metric)
            ws3.append([
                f"T{task_counter:04d}",
                country.name,
                country.region,
                metric.metric_id,
                metric.name,
                category,
                metric.field_type,
                priority,
                method,
                "待开始",
                f"{excel_col_name(country.value_col_idx)}{metric.row_num}",
                f"{excel_col_name(country.source_col_idx)}{metric.row_num}",
                source_pri,
                cn,
                en,
                verification,
            ])
            task_counter += 1
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws3, {"A": 10, "B": 12, "C": 16, "D": 8, "E": 24, "F": 20, "G": 16, "H": 8, "I": 12, "J": 10, "K": 12, "L": 12, "M": 32, "N": 28, "O": 28, "P": 28})

    ws4 = wb.create_sheet("示范国家_新加坡")
    ws4.append(["指标ID", "字段名", "当前值", "状态", "来源链接", "备注"])
    style_header(ws4)
    samples = [
        (1, "外交伙伴关系等级", "全方位高质量的前瞻性伙伴关系", "已初填", "https://www.mfa.gov.sg/newsroom/press-statements-transcripts-and-photos/exchange-of-congratulatory-messages-between-sg-and-china", "来源为新加坡外交部，2023年升级表述在2025年贺电中再次确认。"),
        (2, "中文纳入国民教育体系的层级", "有；基础教育（母语课程必修）", "已初填", "https://www.moe.gov.sg/primary/curriculum/mother-tongue-languages/learning-in-school", "MOE 明确说明 Chinese 为官方母语课程之一，且在小学阶段为 compulsory subject。"),
        (3, "中文教育的地位（中文在所在国的语言地位）", "官方语言之一；官方母语语言之一", "已初填", "https://www.tech.gov.sg/technews/how-singpass-learnt-three-languages", "GovTech 官方表述提到新加坡四种官方语言之一为 Chinese。"),
        (17, "是否拥有专门的学术期刊、研究机构或智库", "有", "已初填", "https://www.sccl.sg/en/about-sccl", "已确认 Singapore Centre for Chinese Language (SCCL)。"),
        (48, "中文教育/教师协会单位数", ">=1（已确认1个）", "已初填", "https://www.ntuc.org.sg/uportal/about-us/affiliated-unions/singapore-chinese-teachers-union", "至少确认 Singapore Chinese Teachers' Union，后续需继续补全其他协会/学会。"),
        (39, "本土中文教师培养项目数", "", "已启动待量化", "https://www.moe.gov.sg/careers/become-teachers/pri-sec-jc-ci/chinese-language-teaching", "已确认存在 MOE 中文教师培养/招募通道，但需补具体项目数量与年度口径。"),
        (42, "开设中文专业高校数", "", "已启动待量化", "", "建议从 NUS / NTU / SUSS / NIE 等高校官网建立名单。"),
        (69, "舆情监测中正面评价占比", "", "待采集", "", "需后续定义监测窗口、平台范围和情感判定口径。"),
    ]
    for row in samples:
        ws4.append(list(row))
    autosize(ws4, {"A": 8, "B": 28, "C": 24, "D": 14, "E": 70, "F": 44})
    for row in ws4.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUTPUT_XLSX)


def build_markdown(metrics: list[Metric], countries: list[Country]) -> None:
    lines = []
    lines.append("# 一带一路中文教育数据采集任务拆解说明")
    lines.append("")
    lines.append(f"- 原始模板：`{SOURCE_XLS}`")
    lines.append(f"- 指标数：`{len(metrics)}`")
    lines.append(f"- 国家数：`{len(countries)}`")
    lines.append(f"- 理论任务总量：`{len(metrics) * len(countries)}`")
    lines.append("")
    lines.append("## 本次已完成")
    lines.append("")
    lines.append("1. 解析原始 `.xls` 模板，提取字段字典与国家列表。")
    lines.append("2. 生成可执行的 `4416` 条任务清单，并保留回填到原总表的单元格坐标。")
    lines.append("3. 选取 `新加坡` 作为示范国家，先行启动基础字段填报。")
    lines.append("")
    lines.append("## 推荐执行顺序")
    lines.append("")
    lines.append("### 第1阶段：先做 P1 基础字段")
    lines.append("- 外交伙伴关系等级")
    lines.append("- 中文是否纳入国民教育体系、语言地位")
    lines.append("- 风险类字段（教育政策环境风险、意识形态风险）")
    lines.append("- 是否存在研究机构/协会/核心机构")
    lines.append("- 学习者比例、考试人数、重点机构数量")
    lines.append("")
    lines.append("### 第2阶段：再做 P2 可量化字段")
    lines.append("- 项目数、课程数、教材数、学校数、教师培训数、合作协议数")
    lines.append("- 数字资源与平台、文化活动、媒体发布数等")
    lines.append("")
    lines.append("### 第3阶段：最后做 P3 深水区字段")
    lines.append("- 论文质量拟合值")
    lines.append("- 高端人才学习者数量")
    lines.append("- 偏远地区覆盖率")
    lines.append("- 正面评价占比等需要自定义口径的指标")
    lines.append("")
    lines.append("## 注意事项")
    lines.append("")
    lines.append("- 原表第 `36` 号指标名称为空，需要业务侧确认是否为漏项。")
    lines.append("- 部分指标必须先统一年份口径，否则后续跨国不可比。")
    lines.append("- `比例/比值` 类字段务必同步记录分子、分母和计算方式。")
    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    metrics, countries = load_template()
    build_workbook(metrics, countries)
    build_markdown(metrics, countries)
    print(OUTPUT_XLSX)
    print(OUTPUT_MD)


if __name__ == "__main__":
    main()
