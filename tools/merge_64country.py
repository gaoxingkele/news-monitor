"""合并 13 份 refined447 xlsx → 单一 64 国总表。

输出：topics/bri_64country_merged_{DATE}.xlsx
列结构：
  A: 序号（指标ID）
  B: 搜集字段
  C: 字段类型
  D: 字段说明
  E: 分类
  F: 优先级
  G..: 64 国 × (值 + 出处) = 128 列
"""
from __future__ import annotations
import io, json, sys
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
TOPICS = ROOT / "topics"
FIELD_DICT = TOPICS / "bri_field_dict.json"
DATE = datetime.now().strftime("%Y%m%d")
OUT = TOPICS / f"bri_64country_merged_{DATE}.xlsx"

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 从 scan 脚本复用 batch 映射
BATCHES = [
    (1,  "“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v11_20260422__refined447_20260424.xlsx",
        ["蒙古","越南","柬埔寨","菲律宾","新加坡"]),
    (2,  "“一带一路”沿线国家中文教育发展指标体系数据总表_批次2_五国_v4_20260422__refined447_20260424.xlsx",
        ["马来西亚","泰国","印尼","老挝","缅甸"]),
    (3,  "“一带一路”沿线国家中文教育发展指标体系数据总表_批次3_南亚I_五国_v_final_20260423__refined447_20260424.xlsx",
        ["文莱","东帝汶","不丹","孟加拉国","尼泊尔"]),
    (4,  "“一带一路”沿线国家中文教育发展指标体系数据总表_批次4_南亚II_五国_v_final_20260423__refined447_20260424.xlsx",
        ["斯里兰卡","印度","巴基斯坦","阿富汗","马尔代夫"]),
    (5,  "“一带一路”沿线国家中文教育发展指标体系数据总表_批次5_中亚_五国_v_final_20260423__refined447_20260424.xlsx",
        ["土库曼斯坦","乌兹别克斯坦","哈萨克斯坦","塔吉克斯坦","吉尔吉斯斯坦"]),
    (6,  "“一带一路”沿线国家中文教育发展指标体系数据总表_批次6_西亚海湾_五国_v_final_20260423__refined447_20260424.xlsx",
        ["阿曼","伊朗","阿联酋","卡塔尔","巴林"]),
    (7,  "“一带一路”沿线国家中文教育发展指标体系数据总表_批次7_西亚阿拉伯_五国_v_final_20260423__refined447_20260424.xlsx",
        ["沙特阿拉伯","科威特","也门","土耳其","伊拉克"]),
    (8,  "“一带一路”沿线国家中文教育发展指标体系数据总表_批次8_地中海东岸_五国_v_final_20260423__refined447_20260424.xlsx",
        ["叙利亚","约旦","黎巴嫩","以色列","格鲁吉亚"]),
    (9,  "“一带一路”沿线国家中文教育发展指标体系数据总表_批次9_高加索北非_五国_v_final_20260423__refined447_20260424.xlsx",
        ["亚美尼亚","阿塞拜疆","埃及","爱沙尼亚","拉脱维亚"]),
    (10, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次10_欧洲中部I_五国_v_final_20260423__refined447_20260424.xlsx",
        ["立陶宛","德国","波兰","捷克","斯洛文尼亚"]),
    (11, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次11_巴尔干I_五国_v_final_20260423__refined447_20260424.xlsx",
        ["克罗地亚","塞尔维亚","波黑","黑山","阿尔巴尼亚"]),
    (12, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次12_欧洲中部II_五国_v_final_20260423__refined447_20260424.xlsx",
        ["匈牙利","斯洛伐克","北马其顿","罗马尼亚","保加利亚"]),
    (13, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次13_东欧_五国_v_final_20260423__refined447_20260424.xlsx",
        ["摩尔多瓦","乌克兰","白俄罗斯","俄罗斯"]),
]


def main():
    # 汇总所有 64 国数据
    all_data = {}  # (country, iid) -> (value, source)
    country_order = []
    for _, fname, countries in BATCHES:
        fp = TOPICS / fname
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        for i, cname in enumerate(countries):
            country_order.append(cname)
            vcol = 5 + 2 * i
            for r in range(2, ws.max_row + 1):
                iid = ws.cell(r, 1).value
                if iid is None: continue
                try: iid = int(iid)
                except: continue
                val = ws.cell(r, vcol).value
                src = ws.cell(r, vcol + 1).value
                all_data[(cname, iid)] = (val, src)
    print(f"读完 {len(country_order)} 国 × 68 指标 = {len(all_data)} 格")

    # 加载字段字典
    with open(FIELD_DICT, encoding='utf-8') as f:
        fields = [x for x in json.load(f) if x.get('指标ID')]
    fields.sort(key=lambda x: int(x['指标ID']))

    # 建新 workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BRI 64国总表"

    # 表头
    headers = ['序号', '搜集字段', '字段类型', '字段说明', '分类', '优先级']
    for c in country_order:
        headers.extend([c, f"{c}_出处"])

    header_fill = PatternFill(start_color="2E5C8A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    hal = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CCCCCC")
    bd = Border(top=thin, bottom=thin, left=thin, right=thin)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = hal
        cell.border = bd

    # 行数据
    for i, fld in enumerate(fields):
        r = i + 2
        iid = int(fld['指标ID'])
        ws.cell(r, 1, iid)
        ws.cell(r, 2, fld.get('字段名', ''))
        ws.cell(r, 3, fld.get('字段类型', ''))
        ws.cell(r, 4, fld.get('字段说明', ''))
        ws.cell(r, 5, fld.get('分类', ''))
        ws.cell(r, 6, fld.get('优先级', ''))
        for j, cname in enumerate(country_order):
            vcol = 7 + 2 * j
            v, s = all_data.get((cname, iid), (None, None))
            if v is not None:
                ws.cell(r, vcol, v)
            if s is not None:
                ws.cell(r, vcol + 1, s)

        # 整行样式
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = bd

    # 列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 36
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 8
    for j in range(len(country_order)):
        ws.column_dimensions[get_column_letter(7 + 2 * j)].width = 18
        ws.column_dimensions[get_column_letter(8 + 2 * j)].width = 22

    ws.freeze_panes = "G2"
    ws.row_dimensions[1].height = 28

    wb.save(OUT)
    print(f"写入 {OUT}")
    print(f"尺寸：{len(fields)+1} 行 × {len(headers)} 列")


if __name__ == "__main__":
    main()
