"""为每个 BRI 国家生成独立完整指标 MD 报告。

输出：topics/per_country/{序号}_{国家}_完整指标数据_{DATE}.md

每份报告包含：
  - 国家概览（总指标数 / 已填 / 数字值率 / 可疑格统计）
  - 68 指标逐行：ID / 字段名 / 分类 / 优先级 / 字段类型 / 字段说明 / 当前值 / 计算依据 / 源 URL / 置信度
  - 精修 447 批次合入 ckpt 的 derivation/rationale/confidence
"""
from __future__ import annotations
import io, json, re, sys
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
TOPICS = ROOT / "topics"
OUT_DIR = TOPICS / "per_country"
OUT_DIR.mkdir(exist_ok=True)
FIELD_DICT = TOPICS / "bri_field_dict.json"
MERGED = TOPICS / f"bri_64country_merged_{datetime.now().strftime('%Y%m%d')}.xlsx"
CKPT = ROOT / "output" / "bri_refine447_ckpt.json"
DATE = datetime.now().strftime("%Y%m%d")

import openpyxl


def parse_cell_value(raw: str) -> dict:
    """从 xlsx 单元格文本提取结构化内容。

    识别前缀：【代理推导】/【精修-广义】/【精修447】/【估算】
    识别"推导："/ "代理："分节。
    """
    if not raw:
        return {'value': '', 'prefix': '', 'proxy': '', 'derivation': ''}
    s = str(raw).strip()
    prefix = ''
    m = re.match(r'^(【[^】]+】)+', s)
    if m:
        prefix = m.group(0)
        s = s[len(prefix):].strip()

    # 取代理：X
    proxy = ''
    mp = re.search(r'[（(]代理[：:]\s*(.+?)[）)]', s)
    if mp:
        proxy = mp.group(1).strip()
        s = s[:mp.start()] + s[mp.end():]

    # 取"推导：" 之后
    derivation = ''
    md = re.search(r'推导[：:]\s*(.+)', s, re.DOTALL)
    if md:
        derivation = md.group(1).strip()
        s = s[:md.start()].strip()

    return {
        'value': s.strip(),
        'prefix': prefix,
        'proxy': proxy,
        'derivation': derivation,
    }


def classify_status(val: str) -> str:
    """判定单元格当前状态：numeric / qualitative / weak / empty"""
    if not val: return '空'
    clean = re.sub(r'^(【[^】]+】\s*)+', '', val).strip()
    first = re.split(r'[（\(\n]', clean, maxsplit=1)[0].strip()
    first = re.split(r'[，,；;：:]', first, maxsplit=1)[0].strip()
    if first in ('0','无','没有','0%','少量','若干','有限','无可核验','无可靠数据'):
        return '弱'
    if any(c.isdigit() for c in first):
        return '数字'
    if len(first) >= 2:
        return '定性'
    return '弱'


def main():
    # 读字段字典
    with open(FIELD_DICT, encoding='utf-8') as f:
        fields = sorted([x for x in json.load(f) if x.get('指标ID')],
                        key=lambda x: int(x['指标ID']))

    # 读 ckpt（精修 447 格）
    ckpt = {}
    if CKPT.exists():
        ckpt = json.loads(CKPT.read_text(encoding='utf-8'))

    # 读总表
    wb = openpyxl.load_workbook(MERGED, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # 国家 → (value_col, source_col)
    country_cols = {}
    for c_idx, h in enumerate(headers, 1):
        if h and not str(h).endswith('_出处') and h not in ('序号','搜集字段','字段类型','字段说明','分类','优先级'):
            country_cols[h] = (c_idx, c_idx + 1)

    # 指标 row map
    indicator_rows = {}
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, 1).value
        if iid is None: continue
        try: iid = int(iid)
        except: continue
        indicator_rows[iid] = r

    print(f"总表：{len(country_cols)} 国 × {len(indicator_rows)} 指标")

    # 对每个国家生成报告
    for idx, (country, (vcol, scol)) in enumerate(country_cols.items(), 1):
        safe_name = country.replace('/', '_')
        out_fp = OUT_DIR / f"{idx:02d}_{safe_name}_完整指标数据_{DATE}.md"

        # 收集数据
        rows = []
        stat = {'数字': 0, '定性': 0, '弱': 0, '空': 0}
        refined_count = 0
        for iid, fld in [(int(f['指标ID']), f) for f in fields]:
            r = indicator_rows.get(iid)
            if not r: continue
            raw = ws.cell(r, vcol).value
            src = ws.cell(r, scol).value
            parsed = parse_cell_value(str(raw) if raw else '')
            status = classify_status(str(raw) if raw else '')
            stat[status] = stat.get(status, 0) + 1

            # 若在 ckpt，取精修结构化内容
            key = f"{country}::{iid}"
            refined = ckpt.get(key)
            if refined:
                refined_count += 1

            rows.append({
                'iid': iid, 'name': fld.get('字段名',''),
                'category': fld.get('分类',''),
                'priority': fld.get('优先级',''),
                'type': fld.get('字段类型',''),
                'desc': fld.get('字段说明','') or '',
                'raw': str(raw) if raw is not None else '',
                'parsed': parsed,
                'src': str(src) if src is not None else '',
                'refined': refined,
                'status': status,
            })

        # 输出 MD
        lines = []
        lines.append(f"# {country} · 中文教育指标体系完整数据")
        lines.append("")
        lines.append(f"**国家**：{country}")
        lines.append(f"**生成时间**：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
        lines.append(f"**指标总数**：{len(rows)}（总表 68 项，含 #36 空字段已跳过）")
        lines.append("")
        lines.append(f"## 数据质量概览")
        lines.append("")
        lines.append(f"| 类型 | 数量 | 占比 |")
        lines.append(f"|---|---|---|")
        tot = len(rows)
        for k in ['数字', '定性', '弱', '空']:
            c = stat.get(k, 0)
            lines.append(f"| {k} | {c} | {c*100//tot if tot else 0}% |")
        lines.append(f"| **含精修 447 轮** | **{refined_count}** | **{refined_count*100//tot if tot else 0}%** |")
        lines.append("")

        # 按分类分组
        lines.append(f"## 按分类分布")
        lines.append("")
        from collections import Counter
        cat_stat = Counter(r['category'] for r in rows)
        for cat, n in sorted(cat_stat.items(), key=lambda x: -x[1]):
            lines.append(f"- **{cat}**：{n} 项")
        lines.append("")
        lines.append("---")
        lines.append("")

        # 详细指标
        lines.append(f"## 指标详情（按 ID 排序）")
        lines.append("")

        cur_cat = None
        for r in rows:
            if r['category'] != cur_cat:
                cur_cat = r['category']
                lines.append(f"### {cur_cat}")
                lines.append("")

            p = r['parsed']
            lines.append(f"#### {r['iid']}. {r['name']}")
            lines.append("")
            lines.append(f"- **优先级**：{r['priority']}")
            lines.append(f"- **字段类型**：{r['type']}")
            lines.append(f"- **字段说明**：{r['desc'][:300]}")
            lines.append(f"- **数据状态**：{r['status']}")
            lines.append("")
            if p['prefix']:
                lines.append(f"**数据来源标识**：{p['prefix']}")
                lines.append("")
            lines.append(f"**当前值**：")
            lines.append("")
            lines.append(f"> {p['value'] or '(空)'}")
            lines.append("")

            if p['proxy']:
                lines.append(f"**代理变量**：{p['proxy']}")
                lines.append("")

            if p['derivation']:
                lines.append(f"**推导过程（xlsx 单元格内）**：")
                lines.append("")
                lines.append(f"{p['derivation']}")
                lines.append("")

            # 精修 447 详情
            if r['refined']:
                ref = r['refined']
                lines.append(f"**精修 447 轮详情**：")
                lines.append("")
                lines.append(f"- 合成值：{ref.get('value','')}")
                if ref.get('proxy_used'):
                    lines.append(f"- 代理：{ref['proxy_used']}")
                if ref.get('derivation'):
                    lines.append(f"- 推导：{ref['derivation']}")
                if ref.get('rationale'):
                    lines.append(f"- 依据：{ref['rationale']}")
                lines.append(f"- 置信度：{ref.get('confidence','')}")
                lines.append(f"- 是否估算：{ref.get('is_estimation','')}")
                urls = ref.get('source_urls', [])
                if urls:
                    if isinstance(urls, str): urls = [urls]
                    lines.append(f"- 源 URL（精修）：")
                    for u in urls[:5]:
                        lines.append(f"  - {u}")
                lines.append("")

            # 出处列
            if r['src']:
                lines.append(f"**xlsx 出处列**：")
                lines.append("")
                urls_in_src = re.findall(r'https?://[^\s;,)]+', r['src'])
                if urls_in_src:
                    for u in urls_in_src[:5]:
                        lines.append(f"- {u}")
                else:
                    lines.append(f"{r['src'][:500]}")
                lines.append("")

            lines.append("---")
            lines.append("")

        out_fp.write_text('\n'.join(lines), encoding='utf-8')
        print(f"[{idx:02d}] {country} → {out_fp.name} ({len(rows)} 项, {refined_count} 精修)")


if __name__ == '__main__':
    main()
