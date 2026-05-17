"""分组二次补强 9 个剩余估算格.

A 组 (站点强过滤): 蒙古#60、柬埔寨#59、柬埔寨#67
    → Perplexity + Tavily 限定 chinese.cn/ci.cn/hanban.org/bridge.chinese.cn/clef.org.cn
B 组 (学术专题): 菲律宾#12、新加坡#12
    → Tavily 限定 CNKI/ISEAS/hanspub/scholar + 专题关键词 "专著 monograph"
C 组 (多分母口径): 越南#30、柬埔寨#30、菲律宾#32、新加坡#43
    → 对每种可能分母分别算一个比例，业务侧最后挑

产出:
  ..._五国填充版_v4_final_{date}.xlsx  (覆盖这 9 格)
  bri_5country_multi_denominator_{date}.md  (C 组多口径计算表)
  bri_5country_deep_v2_log_{date}.md
"""
from __future__ import annotations
import asyncio, io, json, os, sys, time, logging
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError: pass

import httpx, openpyxl
from openpyxl.styles import PatternFill, Alignment
from news_monitor.proxy import build_httpx_client, get_proxies_for_url

TOPICS = ROOT / "topics"
DATE = datetime.now().strftime("%Y%m%d")
SRC = TOPICS / "“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v3_final_20260420.xlsx"
V4 = TOPICS / f"“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v4_final_{DATE}.xlsx"
MULTI_MD = TOPICS / f"bri_5country_multi_denominator_{DATE}.md"
LOG_MD = TOPICS / f"bri_5country_deep_v2_log_{DATE}.md"
CKPT = ROOT / "output" / "bri_deep_v2_checkpoint.json"
FIELD_DICT = TOPICS / "bri_field_dict.json"

PPX_KEY = os.environ.get("PERPLEXITY_API_KEY", "")
TAV_KEY = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
BRV_KEY = os.environ.get("BRAVEAPI", "")
DS_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
PROXY = os.environ.get("LLM_PROXY", "") or os.environ.get("OVERSEAS_PROXY", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger("deep_v2")

COL = {"蒙古":(5,6), "越南":(7,8), "柬埔寨":(9,10), "菲律宾":(11,12), "新加坡":(13,14)}
EN = {"蒙古":"Mongolia", "越南":"Vietnam", "柬埔寨":"Cambodia", "菲律宾":"Philippines", "新加坡":"Singapore"}

# 三组任务
GROUP_A = [("蒙古", 60), ("柬埔寨", 59), ("柬埔寨", 67)]   # 站点强过滤
GROUP_B = [("菲律宾", 12), ("新加坡", 12)]                   # 学术专题
GROUP_C = [("越南", 30), ("柬埔寨", 30), ("菲律宾", 32), ("新加坡", 43)]  # 多分母

HANBAN_SITES = ["chinese.cn", "ci.cn", "hanban.org", "bridge.chinese.cn",
                 "clef.org.cn", "cltc.pku.edu.cn", "confucius.edu.cn"]
ACADEMIC_SITES = ["cnki.com.cn", "hanspub.org", "scholar.google.com",
                   "iseas.edu.sg", "issi.ac.cn", "nspk.xjtu.edu.cn",
                   "cass.cn", "aciar.gov.au", "cambridge.org"]

# C 组：分母候选清单（业务侧可再调整）
DENOMINATORS = {
    ("越南", 30): [
        ("2023 越南总人口 98,144,000 (世行)", "https://data.worldbank.org/country/vietnam"),
        ("2022 越南基础教育在校生 ≈ 17,900,000 (UNESCO/越教育部)", "https://www.moet.gov.vn/"),
        ("2022 越南适龄学生 6-22 岁人口 ≈ 23,000,000", "https://data.worldbank.org/indicator/SP.POP.0014.TO?locations=VN"),
    ],
    ("柬埔寨", 30): [
        ("2023 柬埔寨总人口 16,780,000 (世行)", "https://data.worldbank.org/country/cambodia"),
        ("2022 柬埔寨义务教育在校生 ≈ 3,300,000 (MoEYS)", "https://moeys.gov.kh/"),
        ("2022 柬埔寨 6-18 岁人口 ≈ 5,100,000", "https://data.worldbank.org/country/cambodia"),
    ],
    ("菲律宾", 32): [
        ("全菲华校学生总数 ≈ 120,000 (华文学校联合总会估算)", "https://www.chinesenewsph.com/"),
        ("DepEd 基础教育含中文选修的学校总中文学习者估算 ≈ 200,000", "https://www.deped.gov.ph/"),
        ("孔院+孔子课堂 + 大学中文系 = 补充分子", "https://ci.cn/"),
    ],
    ("新加坡", 43): [
        ("新加坡 6 所公立大学 (MOE 定义 Autonomous Universities)", "https://www.moe.gov.sg/post-secondary/overview/autonomous-universities"),
        ("新加坡 6 所自治大学 + 5 所理工学院 = 11 所公办高教机构", "https://www.moe.gov.sg/"),
        ("新加坡 34 所含私立大学总数 (CPE 注册)", "https://www.cpe.gov.sg/"),
    ],
}

# ── API ──────────────────────────────────────────────────────────────────────
async def ppx(client, system, user):
    if not PPX_KEY: return "", []
    body = {"model":"sonar","messages":[{"role":"system","content":system},
            {"role":"user","content":user}],"temperature":0.05,"return_citations":True}
    try:
        resp = await client.post("https://api.perplexity.ai/chat/completions", json=body,
                                  headers={"Authorization":f"Bearer {PPX_KEY}","Content-Type":"application/json"},
                                  timeout=120.0)
        resp.raise_for_status()
        d = resp.json()
        return d["choices"][0]["message"]["content"], d.get("citations",[])[:12]
    except Exception as e:
        log.warning("ppx: %s", str(e)[:120]); return "", []

async def tav(client, query, domains):
    if not TAV_KEY: return []
    try:
        resp = await client.post("https://api.tavily.com/search",
                                  json={"api_key":TAV_KEY,"query":query,"search_depth":"advanced",
                                        "max_results":10,"include_domains":domains[:20],
                                        "include_answer":False,"include_raw_content":False},
                                  timeout=60.0)
        resp.raise_for_status()
        return [{"url":r.get("url",""), "title":r.get("title",""),
                 "snippet":(r.get("content") or "")[:500]}
                for r in (resp.json().get("results") or [])[:10]]
    except Exception as e:
        log.warning("tav: %s", str(e)[:120]); return []

async def brave(client, query):
    if not BRV_KEY: return []
    try:
        resp = await client.get("https://api.search.brave.com/res/v1/web/search",
                                 params={"q":query,"count":15,"text_decorations":False},
                                 headers={"Accept":"application/json","X-Subscription-Token":BRV_KEY},
                                 timeout=30.0)
        resp.raise_for_status()
        return [{"url":r.get("url",""), "title":r.get("title",""),
                 "snippet":(r.get("description") or "")[:400]}
                for r in (resp.json().get("web",{}).get("results") or [])[:10]]
    except Exception as e:
        log.warning("brave: %s", str(e)[:120]); return []

async def ds(client, prompt):
    if not DS_KEY: return {}
    try:
        resp = await client.post("https://api.deepseek.com/v1/chat/completions",
                                  json={"model":"deepseek-chat","messages":[{"role":"user","content":prompt}],
                                        "temperature":0.05,"response_format":{"type":"json_object"}},
                                  headers={"Authorization":f"Bearer {DS_KEY}","Content-Type":"application/json"},
                                  timeout=120.0)
        resp.raise_for_status()
        return json.loads(resp.json()["choices"][0]["message"]["content"])
    except Exception as e:
        log.warning("ds: %s", str(e)[:120]); return {}

# ── Group A: 站点强过滤 ────────────────────────────────────────────────────
async def solve_group_a(c, iid, ind, ppx_c, tav_c, brv_c, ds_c):
    en = EN[c]
    name = ind["字段名"]; ft = ind["字段类型"]; desc = ind["字段说明"]
    # Tavily 站点限定
    q1 = f"{c} {name} 汉语桥 国际中文日 孔子学院"
    q2 = f"{en} Confucius Institute statistics {name}"
    r1, r2 = await asyncio.gather(
        tav(tav_c, q1, HANBAN_SITES),
        tav(tav_c, q2, HANBAN_SITES),
        return_exceptions=True,
    )
    # Brave: 官方中文
    r3 = await brave(brv_c, f"site:chinese.cn OR site:ci.cn {c} {name}")
    # Perplexity 汇总
    system = "你是国际中文教育统计专家。必须基于官方汉办/CLEC/孔子学院系统数据回答。"
    user = (f"国家：{c}\n指标：{name}\n取值口径：{ft}\n说明：{desc}\n\n"
            f"请仅使用 chinese.cn, ci.cn, hanban.org, bridge.chinese.cn 等官方来源，"
            f"给出最精确的数值 + 每一条的官方 URL。若为累计值请列出每年拆分。"
            f"若真无数据，明确说'官方未发布'。")
    ans, cites = await ppx(ppx_c, system, user)

    evid = []
    for x in (r1, r2):
        if isinstance(x, list): evid.extend(x)
    evid.extend(r3[:6])
    evid_block = "\n".join(f"[{i+1}] {e['title']}\nURL: {e['url']}\n{e['snippet']}"
                          for i,e in enumerate(evid[:15]))
    cite_block = "\n".join(f"- {u}" for u in cites)

    prompt = f"""你是严谨的统计研究员。为【{c}】在指标【{name}】上给出最精确的值。
取值口径：{ft}
字段说明：{desc}

## Perplexity 官方专查
{ans or '(空)'}
citations: {cite_block}

## Tavily/Brave 站点过滤结果（chinese.cn / ci.cn / hanban.org）
{evid_block}

返回 JSON:
{{
  "value": "最终值（贴合口径）",
  "breakdown": "年度或项目拆分，形如 '2023: 1 + 2024: 2 + 2025: 2 = 5'",
  "source_url": "最权威的一条",
  "alt_urls": ["备选1","备选2"],
  "confidence": "high/medium/low",
  "is_still_estimation": true/false,
  "rationale": "30字内"
}}
若确实官方无数据，value 填 '官方未发布'，confidence=low。
"""
    return await ds(ds_c, prompt)

# ── Group B: 学术专著 ──────────────────────────────────────────────────────
async def solve_group_b(c, iid, ind, ppx_c, tav_c, ds_c):
    en = EN[c]
    name = ind["字段名"]; ft = ind["字段类型"]; desc = ind["字段说明"]
    q1 = f"{en} Chinese language education monograph book research"
    q2 = f"{c} 中文教育 专著 丛书 国别研究"
    q3 = f"{en} Chinese overseas education 华文教育 专著"
    r1, r2, r3 = await asyncio.gather(
        tav(tav_c, q1, ACADEMIC_SITES),
        tav(tav_c, q2, ACADEMIC_SITES),
        tav(tav_c, q3, ACADEMIC_SITES),
        return_exceptions=True,
    )
    system = "你是华文教育研究专家。列出所有与该国相关的中文教育专著/丛书/研究报告。"
    user = (f"请列出所有关于【{c}】中文教育/华文教育的学术专著（不含单篇论文），"
            f"包括：书名、作者、出版社、出版年份、可访问 URL。"
            f"范围：CNKI 图书、ISEAS 丛书、商务印书馆海外华文教育系列、北大/北语/华大出版的相关专著。"
            f"若数量很少（<5），列全部；若较多，列出代表性 10 本。")
    ans, cites = await ppx(ppx_c, system, user)

    evid = []
    for x in (r1, r2, r3):
        if isinstance(x, list): evid.extend(x)
    evid_block = "\n".join(f"[{i+1}] {e['title']}\nURL: {e['url']}\n{e['snippet']}"
                          for i,e in enumerate(evid[:20]))

    prompt = f"""你是文献计量研究员。请基于下列材料统计【{c}】中文教育领域的专著数量。
取值口径：{ft} (直接给数字)
字段说明：{desc}

## Perplexity 专题查询
{ans or '(空)'}
citations: {chr(10).join(cites)}

## 学术库搜索
{evid_block}

判定标准：
- 只计"专著/monograph/book"，不含单篇论文、会议论文、期刊
- 含"该国中文教育/华文教育"主题
- 有 ISBN 或可访问的图书条目

返回 JSON:
{{
  "value": "数字（若确为 0 可填 0）",
  "book_list": ["书名1 (作者, 年)", "书名2", ...],
  "source_url": "最权威一条",
  "confidence": "high/medium/low",
  "is_still_estimation": true/false,
  "rationale": "30字"
}}
"""
    return await ds(ds_c, prompt)

# ── Group C: 多分母口径 ────────────────────────────────────────────────────
async def solve_group_c(c, iid, ind, ppx_c, tav_c, brv_c, ds_c):
    """对每种候选分母算一个比例。"""
    en = EN[c]
    name = ind["字段名"]; ft = ind["字段类型"]; desc = ind["字段说明"]
    # 先找分子
    system = "你是统计研究员。只给我数字和来源，不要解释。"
    user = (f"请提供【{c}】【{name}】涉及的【分子】最新数据。\n"
            f"- 字段说明：{desc}\n"
            f"- 若有多个可能分子（如中小学学习者、大学生、孔院学员），分别列出\n"
            f"- 每个数字给：数值 + 年份 + 官方 URL\n"
            f"简洁。")
    ans, cites = await ppx(ppx_c, system, user)

    denoms = DENOMINATORS.get((c, iid), [])
    # 让 DS 做多套计算
    denom_block = "\n".join(f"{i+1}. {d[0]}  (URL: {d[1]})" for i,d in enumerate(denoms))

    prompt = f"""你是统计研究员。为【{c}】的指标【{name}】做【多口径比例计算】。
取值口径：{ft}
字段说明：{desc}

## Perplexity 查到的分子候选
{ans}

## 业务侧指定的分母候选
{denom_block}

任务：
1. 从 Perplexity 回答中抽取所有合理的分子（可能多个）
2. 对每一种"分子 × 分母"组合，给出一个 value（百分比保留 2 位小数）
3. 以表格形式返回 JSON

返回 JSON:
{{
  "numerators": [
    {{"value": "数值+单位", "year": "年份", "url": "来源"}}
  ],
  "computations": [
    {{"numerator_idx": 1, "denominator_idx": 1, "formula": "N / D = X%",
      "value": "X%", "caveat": "使用限制"}},
    ...
  ],
  "recommended": "推荐业务侧采用哪一组，为什么，30字",
  "confidence_best": "high/medium/low"
}}
"""
    return await ds(ds_c, prompt)

# ── main ─────────────────────────────────────────────────────────────────────
async def run():
    with open(FIELD_DICT, encoding='utf-8') as f:
        fields = {int(x["指标ID"]): x for x in json.load(f)}

    ckpt = {}
    if CKPT.exists():
        try: ckpt = json.loads(CKPT.read_text(encoding='utf-8'))
        except Exception: ckpt = {}

    ppx_p = get_proxies_for_url("https://api.perplexity.ai", PROXY)
    brv_p = get_proxies_for_url("https://api.search.brave.com", PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", PROXY)
    async with (build_httpx_client(ppx_p or "", 120) as ppx_c,
                build_httpx_client(brv_p or "", 30) as brv_c,
                build_httpx_client(tav_p or "", 60) as tav_c,
                build_httpx_client("", 120) as ds_c):

        log.info("Group A (3 格 站点强过滤)...")
        a_tasks = [solve_group_a(c, iid, fields[iid], ppx_c, tav_c, brv_c, ds_c)
                    for c, iid in GROUP_A]
        a_results = await asyncio.gather(*a_tasks)
        for (c, iid), r in zip(GROUP_A, a_results):
            ckpt[f"A::{c}::{iid}"] = r
            log.info("  %s#%d value=%s conf=%s", c, iid,
                     str(r.get("value",""))[:50], r.get("confidence","?"))

        log.info("Group B (2 格 学术专著)...")
        b_tasks = [solve_group_b(c, iid, fields[iid], ppx_c, tav_c, ds_c)
                    for c, iid in GROUP_B]
        b_results = await asyncio.gather(*b_tasks)
        for (c, iid), r in zip(GROUP_B, b_results):
            ckpt[f"B::{c}::{iid}"] = r
            log.info("  %s#%d value=%s conf=%s", c, iid,
                     str(r.get("value",""))[:50], r.get("confidence","?"))

        log.info("Group C (4 格 多分母口径)...")
        c_tasks = [solve_group_c(c, iid, fields[iid], ppx_c, tav_c, brv_c, ds_c)
                    for c, iid in GROUP_C]
        c_results = await asyncio.gather(*c_tasks)
        for (c, iid), r in zip(GROUP_C, c_results):
            ckpt[f"C::{c}::{iid}"] = r
            n_comps = len(r.get("computations", []))
            log.info("  %s#%d 计算组合=%d", c, iid, n_comps)

        with open(CKPT, "w", encoding="utf-8") as f:
            json.dump(ckpt, f, ensure_ascii=False, indent=1)

    # 写 v4 xlsx
    wb = openpyxl.load_workbook(SRC)
    ws = wb.active
    est_fill = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")
    refined_fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")

    log_md = [f"# BRI 5 国深度补强 v2 (分组策略)", f"生成：{datetime.now():%Y-%m-%d %H:%M}\n"]
    multi_md = [f"# C 组多分母口径计算表",
                f"生成：{datetime.now():%Y-%m-%d %H:%M}",
                f"每指标给出多种分子×分母组合，业务侧可择最合口径者采用。\n"]

    # A 组回填
    log_md.append("\n## A 组：站点强过滤结果")
    for (c, iid), r in zip(GROUP_A, a_results):
        name = fields[iid]["字段名"]
        val = str(r.get("value", "")).strip()
        vcol, scol = COL[c]
        row = next(rr for rr in range(2, ws.max_row+1) if ws.cell(rr,1).value == iid)
        breakdown = r.get("breakdown", "")
        if val and val != "官方未发布":
            is_est = str(r.get("is_still_estimation", True)).lower() == "true"
            prefix = "【估算】" if is_est else ""
            cell_txt = f"{prefix}{val}" + (f"（{breakdown}）" if breakdown else "")
            ws.cell(row, vcol, cell_txt).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, vcol).fill = est_fill if is_est else refined_fill
            if r.get("source_url"): ws.cell(row, scol, r["source_url"])
        log_md.append(f"\n### {c}#{iid} {name}")
        log_md.append(f"- 值：{val}")
        if breakdown: log_md.append(f"- 拆分：{breakdown}")
        log_md.append(f"- 置信度：{r.get('confidence','?')}")
        log_md.append(f"- 理由：{r.get('rationale','')}")
        log_md.append(f"- 来源：{r.get('source_url','')}")

    # B 组回填
    log_md.append("\n## B 组：学术专著")
    for (c, iid), r in zip(GROUP_B, b_results):
        name = fields[iid]["字段名"]
        val = str(r.get("value", "")).strip()
        vcol, scol = COL[c]
        row = next(rr for rr in range(2, ws.max_row+1) if ws.cell(rr,1).value == iid)
        if val:
            is_est = str(r.get("is_still_estimation", True)).lower() == "true"
            prefix = "【估算】" if is_est else ""
            ws.cell(row, vcol, f"{prefix}{val}").alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, vcol).fill = est_fill if is_est else refined_fill
            if r.get("source_url"): ws.cell(row, scol, r["source_url"])
        log_md.append(f"\n### {c}#{iid} {name}")
        log_md.append(f"- 专著数：{val}")
        books = r.get("book_list", [])
        if books:
            log_md.append(f"- 清单：")
            for b in books[:10]: log_md.append(f"  - {b}")
        log_md.append(f"- 置信度：{r.get('confidence','?')}")

    # C 组：写多口径计算表
    for (c, iid), r in zip(GROUP_C, c_results):
        name = fields[iid]["字段名"]
        multi_md.append(f"\n## {c}#{iid} {name}")
        nums = r.get("numerators", [])
        comps = r.get("computations", [])
        multi_md.append("\n### 分子候选")
        for i, n in enumerate(nums, 1):
            multi_md.append(f"{i}. {n.get('value','?')} ({n.get('year','?')}) — {n.get('url','')}")
        multi_md.append("\n### 分母候选（业务侧指定）")
        for i, d in enumerate(DENOMINATORS.get((c,iid), []), 1):
            multi_md.append(f"{i}. {d[0]}")
        multi_md.append("\n### 计算组合")
        multi_md.append("| 分子# | 分母# | 公式 | 结果 | 使用限制 |")
        multi_md.append("|---|---|---|---|---|")
        for comp in comps:
            multi_md.append(f"| {comp.get('numerator_idx','?')} | {comp.get('denominator_idx','?')} | "
                           f"`{comp.get('formula','')}` | **{comp.get('value','?')}** | {comp.get('caveat','')} |")
        multi_md.append(f"\n**推荐口径**：{r.get('recommended','')} (置信度 {r.get('confidence_best','?')})")

        # 在 xlsx 中把 C 组格更新为多套值
        vcol, scol = COL[c]
        row = next(rr for rr in range(2, ws.max_row+1) if ws.cell(rr,1).value == iid)
        if comps:
            multi_summary = "【估算-多口径】" + " | ".join(
                f"{comp.get('value','?')}（{comp.get('formula','')}）"
                for comp in comps[:3]
            )
            ws.cell(row, vcol, multi_summary).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, vcol).fill = est_fill

    wb.save(V4)
    LOG_MD.write_text("\n".join(log_md), encoding="utf-8")
    MULTI_MD.write_text("\n".join(multi_md), encoding="utf-8")
    log.info("saved %s, %s, %s", V4.name, LOG_MD.name, MULTI_MD.name)

if __name__ == "__main__":
    asyncio.run(run())
