"""统一批次 pipeline：建模板 → 初填 → 代理推导 → 精修。

Usage:
    python tools/run_batch_pipeline.py --batch 3 \
        --countries "文莱,东帝汶,不丹,孟加拉国,尼泊尔"

每批产出:
  topics/..._批次N_五国_v_final_YYYYMMDD.xlsx
  topics/..._批次N_summary_YYYYMMDD.md
"""
from __future__ import annotations
import argparse, asyncio, io, json, logging, os, re, sys, time
from dataclasses import dataclass, asdict, field
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "tools"))
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError: pass

import httpx, openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from news_monitor.proxy import build_httpx_client, get_proxies_for_url
from academic_search import academic_search, format_results

TOPICS = ROOT / "topics"
FIELD_DICT = TOPICS / "bri_field_dict.json"
DATE = datetime.now().strftime("%Y%m%d")

PPX = os.environ.get("PERPLEXITY_API_KEY", "")
TAV = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
BRV = os.environ.get("BRAVEAPI", "")
DS = os.environ.get("DEEPSEEK_API_KEY", "")
PROXY = os.environ.get("LLM_PROXY", "")

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger("pipe")

QUALITATIVE = {2,3,17,51,52,64}
BAD_TERMS = ['无可核验','无确切','无具体','无官方','无数据','无统计','无公开','无法','未发布',
             '未公布','未披露','未见','未找到','未统计','未知','暂无','尚无','不公开','不详',
             '待核','N/A','无法确定','无法计算','无从查证','缺乏','无直接','无明确','无系统']
CHINA_SIDE = ["hanban.org","chinese.cn","ci.cn","ccn.edu.cn","confucius.edu.cn",
              "mfa.gov.cn","fmprc.gov.cn","moe.gov.cn","edu.cn","xinhuanet.com",
              "chinanews.com","people.com.cn"]

# 国家 → (官方域, Brave CC, EN 名称)
COUNTRY_META = {
    # 南亚
    "文莱": (["mfa.gov.bn","moe.gov.bn","gov.bn","ubd.edu.bn"], "", "Brunei"),
    "东帝汶": (["mfac.gov.tl","gov.tl","unp.edu.tl"], "", "Timor-Leste"),
    "不丹": (["moe.gov.bt","mfa.gov.bt","gov.bt","rub.edu.bt"], "", "Bhutan"),
    "孟加拉国": (["moedu.gov.bd","mofa.gov.bd","gov.bd","du.ac.bd"], "", "Bangladesh"),
    "尼泊尔": (["moe.gov.np","mofa.gov.np","gov.np","tribhuvan-university.edu.np"], "", "Nepal"),
    "斯里兰卡": (["moe.gov.lk","mfa.gov.lk","gov.lk","cmb.ac.lk"], "", "Sri Lanka"),
    "印度": (["education.gov.in","mea.gov.in","gov.in","du.ac.in"], "IN", "India"),
    "巴基斯坦": (["mofept.gov.pk","mofa.gov.pk","gov.pk","qau.edu.pk"], "", "Pakistan"),
    "阿富汗": (["moe.gov.af","mfa.gov.af","gov.af"], "", "Afghanistan"),
    "马尔代夫": (["moe.gov.mv","foreign.gov.mv","gov.mv","mnu.edu.mv"], "", "Maldives"),
    # 中亚
    "土库曼斯坦": (["edu.gov.tm","mfa.gov.tm","gov.tm"], "", "Turkmenistan"),
    "乌兹别克斯坦": (["edu.uz","mfa.uz","gov.uz","nuu.uz"], "", "Uzbekistan"),
    "哈萨克斯坦": (["edu.gov.kz","mfa.gov.kz","gov.kz","kaznu.kz"], "", "Kazakhstan"),
    "塔吉克斯坦": (["moa.tj","mfa.tj","gov.tj","tnu.tj"], "", "Tajikistan"),
    "吉尔吉斯斯坦": (["edu.gov.kg","mfa.gov.kg","gov.kg","kyrgyz-natgeo.edu.kg"], "", "Kyrgyzstan"),
    # 西亚
    "阿曼": (["moe.gov.om","fm.gov.om","omanportal.gov.om","squ.edu.om"], "", "Oman"),
    "伊朗": (["medu.ir","mfa.ir","ir","ut.ac.ir"], "", "Iran"),
    "阿联酋": (["moe.gov.ae","mofaic.gov.ae","government.ae","uaeu.ac.ae"], "", "UAE"),
    "卡塔尔": (["edu.gov.qa","mofa.gov.qa","qu.edu.qa"], "", "Qatar"),
    "巴林": (["moe.gov.bh","mofa.gov.bh","uob.edu.bh"], "", "Bahrain"),
    "沙特阿拉伯": (["moe.gov.sa","mofa.gov.sa","gov.sa","ksu.edu.sa"], "SA", "Saudi Arabia"),
    "科威特": (["moe.edu.kw","mofa.gov.kw","kuniv.edu.kw"], "", "Kuwait"),
    "也门": (["yemen.gov.ye","mofa.gov.ye","ye"], "", "Yemen"),
    "土耳其": (["meb.gov.tr","mfa.gov.tr","turkiye.gov.tr","istanbul.edu.tr"], "TR", "Turkey"),
    "伊拉克": (["moedu.gov.iq","mofa.gov.iq","iq","uobaghdad.edu.iq"], "", "Iraq"),
    "叙利亚": (["syrianeducation.org.sy","mofaex.gov.sy","damascus-university.edu.sy"], "", "Syria"),
    "约旦": (["moe.gov.jo","mfa.gov.jo","jo","ju.edu.jo"], "", "Jordan"),
    "黎巴嫩": (["mehe.gov.lb","general-security.gov.lb","lb","ul.edu.lb"], "", "Lebanon"),
    "以色列": (["edu.gov.il","mfa.gov.il","gov.il","huji.ac.il","tau.ac.il"], "IL", "Israel"),
    "格鲁吉亚": (["mes.gov.ge","mfa.gov.ge","tsu.ge"], "", "Georgia"),
    "亚美尼亚": (["edu.am","mfa.am","ysu.am"], "", "Armenia"),
    "阿塞拜疆": (["edu.gov.az","mfa.gov.az","gov.az","bsu.edu.az"], "", "Azerbaijan"),
    "埃及": (["moe.gov.eg","mfa.gov.eg","egypt.gov.eg","cu.edu.eg"], "EG", "Egypt"),
    # 欧洲
    "爱沙尼亚": (["hm.ee","vm.ee","ee","ut.ee"], "EE", "Estonia"),
    "拉脱维亚": (["izm.gov.lv","mfa.gov.lv","lv","lu.lv"], "LV", "Latvia"),
    "立陶宛": (["smm.lrv.lt","urm.lt","lt","vu.lt"], "LT", "Lithuania"),
    "德国": (["bmbf.de","auswaertiges-amt.de","de","uni-heidelberg.de","hu-berlin.de"], "DE", "Germany"),
    "波兰": (["gov.pl","msz.gov.pl","pl","uw.edu.pl"], "PL", "Poland"),
    "捷克": (["msmt.cz","mzv.cz","cz","cuni.cz"], "CZ", "Czech Republic"),
    "斯洛文尼亚": (["mvzi.gov.si","mzz.gov.si","si","uni-lj.si"], "SI", "Slovenia"),
    "克罗地亚": (["mzo.gov.hr","mvep.gov.hr","hr","unizg.hr"], "HR", "Croatia"),
    "塞尔维亚": (["mpn.gov.rs","mfa.gov.rs","srbija.gov.rs","bg.ac.rs"], "RS", "Serbia"),
    "波黑": (["mcp.gov.ba","mvp.gov.ba","ba","unsa.ba"], "BA", "Bosnia and Herzegovina"),
    "黑山": (["mpin.gov.me","mvp.gov.me","me","ucg.ac.me"], "ME", "Montenegro"),
    "阿尔巴尼亚": (["arsimi.gov.al","punetejashtme.gov.al","al","unitir.edu.al"], "AL", "Albania"),
    "匈牙利": (["kormany.hu","mfa.gov.hu","hu","elte.hu"], "HU", "Hungary"),
    "斯洛伐克": (["minedu.sk","mzv.sk","slovensko.sk","uniba.sk"], "SK", "Slovakia"),
    "北马其顿": (["mon.gov.mk","mfa.gov.mk","mk","ukim.edu.mk"], "MK", "North Macedonia"),
    "罗马尼亚": (["edu.ro","mae.ro","gov.ro","unibuc.ro"], "RO", "Romania"),
    "保加利亚": (["mon.bg","mfa.bg","bg","uni-sofia.bg"], "BG", "Bulgaria"),
    "摩尔多瓦": (["mecc.gov.md","mfa.gov.md","gov.md","usm.md"], "MD", "Moldova"),
    "乌克兰": (["mon.gov.ua","mfa.gov.ua","gov.ua","knu.ua"], "UA", "Ukraine"),
    "白俄罗斯": (["edu.gov.by","mfa.gov.by","belarus.by","bsu.by"], "BY", "Belarus"),
    "俄罗斯": (["edu.gov.ru","mid.ru","gov.ru","msu.ru","spbu.ru"], "RU", "Russia"),
}

# ── API ─────────────────────────────────────────────────────────────────────
async def ppx_ask(client, q, sys_msg=None):
    if not PPX: return "", []
    msg = sys_msg or "你是研究员。给事实+URL。若无直接数据给代理推导值，绝不返回'无可核验数据'。"
    try:
        r = await client.post("https://api.perplexity.ai/chat/completions",
            json={"model":"sonar","messages":[
                {"role":"system","content":msg},
                {"role":"user","content":q}],
                "temperature":0.1,"return_citations":True},
            headers={"Authorization":f"Bearer {PPX}","Content-Type":"application/json"},
            timeout=90.0)
        r.raise_for_status()
        d = r.json()
        return d["choices"][0]["message"]["content"], (d.get("citations",[]) or [])[:10]
    except Exception as e:
        log.warning("ppx: %s", str(e)[:80]); return "", []

async def tav(client, q, domains=None):
    if not TAV: return ([], "")
    try:
        body = {"api_key":TAV,"query":q,"search_depth":"advanced","max_results":8,
                "include_answer":True,"include_raw_content":False}
        if domains: body["include_domains"] = domains[:15]
        r = await client.post("https://api.tavily.com/search", json=body, timeout=60.0)
        r.raise_for_status()
        d = r.json()
        items = [{"url":x.get("url",""),"title":x.get("title",""),
                 "snippet":(x.get("content") or "")[:400]}
                for x in (d.get("results") or [])[:8]]
        return (items, d.get("answer","") or "")
    except Exception as e:
        log.warning("tav: %s", str(e)[:80]); return ([], "")

async def brave(client, q, cc=""):
    if not BRV: return []
    params = {"q":q,"count":10,"text_decorations":False}
    valid_cc = {"AR","AU","BR","CA","CN","DE","FR","GB","HK","IN","ID","IT","JP","KR",
                "MX","MY","NL","PH","RU","SG","TW","US","TH","TR","SA","SI","HR","RS",
                "BA","ME","AL","HU","SK","MK","RO","BG","MD","UA","BY","IL","EG","EE",
                "LV","LT","PL","CZ"}
    if cc and cc in valid_cc: params["country"] = cc
    try:
        r = await client.get("https://api.search.brave.com/res/v1/web/search", params=params,
            headers={"Accept":"application/json","X-Subscription-Token":BRV}, timeout=30.0)
        r.raise_for_status()
        return [{"url":x.get("url",""),"title":x.get("title",""),
                 "snippet":(x.get("description") or "")[:300]}
                for x in (r.json().get("web",{}).get("results") or [])[:10]]
    except Exception as e:
        log.warning("brave: %s", str(e)[:80]); return []

async def ds_call(client, prompt, json_mode=True):
    if not DS: return {}
    try:
        payload = {"model":"deepseek-chat","messages":[{"role":"user","content":prompt}],
                   "temperature":0.1}
        if json_mode: payload["response_format"] = {"type":"json_object"}
        r = await client.post("https://api.deepseek.com/v1/chat/completions",
            json=payload,
            headers={"Authorization":f"Bearer {DS}","Content-Type":"application/json"},
            timeout=120.0)
        r.raise_for_status()
        content = r.json()["choices"][0]["message"]["content"]
        return json.loads(content) if json_mode else content
    except Exception as e:
        log.warning("ds: %s", str(e)[:80]); return {}

# ── Build template ──────────────────────────────────────────────────────────
def build_template(batch_id, countries, out_path):
    with open(FIELD_DICT, encoding='utf-8') as f:
        fields = [x for x in json.load(f) if x.get('指标ID')]
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"批次{batch_id}-5国"
    headers = ['序号','搜集字段','字段类型','字段说明']
    for c in countries:
        headers.extend([c, f"{c}出处"])
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    hal = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="999999")
    bd = Border(top=thin, bottom=thin, left=thin, right=thin)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = hal; cell.border = bd
    for i, f in enumerate(fields):
        r = i + 2
        ws.cell(r, 1, f.get('指标ID'))
        ws.cell(r, 2, f.get('字段名'))
        ws.cell(r, 3, f.get('字段类型'))
        ws.cell(r, 4, f.get('字段说明'))
        for c in range(1, 4 + 2*len(countries) + 1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, c).border = bd
    widths = [6,22,22,40] + [20,30] * len(countries)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "E2"; ws.row_dimensions[1].height = 28
    wb.save(out_path)
    log.info("template built: %s", out_path.name)

def col_map(countries):
    """returns {country: (value_col, source_col)}"""
    return {c: (5 + 2*i, 6 + 2*i) for i, c in enumerate(countries)}

# ── Stage 1: initial fill ───────────────────────────────────────────────────
async def initial_fill(xlsx_path, countries, ckpt_path):
    with open(FIELD_DICT, encoding='utf-8') as f:
        fields = [x for x in json.load(f) if x.get('指标ID')]
    cmap = col_map(countries)

    ckpt = {}
    if ckpt_path.exists():
        try: ckpt = json.loads(ckpt_path.read_text(encoding='utf-8'))
        except: pass

    # find gaps
    wb = openpyxl.load_workbook(xlsx_path, data_only=True); ws = wb.active
    gaps = []
    for cname in countries:
        vcol, _ = cmap[cname]
        for i, fld in enumerate(fields):
            row = i + 2
            v = ws.cell(row, vcol).value
            if v is None or (isinstance(v,str) and not v.strip()):
                iid = int(fld.get('指标ID') or 0)
                if iid == 0: continue
                gaps.append({
                    'country': cname, 'id': iid, 'row': row, 'vcol': vcol,
                    'name': fld.get('字段名',''), 'type': fld.get('字段类型',''),
                    'desc': (fld.get('字段说明','') or '')[:200],
                    'category': fld.get('分类',''), 'priority': fld.get('优先级','P?'),
                    'hint_cn': fld.get('中文检索提示',''), 'hint_en': fld.get('英文检索提示',''),
                })
    log.info("Stage 1 initial fill: %d gaps", len(gaps))

    ppx_p = get_proxies_for_url("https://api.perplexity.ai", PROXY)
    brv_p = get_proxies_for_url("https://api.search.brave.com", PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", PROXY)
    async with (build_httpx_client(ppx_p or "", 90) as ppx_c,
                build_httpx_client(brv_p or "", 30) as brv_c,
                build_httpx_client(tav_p or "", 60) as tav_c,
                build_httpx_client("", 90) as ds_c):
        sem = asyncio.Semaphore(5)
        async def _task(g):
            key = f"{g['country']}::{g['id']}"
            if key in ckpt: return ckpt[key]
            async with sem:
                t0 = time.time()
                en = COUNTRY_META.get(g['country'], ([], "", g['country']))[2]
                official = COUNTRY_META.get(g['country'], ([], "", ""))[0] + CHINA_SIDE
                cc = COUNTRY_META.get(g['country'], ([], "", ""))[1]
                q = (f"请提供关于【{g['country']}】在指标【{g['name']}】上的最权威信息。\n"
                     f"- 取值口径：{g['type']}\n- 说明：{g['desc']}\n"
                     f"- 优先来源：官方政府/教育部/驻外使馆/双边声明/权威研究\n"
                     f"给出事实+来源 URL。")
                ppx_a, ppx_cites = await ppx_ask(ppx_c, q)
                q_en = f"{en} Chinese language education {g['name']}"
                q_cn = f"{g['country']} {g['name']}"
                br_r, tav_o, tav_g = await asyncio.gather(
                    brave(brv_c, q_en, cc), tav(tav_c, q_en, official),
                    tav(tav_c, q_cn, None), return_exceptions=True)
                evidence = [{'source':'perplexity','url':u} for u in ppx_cites]
                if isinstance(br_r, list): evidence.extend([{'source':'brave',**x} for x in br_r[:5]])
                if isinstance(tav_o, tuple): evidence.extend([{'source':'tavily',**x} for x in tav_o[0][:5]])
                if isinstance(tav_g, tuple): evidence.extend([{'source':'tavily',**x} for x in tav_g[0][:5]])
                evid_str = "\n".join(f"[{e.get('source')}] {e.get('title','')[:80]} | {e.get('url','')}\n  {e.get('snippet','')[:200]}" for e in evidence[:10])
                ext = await ds_call(ds_c, f"""为【{g['country']}】【{g['name']}】从证据中抽取事实。
取值口径：{g['type']}
说明：{g['desc']}

Perplexity 回答:
{ppx_a or '(空)'}

其他证据:
{evid_str}

返回 JSON: {{value, source_url, caliber, confidence, is_estimation, rationale}}""")
                r = {
                    'value': str(ext.get('value',''))[:500] or (ppx_a or '')[:300],
                    'source_url': ext.get('source_url') or (ppx_cites[0] if ppx_cites else ''),
                    'caliber': ext.get('caliber',''),
                    'confidence': ext.get('confidence','low'),
                    'is_estimation': bool(ext.get('is_estimation', False)),
                    'evidence': evidence[:12],
                }
                ckpt[key] = r
                with open(ckpt_path, 'w', encoding='utf-8') as f:
                    json.dump(ckpt, f, ensure_ascii=False)
                log.info("[S1 %s#%d] %.1fs conf=%s value=%s",
                         g['country'], g['id'], time.time()-t0, r['confidence'], str(r['value'])[:40])
                return r
        tasks = [_task(g) for g in gaps]
        await asyncio.gather(*tasks)

    # write back
    wb = openpyxl.load_workbook(xlsx_path); ws = wb.active
    est_fill = PatternFill(start_color="FFF5E6", fill_type="solid")
    low_fill = PatternFill(start_color="FFEAEA", fill_type="solid")
    for g in gaps:
        key = f"{g['country']}::{g['id']}"
        r = ckpt.get(key);
        if not r or not r.get('value'): continue
        existing = ws.cell(g['row'], g['vcol']).value
        if existing and str(existing).strip(): continue
        vcol, scol = g['vcol'], g['vcol']+1
        prefix = "【估算】" if r.get('is_estimation') else ""
        suffix = f"（口径：{r['caliber']}）" if r.get('caliber') else ""
        cell = ws.cell(g['row'], vcol, f"{prefix}{r['value']}{suffix}")
        src_u = r.get('source_url','')
        if isinstance(src_u, (list, tuple)): src_u = "; ".join(str(x) for x in src_u[:3])
        ws.cell(g['row'], scol, str(src_u)[:500])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if r.get('is_estimation'): cell.fill = est_fill
        elif r.get('confidence') == 'low': cell.fill = low_fill
    wb.save(xlsx_path)
    log.info("Stage 1 complete, %s", xlsx_path.name)

# ── Stage 2 + 3: proxy & refine (same algorithm) ────────────────────────────
async def proxy_or_refine(src_xlsx, dst_xlsx, countries, ckpt_path, stage_name="proxy"):
    """识别空/语义空/0/无，调 DS 扩展口径 + 多源搜索 + 合成."""
    with open(FIELD_DICT, encoding='utf-8') as f:
        fd = {int(x['指标ID']):x for x in json.load(f) if x.get('指标ID')}
    cmap = col_map(countries)

    wb = openpyxl.load_workbook(src_xlsx, data_only=True); ws = wb.active
    gaps = []
    for cname in countries:
        vcol, _ = cmap[cname]
        for r in range(2, ws.max_row+1):
            v_raw = ws.cell(r, vcol).value
            iid = ws.cell(r, 1).value
            if iid is None or iid == 36: continue
            fld = fd.get(int(iid));
            if not fld: continue
            v = str(v_raw or '').strip()
            is_empty = not v
            is_soft = any(b in v for b in BAD_TERMS)
            clean = re.sub(r'^(【估算】|【代理推导】|【代理】)\s*','',v)
            first = re.split(r'[（\(]', clean)[0].strip()
            is_zero = bool(re.match(r'^0(\.\d+)?(%|人|所|项|个|次|篇)?$', first)) or first in ('0','无','没有')
            # 精修模式：也处理软定性
            soft_terms = ['少量','若干','有限'] if stage_name == "refine" else []
            is_very_soft = first in soft_terms
            if int(iid) in QUALITATIVE and is_zero: continue
            if is_empty or is_soft or is_zero or is_very_soft:
                gaps.append({'country':cname,'id':int(iid),'row':r,'vcol':vcol,
                    'name':fld.get('字段名',''),'type':fld.get('字段类型',''),
                    'desc':(fld.get('字段说明','') or '')[:200],
                    'category':fld.get('分类',''),'priority':fld.get('优先级','?'),
                    'current':v[:100] if v else '(空)'})
    log.info("Stage %s: %d gaps", stage_name, len(gaps))
    if not gaps:
        wb.save(dst_xlsx); return 0

    ckpt = {}
    if ckpt_path.exists():
        try: ckpt = json.loads(ckpt_path.read_text(encoding='utf-8'))
        except: pass

    ppx_p = get_proxies_for_url("https://api.perplexity.ai", PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", PROXY)
    async with (build_httpx_client(ppx_p or "", 90) as ppx_c,
                build_httpx_client(tav_p or "", 60) as tav_c,
                build_httpx_client("", 90) as ds_c):
        sem = asyncio.Semaphore(4)
        async def _proc(g):
            key = f"{g['country']}::{g['id']}"
            if key in ckpt: return ckpt[key]
            async with sem:
                t0 = time.time()
                plan = await ds_call(ds_c, f"""为该"搜不到数据"的指标设计扩展代理方案。
国家：{g['country']}
字段：{g['name']}
取值口径：{g['type']}
说明：{g['desc']}
当前错误值：{g['current']}

返回 JSON: {{
  "extended_definition": "扩展口径说明",
  "queries": {{"perplexity":"具体提问","tavily_zh":"","tavily_en":"","academic":""}},
  "priority_domains": ["域名1","域名2"]
}}""")
                if not plan.get('queries'):
                    ckpt[key] = {"value":"无可用策略","confidence":"low"}
                    _persist_ckpt(ckpt, ckpt_path); return ckpt[key]
                q = plan['queries']; doms = plan.get('priority_domains', [])
                results = await asyncio.gather(
                    ppx_ask(ppx_c, q.get('perplexity','')),
                    tav(tav_c, q.get('tavily_zh',''), doms),
                    tav(tav_c, q.get('tavily_en',''), None),
                    academic_search(q.get('academic',''), max_per_source=3, proxy_url=PROXY, include_scholar=True),
                    return_exceptions=True)
                ppx_r = results[0] if isinstance(results[0], tuple) else ("", [])
                tav_zh = results[1] if isinstance(results[1], tuple) and len(results[1])==2 else ([],"")
                tav_en = results[2] if isinstance(results[2], tuple) and len(results[2])==2 else ([],"")
                acad = results[3] if isinstance(results[3], dict) else {}
                evid = f"""## Perplexity\n{ppx_r[0]}\nCitations: {chr(10).join(ppx_r[1][:6])}
## Tavily-zh answer: {tav_zh[1]}
{chr(10).join(f'  - {x["title"]} | {x["url"]}' for x in tav_zh[0][:4])}
## Tavily-en answer: {tav_en[1]}
{chr(10).join(f'  - {x["title"]} | {x["url"]}' for x in tav_en[0][:4])}
## Academic
{format_results(acad)[:3500]}"""[:12000]
                synth = await ds_call(ds_c, f"""根据扩展口径+证据给【{g['country']}】【{g['name']}】最终值。
扩展口径: {plan.get('extended_definition','')}
取值口径: {g['type']}
证据:
{evid}

绝不返回 0/无/合成失败。有数字给数字，无数字给区间，都没有给定性分级+依据。

返回 JSON: {{value, proxy_used, derivation, source_urls, confidence, is_estimation, rationale}}""")
                ckpt[key] = synth
                _persist_ckpt(ckpt, ckpt_path)
                log.info("[%s %s#%d] %.1fs value=%s", stage_name, g['country'], g['id'], time.time()-t0, str(synth.get('value',''))[:40])
                return synth
        tasks = [_proc(g) for g in gaps]
        await asyncio.gather(*tasks)

    # write back
    wb = openpyxl.load_workbook(src_xlsx); ws = wb.active
    fill_color = "FFF9E6" if stage_name == "proxy" else "E6FFFA"
    proxy_fill = PatternFill(start_color=fill_color, fill_type="solid")
    updated = 0
    for g in gaps:
        r = ckpt.get(f"{g['country']}::{g['id']}", {})
        val = str(r.get('value','')).strip()
        if not val or val in ('合成失败','无可用策略'): continue
        if re.match(r'^(0|无|没有|0%)$', val): continue
        updated += 1
        prefix = "【代理推导】" if stage_name == "proxy" else "【精修-广义】"
        proxy = r.get('proxy_used',''); deriv = r.get('derivation','')
        txt = f"{prefix}{val}"
        if proxy: txt += f"（代理：{proxy}）"
        if deriv: txt += f"\n推导：{deriv}"
        ws.cell(g['row'], g['vcol'], txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(g['row'], g['vcol']).fill = proxy_fill
        urls = r.get('source_urls', [])
        if urls:
            if isinstance(urls, str): urls = [urls]
            ws.cell(g['row'], g['vcol']+1, "; ".join(str(u) for u in urls[:3])[:500])
    wb.save(dst_xlsx)
    log.info("Stage %s complete: %d/%d updated", stage_name, updated, len(gaps))
    return updated

def _persist_ckpt(ckpt, path):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(ckpt, f, ensure_ascii=False)

# ── Main pipeline ───────────────────────────────────────────────────────────
async def run_pipeline(batch_id, countries, region_tag=""):
    tag = f"批次{batch_id}" + (f"_{region_tag}" if region_tag else "")
    base_name = f"“一带一路”沿线国家中文教育发展指标体系数据总表_{tag}_五国"
    tmpl_path = TOPICS / f"{base_name}_{DATE}.xlsx"
    v2_path = TOPICS / f"{base_name}_v2_{DATE}.xlsx"
    v3_path = TOPICS / f"{base_name}_v3_{DATE}.xlsx"
    v4_path = TOPICS / f"{base_name}_v_final_{DATE}.xlsx"

    ckpt_s1 = ROOT / "output" / f"bri_{tag}_s1_ckpt.json"
    ckpt_s2 = ROOT / "output" / f"bri_{tag}_s2_ckpt.json"
    ckpt_s3 = ROOT / "output" / f"bri_{tag}_s3_ckpt.json"

    t0 = time.time()
    # Stage 0: build template
    if not tmpl_path.exists():
        build_template(batch_id, countries, tmpl_path)
    import shutil
    if not v2_path.exists():
        shutil.copy(tmpl_path, v2_path)

    # Stage 1: initial fill
    log.info("===== Stage 1 (initial fill) =====")
    await initial_fill(v2_path, countries, ckpt_s1)
    log.info("Stage 1 done in %.1fs", time.time()-t0)

    # Stage 2: proxy
    log.info("===== Stage 2 (proxy derivation) =====")
    t1 = time.time()
    import shutil
    if not v3_path.exists(): shutil.copy(v2_path, v3_path)
    await proxy_or_refine(v2_path, v3_path, countries, ckpt_s2, "proxy")
    log.info("Stage 2 done in %.1fs", time.time()-t1)

    # Stage 3: refine
    log.info("===== Stage 3 (refinement) =====")
    t2 = time.time()
    if not v4_path.exists(): shutil.copy(v3_path, v4_path)
    await proxy_or_refine(v3_path, v4_path, countries, ckpt_s3, "refine")
    log.info("Stage 3 done in %.1fs", time.time()-t2)

    # Summary
    log.info("===== Pipeline complete for batch %s in %.1fs =====",
             batch_id, time.time()-t0)
    log.info("Final: %s", v4_path.name)

    # stats
    wb = openpyxl.load_workbook(v4_path, data_only=True); ws = wb.active
    cmap = col_map(countries)
    per_country = {}
    for cname in countries:
        vcol, _ = cmap[cname]
        d = {'filled':0,'proxy':0,'refined':0,'est':0,'empty':0}
        for r in range(2, ws.max_row+1):
            v = str(ws.cell(r, vcol).value or '').strip()
            if not v: d['empty']+=1; continue
            if '【精修-广义】' in v: d['refined']+=1
            elif '【代理推导】' in v or '【代理】' in v: d['proxy']+=1
            elif '【估算' in v: d['est']+=1
            else: d['filled']+=1
        per_country[cname] = d
        log.info("  %s: %s", cname, d)
    return v4_path

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--batch", type=int, required=True)
    ap.add_argument("--countries", required=True, help="逗号分隔中文国家名")
    ap.add_argument("--region", default="")
    args = ap.parse_args()
    countries = [c.strip() for c in args.countries.split(",")]
    asyncio.run(run_pipeline(args.batch, countries, args.region))
