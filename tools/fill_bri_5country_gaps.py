"""一带一路 5 国中文教育指标缺失回填脚本.

管道：Perplexity 初查 (带 citations) → Brave + Tavily 扩证（官方域优先）→ DeepSeek 结构化抽取

运行:
    # 试跑：每国前 2 个指标
    python tools/fill_bri_5country_gaps.py --sample 2
    # 全量：
    python tools/fill_bri_5country_gaps.py
    # 仅跑指定国家:
    python tools/fill_bri_5country_gaps.py --country 蒙古

输出:
    topics/“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v2_{date}.xlsx
    topics/bri_5country_evidence_log_{date}.md
    topics/bri_5country_estimation_notes_{date}.md
    output/bri_gap_fill_checkpoint.json   (断点续跑)
"""
from __future__ import annotations

import argparse
import asyncio
import io
import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Any

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))

try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass

import httpx
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from news_monitor.proxy import build_httpx_client, get_proxies_for_url

# ── paths ────────────────────────────────────────────────────────────────────
TOPICS = ROOT / "topics"
OUTPUT_DIR = ROOT / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
SRC_XLSX = TOPICS / "“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版.xlsx"
FIELD_DICT = TOPICS / "bri_field_dict.json"
CHECKPOINT = OUTPUT_DIR / "bri_gap_fill_checkpoint.json"

DATE_TAG = datetime.now().strftime("%Y%m%d")
V2_XLSX = TOPICS / f"“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v2_{DATE_TAG}.xlsx"
EVIDENCE_MD = TOPICS / f"bri_5country_evidence_log_{DATE_TAG}.md"
ESTIMATION_MD = TOPICS / f"bri_5country_estimation_notes_{DATE_TAG}.md"

# ── constants ────────────────────────────────────────────────────────────────
COUNTRIES = [
    # (name_cn, value_col, source_col, official_domains, lang_code)
    ("蒙古",   5,  6,  ["gov.mn", "mecss.gov.mn", "edu.mn", "mfa.gov.mn", "ulaanbaatar.mn"],
     "mn"),
    ("越南",   7,  8,  ["moet.gov.vn", "mofa.gov.vn", "chinhphu.vn", "gov.vn", "vneconomy.vn", "vnexpress.net"],
     "vi"),
    ("柬埔寨", 9,  10, ["moeys.gov.kh", "mfaic.gov.kh", "rupp.edu.kh", "pressocm.gov.kh", "gov.kh"],
     "km"),
    ("菲律宾", 11, 12, ["deped.gov.ph", "ched.gov.ph", "officialgazette.gov.ph", "dfa.gov.ph", "gov.ph"],
     "en"),
    ("新加坡", 13, 14, ["moe.gov.sg", "mfa.gov.sg", "gov.sg", "nus.edu.sg", "ntu.edu.sg", "straitstimes.com"],
     "en"),
]
COUNTRY_MAP = {c[0]: c for c in COUNTRIES}
CHINA_SIDE_DOMAINS = [
    "hanban.org", "chinese.cn", "ci.cn", "ccn.edu.cn", "confucius.edu.cn",
    "mfa.gov.cn", "fmprc.gov.cn", "moe.gov.cn", "edu.cn",
    "xinhuanet.com", "chinanews.com", "people.com.cn",
]

# ── API keys ─────────────────────────────────────────────────────────────────
PERPLEXITY_API_KEY = os.environ.get("PERPLEXITY_API_KEY", "")
BRAVE_API_KEY = os.environ.get("BRAVEAPI", "") or os.environ.get("BRAVE_API_KEY", "")
TAVILY_API_KEY = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
QWEN_API_KEY = os.environ.get("QWEN_API_KEY", "")
OVERSEAS_PROXY = os.environ.get("LLM_PROXY", "") or os.environ.get("OVERSEAS_PROXY", "")

# ── logging ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("fill_bri")

# ── data classes ─────────────────────────────────────────────────────────────
@dataclass
class Indicator:
    id: int
    row: int
    name: str
    ftype: str
    category: str
    priority: str
    desc: str
    hint_cn: str
    hint_en: str

@dataclass
class Evidence:
    source: str  # perplexity / brave / tavily
    url: str = ""
    title: str = ""
    snippet: str = ""

@dataclass
class CellResult:
    country: str
    indicator_id: int
    indicator_name: str
    value: str = ""
    source_url: str = ""
    caliber: str = ""        # 口径说明 (time point/unit/numerator-denominator)
    confidence: str = "low"  # high / medium / low
    is_estimation: bool = False
    evidence: list[Evidence] = field(default_factory=list)
    error: str = ""

# ── API clients (direct HTTP, tailored for fact QA) ──────────────────────────
async def ask_perplexity(client: httpx.AsyncClient, question: str) -> tuple[str, list[str]]:
    """One-shot fact QA. Returns (answer, citations)."""
    if not PERPLEXITY_API_KEY:
        return "", []
    body = {
        "model": "sonar",
        "messages": [
            {"role": "system", "content":
             "You are a meticulous research assistant. Answer the user's factual question about education/policy "
             "with a concise answer and cite every claim. Prefer official government and authoritative sources. "
             "If the fact is uncertain or unavailable, say so explicitly."},
            {"role": "user", "content": question},
        ],
        "temperature": 0.1,
        "return_citations": True,
    }
    try:
        resp = await client.post(
            "https://api.perplexity.ai/chat/completions",
            json=body,
            headers={"Authorization": f"Bearer {PERPLEXITY_API_KEY}", "Content-Type": "application/json"},
            timeout=60.0,
        )
        resp.raise_for_status()
        data = resp.json()
        content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        citations = data.get("citations", []) or []
        return content, citations[:8]
    except Exception as exc:
        log.warning("perplexity err: %s", str(exc)[:120])
        return "", []

async def search_brave(client: httpx.AsyncClient, query: str, country_code: str = "") -> list[Evidence]:
    if not BRAVE_API_KEY:
        return []
    _BRAVE_CC = {"mn": "", "vi": "", "km": "", "en": "US"}  # Brave lacks most SE Asian codes
    params = {"q": query, "count": 10, "text_decorations": False}
    if country_code and country_code in {"AR","AU","BR","CA","CN","DE","FR","GB","HK","IN","ID","IT","JP","KR","MX","MY","NL","PH","RU","SG","TW","US"}:
        params["country"] = country_code
    try:
        resp = await client.get(
            "https://api.search.brave.com/res/v1/web/search",
            params=params,
            headers={"Accept": "application/json", "X-Subscription-Token": BRAVE_API_KEY},
            timeout=30.0,
        )
        resp.raise_for_status()
        data = resp.json()
        out = []
        for r in (data.get("web", {}).get("results") or [])[:10]:
            out.append(Evidence(source="brave", url=r.get("url", ""),
                                title=r.get("title", ""), snippet=(r.get("description") or "")[:300]))
        return out
    except Exception as exc:
        log.warning("brave err: %s", str(exc)[:120])
        return []

async def search_tavily(client: httpx.AsyncClient, query: str, include_domains: list[str] | None = None) -> list[Evidence]:
    if not TAVILY_API_KEY:
        return []
    body = {
        "api_key": TAVILY_API_KEY,
        "query": query,
        "search_depth": "advanced",  # advanced gives more content
        "max_results": 8,
        "include_answer": False,
        "include_raw_content": False,
    }
    if include_domains:
        body["include_domains"] = include_domains[:20]
    try:
        resp = await client.post("https://api.tavily.com/search", json=body, timeout=45.0)
        resp.raise_for_status()
        data = resp.json()
        out = []
        for r in (data.get("results") or [])[:8]:
            out.append(Evidence(source="tavily", url=r.get("url", ""),
                                title=r.get("title", ""), snippet=(r.get("content") or "")[:400]))
        return out
    except Exception as exc:
        log.warning("tavily err: %s", str(exc)[:120])
        return []

async def llm_extract(client_direct: httpx.AsyncClient, ind: Indicator, country: str,
                      perplexity_answer: str, web_results: list[Evidence]) -> dict:
    """DeepSeek: extract structured {value, source_url, caliber, confidence, is_estimation}."""
    if not DEEPSEEK_API_KEY:
        # Fallback: just take perplexity answer as-is
        return {"value": (perplexity_answer or "").strip()[:500],
                "source_url": "", "caliber": "", "confidence": "low", "is_estimation": False}

    evidence_block = "\n".join(
        f"[{i+1}] ({e.source}) {e.title}\nURL: {e.url}\n摘要: {e.snippet}"
        for i, e in enumerate(web_results[:10])
    )
    prompt = f"""你是专门抽取事实的助手。请从下面的证据中为【{country}】在指标【{ind.name}】上给出回答。

## 指标信息
- 字段类型/取值口径：{ind.ftype}
- 字段说明：{ind.desc}
- 优先级：{ind.priority}  分类：{ind.category}

## Perplexity 初查摘要
{perplexity_answer or '(空)'}

## 扩证网页结果
{evidence_block or '(无)'}

## 产出要求
严格返回 JSON（不要多余文本），字段：
- value: 回答内容，务必贴合"字段类型"口径。若是数值/比值，写出具体数字 + 单位；若为分类，选最贴近的一项。若证据不足，允许写估算值但 is_estimation=true 并注明口径。
- source_url: 首选最具权威性的一条出处（官方/教育部/驻外使馆/权威研究）。
- caliber: 口径说明（年份、分子分母、统计范围等），没有则写空串。
- confidence: high|medium|low
- is_estimation: 是否基于估算而非原始数据（布尔）
- rationale: 30字以内简要说明依据。

只输出 JSON 对象。
"""
    try:
        resp = await client_direct.post(
            "https://api.deepseek.com/v1/chat/completions",
            json={
                "model": "deepseek-chat",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.1,
                "response_format": {"type": "json_object"},
            },
            headers={"Authorization": f"Bearer {DEEPSEEK_API_KEY}", "Content-Type": "application/json"},
            timeout=60.0,
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        parsed = json.loads(content)
        return {
            "value": str(parsed.get("value", "")).strip(),
            "source_url": str(parsed.get("source_url", "")).strip(),
            "caliber": str(parsed.get("caliber", "")).strip(),
            "confidence": str(parsed.get("confidence", "low")).lower(),
            "is_estimation": bool(parsed.get("is_estimation", False)),
            "rationale": str(parsed.get("rationale", "")).strip(),
        }
    except Exception as exc:
        log.warning("deepseek err (%s/%s): %s", country, ind.id, str(exc)[:120])
        # Fallback: raw perplexity
        first_url = web_results[0].url if web_results else ""
        return {"value": (perplexity_answer or "")[:300], "source_url": first_url,
                "caliber": "", "confidence": "low", "is_estimation": False, "rationale": "LLM 抽取失败，用原始摘要"}

# ── query builder ───────────────────────────────────────────────────────────
def build_perplexity_question(country: str, ind: Indicator) -> str:
    return (
        f"请提供关于【{country}】在指标【{ind.name}】上的最权威信息。\n"
        f"- 字段取值口径：{ind.ftype}\n"
        f"- 字段说明：{ind.desc}\n"
        f"- 推荐优先来源：官方政府/教育部/驻外使馆/双边声明/权威研究\n\n"
        f"请给出：(1) 明确的事实/数据（按口径）；(2) 对应的官方来源 URL；"
        f"(3) 若有多个口径/年份，列出差异。中文回答。"
    )

def build_web_query(country: str, ind: Indicator, lang: str = "cn") -> str:
    if lang == "en":
        en_hint = ind.hint_en.replace("国家名", country)
        return f"{en_hint} Chinese language education {country}"
    return ind.hint_cn.replace("国家名", country)

# ── core pipeline ────────────────────────────────────────────────────────────
async def fill_one_cell(sem: asyncio.Semaphore, country: str, ind: Indicator,
                         ppx_client: httpx.AsyncClient, brave_client: httpx.AsyncClient,
                         tavily_client: httpx.AsyncClient, ds_client: httpx.AsyncClient,
                         checkpoint: dict) -> CellResult:
    key = f"{country}::{ind.id}"
    if key in checkpoint:
        cached = checkpoint[key]
        r = CellResult(country=country, indicator_id=ind.id, indicator_name=ind.name,
                       value=cached.get("value", ""), source_url=cached.get("source_url", ""),
                       caliber=cached.get("caliber", ""), confidence=cached.get("confidence", "low"),
                       is_estimation=bool(cached.get("is_estimation")),
                       error=cached.get("error", ""))
        r.evidence = [Evidence(**e) for e in cached.get("evidence", [])]
        return r

    async with sem:
        result = CellResult(country=country, indicator_id=ind.id, indicator_name=ind.name)
        country_meta = COUNTRY_MAP[country]
        official_domains = country_meta[3] + CHINA_SIDE_DOMAINS
        country_cc = {"蒙古": "", "越南": "", "柬埔寨": "", "菲律宾": "PH", "新加坡": "SG"}[country]

        # Step 1: Perplexity
        ppx_q = build_perplexity_question(country, ind)
        ppx_answer, ppx_cites = await ask_perplexity(ppx_client, ppx_q)
        for url in ppx_cites:
            result.evidence.append(Evidence(source="perplexity", url=url, title="", snippet=""))

        # Step 2: Brave + Tavily (concurrent)
        q_cn = build_web_query(country, ind, "cn")
        q_en = build_web_query(country, ind, "en")
        brave_res, tavily_official, tavily_general = await asyncio.gather(
            search_brave(brave_client, q_en, country_cc),
            search_tavily(tavily_client, q_en, include_domains=official_domains),
            search_tavily(tavily_client, q_cn, include_domains=None),
            return_exceptions=True,
        )
        for r in (brave_res, tavily_official, tavily_general):
            if isinstance(r, list):
                result.evidence.extend(r[:5])
            elif isinstance(r, Exception):
                log.warning("  search err: %s", str(r)[:80])

        if not ppx_answer and not any(e.url for e in result.evidence):
            result.error = "no evidence"
            _save_cell(checkpoint, key, result)
            return result

        # Step 3: LLM synthesize
        extracted = await llm_extract(ds_client, ind, country, ppx_answer, result.evidence)
        result.value = extracted["value"]
        result.source_url = extracted["source_url"] or (result.evidence[0].url if result.evidence else "")
        result.caliber = extracted["caliber"]
        result.confidence = extracted["confidence"]
        result.is_estimation = extracted["is_estimation"]
        _save_cell(checkpoint, key, result)
        return result

def _save_cell(checkpoint: dict, key: str, r: CellResult):
    checkpoint[key] = {
        "value": r.value, "source_url": r.source_url, "caliber": r.caliber,
        "confidence": r.confidence, "is_estimation": r.is_estimation,
        "evidence": [asdict(e) for e in r.evidence[:12]],
        "error": r.error,
    }
    # Persist after each cell
    with open(CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump(checkpoint, f, ensure_ascii=False, indent=1)

# ── load/write Excel ─────────────────────────────────────────────────────────
def load_indicators() -> list[Indicator]:
    with open(FIELD_DICT, "r", encoding="utf-8") as f:
        raw = json.load(f)
    inds = []
    for i, f_ in enumerate(raw):
        inds.append(Indicator(
            id=int(f_["指标ID"]), row=i + 2,
            name=str(f_["字段名"] or ""),
            ftype=str(f_["字段类型"] or ""),
            category=str(f_["分类"] or ""),
            priority=str(f_["优先级"] or ""),
            desc=str(f_["字段说明"] or ""),
            hint_cn=str(f_["中文检索提示"] or ""),
            hint_en=str(f_["英文检索提示"] or ""),
        ))
    return inds

def find_gaps(inds: list[Indicator], only_countries: list[str] | None = None) -> list[tuple[str, Indicator]]:
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb.active
    gaps: list[tuple[str, Indicator]] = []
    for country, vcol, _, _, _ in COUNTRIES:
        if only_countries and country not in only_countries:
            continue
        for ind in inds:
            v = ws.cell(row=ind.row, column=vcol).value
            if v is None or (isinstance(v, str) and not v.strip()):
                gaps.append((country, ind))
    return gaps

def write_v2(inds: list[Indicator], results_by_key: dict[str, CellResult]):
    """Read src xlsx as template, fill gaps, save to V2_XLSX."""
    wb = openpyxl.load_workbook(SRC_XLSX)
    ws = wb.active
    est_fill = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")
    low_fill = PatternFill(start_color="FFEAEA", end_color="FFEAEA", fill_type="solid")

    for country, vcol, scol, _, _ in COUNTRIES:
        for ind in inds:
            key = f"{country}::{ind.id}"
            r = results_by_key.get(key)
            if not r or not r.value:
                continue
            # Only write if currently empty (don't overwrite existing data)
            existing = ws.cell(row=ind.row, column=vcol).value
            if existing is not None and str(existing).strip():
                continue
            prefix = "【估算】" if r.is_estimation else ""
            caliber_suffix = f"（口径：{r.caliber}）" if r.caliber else ""
            cell_v = ws.cell(row=ind.row, column=vcol, value=f"{prefix}{r.value}{caliber_suffix}")
            ws.cell(row=ind.row, column=scol, value=r.source_url)
            cell_v.alignment = Alignment(wrap_text=True, vertical="top")
            if r.is_estimation:
                cell_v.fill = est_fill
            elif r.confidence == "low":
                cell_v.fill = low_fill

    wb.save(V2_XLSX)
    log.info("saved %s", V2_XLSX.name)

def write_evidence_md(inds: list[Indicator], results_by_key: dict[str, CellResult]):
    lines = [f"# 一带一路 5 国中文教育指标证据日志", f"生成时间：{datetime.now():%Y-%m-%d %H:%M}", ""]
    for country, _, _, _, _ in COUNTRIES:
        lines.append(f"\n## {country}")
        for ind in inds:
            r = results_by_key.get(f"{country}::{ind.id}")
            if not r:
                continue
            lines.append(f"\n### #{ind.id} {ind.name} [{ind.priority}]")
            if r.error:
                lines.append(f"> ⚠️ {r.error}")
                continue
            flag = " 【估算】" if r.is_estimation else ""
            lines.append(f"- **值**{flag}：{r.value}")
            if r.caliber:
                lines.append(f"- **口径**：{r.caliber}")
            lines.append(f"- **置信度**：{r.confidence}")
            lines.append(f"- **首选出处**：{r.source_url}")
            if r.evidence:
                lines.append(f"- **证据链**：")
                for e in r.evidence[:6]:
                    t = (e.title or e.snippet[:60] or "").replace("\n", " ")
                    lines.append(f"  - ({e.source}) {t} — {e.url}")
    EVIDENCE_MD.write_text("\n".join(lines), encoding="utf-8")
    log.info("saved %s", EVIDENCE_MD.name)

def write_estimation_md(inds: list[Indicator], results_by_key: dict[str, CellResult]):
    lines = [f"# 估算类指标口径说明", f"生成时间：{datetime.now():%Y-%m-%d %H:%M}",
             "下列单元格标记为【估算】，非原始官方数据，仅作参考。", ""]
    total = 0
    for country, _, _, _, _ in COUNTRIES:
        rows = []
        for ind in inds:
            r = results_by_key.get(f"{country}::{ind.id}")
            if not r or not r.is_estimation:
                continue
            total += 1
            rows.append(f"- #{ind.id} **{ind.name}**（{ind.priority}）：{r.value}  \n  口径：{r.caliber or '未说明'}  \n  来源：{r.source_url}")
        if rows:
            lines.append(f"\n## {country}")
            lines.extend(rows)
    lines.insert(3, f"合计估算单元格：**{total}** 个")
    ESTIMATION_MD.write_text("\n".join(lines), encoding="utf-8")
    log.info("saved %s", ESTIMATION_MD.name)

# ── main ────────────────────────────────────────────────────────────────────
async def run(only_countries: list[str] | None, sample: int | None, concurrency: int):
    inds = load_indicators()
    log.info("loaded %d indicators", len(inds))

    gaps = find_gaps(inds, only_countries)
    if sample:
        keep_per_country: dict[str, int] = {}
        gaps2 = []
        for country, ind in gaps:
            if keep_per_country.get(country, 0) < sample:
                gaps2.append((country, ind))
                keep_per_country[country] = keep_per_country.get(country, 0) + 1
        gaps = gaps2
    log.info("gaps to fill: %d", len(gaps))

    # Load checkpoint
    checkpoint: dict = {}
    if CHECKPOINT.exists():
        try:
            checkpoint = json.loads(CHECKPOINT.read_text(encoding="utf-8"))
            log.info("checkpoint: %d cells cached", len(checkpoint))
        except Exception:
            checkpoint = {}

    # Clients: 3 with proxy, 1 direct (deepseek)
    ppx_proxy = get_proxies_for_url("https://api.perplexity.ai", OVERSEAS_PROXY)
    brave_proxy = get_proxies_for_url("https://api.search.brave.com", OVERSEAS_PROXY)
    tav_proxy = get_proxies_for_url("https://api.tavily.com", OVERSEAS_PROXY)

    ppx_client = build_httpx_client(ppx_proxy or "", timeout=60.0)
    brave_client = build_httpx_client(brave_proxy or "", timeout=30.0)
    tavily_client = build_httpx_client(tav_proxy or "", timeout=45.0)
    ds_client = build_httpx_client("", timeout=60.0)

    sem = asyncio.Semaphore(concurrency)

    async with ppx_client, brave_client, tavily_client, ds_client:
        t0 = time.monotonic()
        total = len(gaps)
        results: list[CellResult] = []

        async def _task(country, ind, idx):
            r = await fill_one_cell(sem, country, ind, ppx_client, brave_client,
                                     tavily_client, ds_client, checkpoint)
            dt = time.monotonic() - t0
            marker = "✓" if r.value else "✗"
            est = "[EST]" if r.is_estimation else ""
            log.info("%s [%d/%d] %s#%d %s %s %s (%.1fs)",
                     marker, idx + 1, total, country, ind.id, ind.name[:20], r.confidence, est, dt)
            return r

        tasks = [_task(c, i, idx) for idx, (c, i) in enumerate(gaps)]
        results = await asyncio.gather(*tasks)

    # Merge into dict
    results_by_key: dict[str, CellResult] = {}
    for r in results:
        results_by_key[f"{r.country}::{r.indicator_id}"] = r
    # Include cached-only keys
    for k, v in checkpoint.items():
        if k in results_by_key:
            continue
        country, iid = k.split("::")
        ind = next((i for i in inds if i.id == int(iid)), None)
        if not ind:
            continue
        r = CellResult(country=country, indicator_id=ind.id, indicator_name=ind.name,
                       value=v.get("value", ""), source_url=v.get("source_url", ""),
                       caliber=v.get("caliber", ""), confidence=v.get("confidence", "low"),
                       is_estimation=bool(v.get("is_estimation")), error=v.get("error", ""))
        r.evidence = [Evidence(**e) for e in v.get("evidence", [])]
        results_by_key[k] = r

    # Write outputs
    write_v2(inds, results_by_key)
    write_evidence_md(inds, results_by_key)
    write_estimation_md(inds, results_by_key)

    # Stats
    filled = sum(1 for r in results_by_key.values() if r.value and not r.error)
    est = sum(1 for r in results_by_key.values() if r.is_estimation)
    log.info("DONE: %d filled, %d estimations, %d total cached", filled, est, len(results_by_key))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sample", type=int, default=None, help="每国取前 N 个指标做试跑")
    ap.add_argument("--country", type=str, default=None, help="只跑指定国家（中文名，逗号分隔）")
    ap.add_argument("--concurrency", type=int, default=5, help="并发数")
    ap.add_argument("--reset", action="store_true", help="清空 checkpoint 重跑")
    args = ap.parse_args()

    if args.reset and CHECKPOINT.exists():
        CHECKPOINT.unlink()
        log.info("checkpoint cleared")

    only = [c.strip() for c in args.country.split(",")] if args.country else None

    asyncio.run(run(only, args.sample, args.concurrency))


if __name__ == "__main__":
    main()
