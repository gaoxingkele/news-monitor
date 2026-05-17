"""对 P1+P2 语义空格做代理指标推导 + 学术搜索。

流程:
1. DeepSeek 规划: 原指标 → 代理指标 + 推导路径 + 搜索关键词
2. 并发查询: academic_search(CNKI/Scholar/OpenAlex/CrossRef/Semantic) + Perplexity + Tavily
3. DeepSeek 合成: 从代理值反推原指标值 + 置信度 + 不确定性

输入: output/bri_p_layered_gaps.json (P1+P2 格)
产出:
  ..._五国填充版_v8_{date}.xlsx
  bri_proxy_derivation_{date}.md
"""
from __future__ import annotations
import asyncio, io, json, os, sys, logging, time
from dataclasses import asdict
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "tools"))
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError: pass

import httpx, openpyxl
from openpyxl.styles import PatternFill, Alignment
from news_monitor.proxy import build_httpx_client, get_proxies_for_url
from academic_search import academic_search, format_results

TOPICS = ROOT / "topics"
DATE = datetime.now().strftime("%Y%m%d")
SRC = TOPICS / "“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v9_20260421.xlsx"
V8 = TOPICS / f"“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v10_{DATE}.xlsx"
LOG_MD = TOPICS / f"bri_proxy_derivation_{DATE}.md"
CKPT = ROOT / "output" / "bri_proxy_derivation_ckpt.json"
GAPS_FILE = ROOT / "output" / "bri_p_layered_gaps.json"

PPX = os.environ.get("PERPLEXITY_API_KEY", "")
TAV = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
DS = os.environ.get("DEEPSEEK_API_KEY", "")
PROXY = os.environ.get("LLM_PROXY", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s", datefmt="%H:%M:%S",
                    handlers=[logging.FileHandler(ROOT/'output'/f'bri_proxy_{DATE}.log', encoding='utf-8'),
                              logging.StreamHandler()])
log = logging.getLogger("proxy")

COL = {"蒙古":(5,6), "越南":(7,8), "柬埔寨":(9,10), "菲律宾":(11,12), "新加坡":(13,14)}
COUNTRY_EN = {"蒙古":"Mongolia", "越南":"Vietnam", "柬埔寨":"Cambodia",
              "菲律宾":"Philippines", "新加坡":"Singapore"}

# ── Step 1: LLM plans proxy strategy ────────────────────────────────────────
async def plan_proxy(client: httpx.AsyncClient, gap: dict) -> dict:
    """让 DS 规划代理指标 + 搜索查询."""
    prompt = f"""你是研究方法论专家。为下列"搜不到直接数据"的指标设计代理推导方案。

## 目标指标
- 国家：{gap['country']}
- 字段名：{gap['name']}
- 取值口径：{gap['type']}
- 字段说明：{gap['desc']}
- 优先级：P{gap.get('priority','?')}

## 任务
请给出 2-3 个合理的代理推导方案，优先级从强到弱。代理应该：
1. 代理指标本身有公开可查的数据（官方统计、学术研究、汉办/孔院数据、CNKI 期刊等）
2. 与目标指标有明确的可计算映射关系（比值、累加、区域类比）
3. 映射的假设前提明确可论证

返回 JSON（不要其他文本）:
{{
  "proxy_plans": [
    {{
      "rank": 1,
      "proxy_name": "代理指标名（如：孔子学院/课堂 学员数占比）",
      "derivation_logic": "如何从代理值推算目标值，30字内",
      "queries": {{
        "academic": "学术数据库查询词（中英文混合）",
        "perplexity": "问 Perplexity 的具体问题",
        "tavily_zh": "中文 Tavily 查询",
        "tavily_en": "英文 Tavily 查询"
      }},
      "priority_domains": ["优先搜索的域名", "..."],
      "confidence_ceiling": "high/medium/low"
    }},
    {{"rank": 2, ...}}
  ],
  "fallback": "如果上述代理都无数据，给出一个兜底的定性回答（如'有限 / 无 / 有'），并说明推理依据"
}}
"""
    try:
        resp = await client.post(
            "https://api.deepseek.com/v1/chat/completions",
            json={"model":"deepseek-chat","messages":[{"role":"user","content":prompt}],
                  "temperature":0.1,"response_format":{"type":"json_object"}},
            headers={"Authorization":f"Bearer {DS}","Content-Type":"application/json"},
            timeout=90.0,
        )
        resp.raise_for_status()
        return json.loads(resp.json()["choices"][0]["message"]["content"])
    except Exception as e:
        log.warning("plan err %s#%d: %s", gap['country'], gap['id'], str(e)[:100])
        return {"proxy_plans": [], "fallback": "无可推导方案"}

# ── Step 2: Execute queries in parallel ─────────────────────────────────────
async def ppx_ask(client, q):
    if not PPX: return "", []
    try:
        r = await client.post(
            "https://api.perplexity.ai/chat/completions",
            json={"model":"sonar","messages":[
                {"role":"system","content":"你是代理研究员。当直接数据不可得时，用代理指标推导原指标值。给出具体数字+来源URL。"},
                {"role":"user","content":q}],
                "temperature":0.1,"return_citations":True},
            headers={"Authorization":f"Bearer {PPX}","Content-Type":"application/json"},
            timeout=90.0,
        )
        r.raise_for_status()
        d = r.json()
        return d["choices"][0]["message"]["content"], d.get("citations",[])[:8]
    except Exception as e:
        return "", []

async def tav(client, q, domains=None):
    if not TAV: return []
    try:
        body = {"api_key":TAV,"query":q,"search_depth":"advanced",
                "max_results":8,"include_answer":False,"include_raw_content":False}
        if domains: body["include_domains"] = domains[:15]
        r = await client.post("https://api.tavily.com/search", json=body, timeout=60.0)
        r.raise_for_status()
        return [{"url":x.get("url",""),"title":x.get("title",""),
                 "snippet":(x.get("content") or "")[:400]}
                for x in (r.json().get("results") or [])[:8]]
    except Exception as e:
        return []

# ── Step 3: DS synthesize final value ───────────────────────────────────────
async def synthesize(client, gap, plan, evidence: dict) -> dict:
    """根据代理计划 + 证据合成最终值."""
    plans_json = json.dumps(plan.get("proxy_plans", [])[:3], ensure_ascii=False, indent=1)
    # 组装证据
    ppx_blocks = []
    for name, (ans, cites) in evidence.get("ppx", {}).items():
        ppx_blocks.append(f"### Perplexity [{name}]\n{ans}\nCitations:\n" + "\n".join(cites))
    tav_blocks = []
    for name, items in evidence.get("tav", {}).items():
        sub = []
        for x in items[:5]:
            sub.append(f"  - {x['title']} | {x['url']}\n    {x['snippet'][:200]}")
        tav_blocks.append(f"### Tavily [{name}]\n" + "\n".join(sub))
    acad_block = evidence.get("academic_formatted", "")

    prompt = f"""你是研究员。请根据代理推导计划 + 所有证据，给出【{gap['country']}】【{gap['name']}】的最终值。

## 目标指标
- 字段：{gap['name']}
- 取值口径：{gap['type']}
- 字段说明：{gap['desc']}

## 代理推导方案（从 Step1 得出）
{plans_json}

## 证据池
{chr(10).join(ppx_blocks)}

{chr(10).join(tav_blocks)}

## 学术数据库结果
{acad_block[:8000]}

## 任务
选用最强可行的代理方案，给出目标指标的数值答案（带代理推导过程）。

返回 JSON:
{{
  "value": "最终回答（贴合取值口径）",
  "proxy_used": "实际使用的代理指标名",
  "derivation": "完整推导过程：代理值 X → 映射规则 Y → 目标值 Z",
  "source_urls": ["证据URL1", "URL2"],
  "confidence": "high/medium/low",
  "is_proxy_derivation": true,
  "rationale": "30字内说明"
}}

若所有代理都无可用证据，value 填 plan.fallback 的定性答案（如"有限"/"无"），confidence=low。
不要给出 "无可核验数据" —— 一定要有一个答案，哪怕是定性的。
"""
    try:
        r = await client.post(
            "https://api.deepseek.com/v1/chat/completions",
            json={"model":"deepseek-chat","messages":[{"role":"user","content":prompt}],
                  "temperature":0.05,"response_format":{"type":"json_object"}},
            headers={"Authorization":f"Bearer {DS}","Content-Type":"application/json"},
            timeout=120.0,
        )
        r.raise_for_status()
        return json.loads(r.json()["choices"][0]["message"]["content"])
    except Exception as e:
        log.warning("synth err %s#%d: %s", gap['country'], gap['id'], str(e)[:100])
        return {"value":"合成失败","confidence":"low","is_proxy_derivation":True}

# ── Main per-gap pipeline ───────────────────────────────────────────────────
async def fill_gap(sem, gap, clients, ckpt) -> dict:
    key = f"{gap['country']}::{gap['id']}"
    if key in ckpt:
        return ckpt[key]
    async with sem:
        t0 = time.time()
        ppx_c, tav_c, ds_c = clients

        # Step 1: plan
        plan = await plan_proxy(ds_c, gap)
        plans = plan.get("proxy_plans", [])[:2]  # use top 2
        if not plans:
            ckpt[key] = {"value":plan.get("fallback","不可推导"),"confidence":"low",
                        "is_proxy_derivation":True,"rationale":"无方案"}
            _save_ckpt(ckpt)
            return ckpt[key]

        # Step 2: run queries for each plan in parallel
        ppx_tasks = {}
        tav_tasks = {}
        acad_tasks = {}
        for pp in plans:
            pname = pp.get("proxy_name","?")[:30]
            qs = pp.get("queries", {})
            if qs.get("perplexity"):
                ppx_tasks[pname] = ppx_ask(ppx_c, qs["perplexity"])
            if qs.get("tavily_zh"):
                doms = pp.get("priority_domains") or None
                tav_tasks[pname+"_zh"] = tav(tav_c, qs["tavily_zh"], doms)
            if qs.get("tavily_en"):
                tav_tasks[pname+"_en"] = tav(tav_c, qs["tavily_en"], None)
            if qs.get("academic"):
                acad_tasks[pname] = academic_search(qs["academic"], max_per_source=3,
                                                    proxy_url=PROXY, include_scholar=True)

        ppx_results = dict(zip(ppx_tasks.keys(),
                              await asyncio.gather(*ppx_tasks.values(), return_exceptions=True)))
        tav_results = dict(zip(tav_tasks.keys(),
                              await asyncio.gather(*tav_tasks.values(), return_exceptions=True)))
        acad_results = dict(zip(acad_tasks.keys(),
                               await asyncio.gather(*acad_tasks.values(), return_exceptions=True)))

        # Format academic results
        acad_fmt = ""
        for pname, r in acad_results.items():
            if isinstance(r, dict):
                acad_fmt += f"\n\n== 代理 [{pname}] 学术 ==\n" + format_results(r)[:4000]

        evidence = {
            "ppx": {k: v for k, v in ppx_results.items() if isinstance(v, tuple)},
            "tav": {k: v for k, v in tav_results.items() if isinstance(v, list)},
            "academic_formatted": acad_fmt,
        }

        # Step 3: synthesize
        result = await synthesize(ds_c, gap, plan, evidence)
        result["plan_used"] = plans[0] if plans else {}
        ckpt[key] = result
        _save_ckpt(ckpt)
        log.info("%s#%d %.1fs value=%s conf=%s", gap['country'], gap['id'],
                 time.time()-t0, str(result.get("value",""))[:50], result.get("confidence","?"))
        return result

def _save_ckpt(ckpt):
    with open(CKPT, "w", encoding="utf-8") as f:
        json.dump(ckpt, f, ensure_ascii=False, indent=1)

# ── Main ────────────────────────────────────────────────────────────────────
async def run(only_priority=None, limit=None):
    with open(GAPS_FILE, encoding='utf-8') as f:
        gaps_all = json.load(f)
    # Flatten P1 + P2 (optionally P3 if user wants)
    targets = []
    for p in ['P1','P2'] if only_priority is None else [only_priority]:
        for g in gaps_all.get(p, []):
            g['priority'] = p[1]
            targets.append(g)
    if limit: targets = targets[:limit]
    log.info("填 %d 格 (P1=%d, P2=%d)",
             len(targets), sum(1 for t in targets if t['priority']=='1'),
             sum(1 for t in targets if t['priority']=='2'))

    ckpt = {}
    if CKPT.exists():
        try: ckpt = json.loads(CKPT.read_text(encoding='utf-8'))
        except Exception: pass
    log.info("ckpt: %d 已缓存", len(ckpt))

    ppx_p = get_proxies_for_url("https://api.perplexity.ai", PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", PROXY)

    async with (build_httpx_client(ppx_p or "", 90) as ppx_c,
                build_httpx_client(tav_p or "", 60) as tav_c,
                build_httpx_client("", 120) as ds_c):
        sem = asyncio.Semaphore(4)
        clients = (ppx_c, tav_c, ds_c)
        tasks = [fill_gap(sem, g, clients, ckpt) for g in targets]
        results = await asyncio.gather(*tasks)

    # Write v8 xlsx
    wb = openpyxl.load_workbook(SRC); ws = wb.active
    proxy_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    log_md = [f"# BRI 5 国代理推导填报 (v8)", f"生成：{datetime.now():%Y-%m-%d %H:%M}",
              f"对 P1+P2 优先级共 {len(targets)} 个语义空格做代理指标推导。\n"]

    for g, r in zip(targets, results):
        val = str(r.get("value","")).strip()
        if not val or val == "合成失败": continue
        proxy = r.get("proxy_used","")
        deriv = r.get("derivation","")
        row = g['row']
        vcol = g['vcol']
        prefix = "【代理推导】"
        txt = f"{prefix}{val}"
        if proxy: txt += f"（代理：{proxy}）"
        if deriv: txt += f"\n推导：{deriv}"
        ws.cell(row, vcol, txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, vcol).fill = proxy_fill
        urls = r.get("source_urls",[])
        if urls:
            ws.cell(row, vcol+1, "; ".join(urls[:3]))

        log_md.append(f"\n## {g['country']}#{g['id']} {g['name']} (P{g['priority']})")
        log_md.append(f"- **值**：{val}")
        log_md.append(f"- **代理**：{proxy}")
        log_md.append(f"- **推导**：{deriv}")
        log_md.append(f"- **置信度**：{r.get('confidence','?')}")
        log_md.append(f"- **理由**：{r.get('rationale','')}")
        if urls:
            log_md.append(f"- **来源**：")
            for u in urls[:5]: log_md.append(f"  - {u}")

    wb.save(V8)
    LOG_MD.write_text("\n".join(log_md), encoding="utf-8")
    log.info("saved %s", V8.name)
    log.info("saved %s", LOG_MD.name)

    # Stats
    filled = sum(1 for r in results if r.get("value") and r.get("value") not in ("合成失败","不可推导"))
    log.info("DONE: %d/%d 成功代理推导", filled, len(targets))


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--priority", default=None, help="只跑 P1 或 P2")
    ap.add_argument("--limit", type=int, default=None)
    args = ap.parse_args()
    asyncio.run(run(args.priority, args.limit))
