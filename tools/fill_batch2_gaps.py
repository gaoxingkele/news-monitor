"""批次 2 初填: 马来西亚/泰国/印尼/老挝/缅甸 5 国 × 69 指标.

基于 fill_bri_5country_gaps.py 改造，只修改国家配置。
"""
from __future__ import annotations
import argparse, asyncio, io, json, logging, os, sys, time
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))

try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError: pass

import httpx, openpyxl
from openpyxl.styles import Alignment, PatternFill
from news_monitor.proxy import build_httpx_client, get_proxies_for_url

TOPICS = ROOT / "topics"
OUTPUT_DIR = ROOT / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
SRC_XLSX = TOPICS / "“一带一路”沿线国家中文教育发展指标体系数据总表_批次2_五国_20260422.xlsx"
FIELD_DICT = TOPICS / "bri_field_dict.json"
CHECKPOINT = OUTPUT_DIR / "bri_batch2_gap_fill_checkpoint.json"

DATE_TAG = datetime.now().strftime("%Y%m%d")
V2_XLSX = TOPICS / f"“一带一路”沿线国家中文教育发展指标体系数据总表_批次2_五国_v2_{DATE_TAG}.xlsx"
EVIDENCE_MD = TOPICS / f"bri_batch2_evidence_log_{DATE_TAG}.md"
ESTIMATION_MD = TOPICS / f"bri_batch2_estimation_notes_{DATE_TAG}.md"

# ── NEW 5 countries: Malaysia/Thailand/Indonesia/Laos/Myanmar ────────────────
COUNTRIES = [
    ("马来西亚", 5, 6,
     ["moe.gov.my", "kln.gov.my", "gov.my", "um.edu.my", "utar.edu.my",
      "sin.chineseembassy.gov.cn", "kedutaanchina.org.my", "dongzong.my",
      "jiaozong.org.my"], "ms"),
    ("泰国", 7, 8,
     ["moe.go.th", "mfa.go.th", "go.th", "chula.ac.th", "tu.ac.th",
      "bangkok.cn", "hanban.org", "ci.cn"], "th"),
    ("印尼", 9, 10,
     ["kemdikbud.go.id", "kemlu.go.id", "go.id", "ui.ac.id", "ugm.ac.id",
      "binus.edu", "petrachristian.ac.id"], "id"),
    ("老挝", 11, 12,
     ["moes.edu.la", "mofa.gov.la", "nuol.edu.la", "gov.la",
      "kpl.gov.la"], "lo"),
    ("缅甸", 13, 14,
     ["moe.gov.mm", "mofa.gov.mm", "gov.mm", "ucsy.edu.mm",
      "uym.edu.mm"], "my"),
]
COUNTRY_MAP = {c[0]: c for c in COUNTRIES}
CHINA_SIDE_DOMAINS = [
    "hanban.org", "chinese.cn", "ci.cn", "ccn.edu.cn", "confucius.edu.cn",
    "mfa.gov.cn", "fmprc.gov.cn", "moe.gov.cn", "edu.cn",
    "xinhuanet.com", "chinanews.com", "people.com.cn",
]

PERPLEXITY_API_KEY = os.environ.get("PERPLEXITY_API_KEY", "")
BRAVE_API_KEY = os.environ.get("BRAVEAPI", "") or os.environ.get("BRAVE_API_KEY", "")
TAVILY_API_KEY = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
OVERSEAS_PROXY = os.environ.get("LLM_PROXY", "") or os.environ.get("OVERSEAS_PROXY", "")

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger("batch2")

@dataclass
class Indicator:
    id: int; row: int; name: str; ftype: str; category: str; priority: str
    desc: str; hint_cn: str; hint_en: str

@dataclass
class Evidence:
    source: str; url: str = ""; title: str = ""; snippet: str = ""

@dataclass
class CellResult:
    country: str; indicator_id: int; indicator_name: str
    value: str = ""; source_url: str = ""; caliber: str = ""
    confidence: str = "low"; is_estimation: bool = False
    evidence: list = field(default_factory=list); error: str = ""

async def ask_perplexity(client, question):
    if not PERPLEXITY_API_KEY: return "", []
    body = {"model":"sonar","messages":[
        {"role":"system","content":
         "You are a meticulous research assistant. Answer the user's factual question about education/policy "
         "with a concise answer and cite every claim. Prefer official government and authoritative sources. "
         "If the fact is uncertain or unavailable, say so explicitly."},
        {"role":"user","content":question}],
        "temperature":0.1,"return_citations":True}
    try:
        r = await client.post("https://api.perplexity.ai/chat/completions", json=body,
            headers={"Authorization":f"Bearer {PERPLEXITY_API_KEY}","Content-Type":"application/json"},
            timeout=60.0)
        r.raise_for_status()
        d = r.json()
        return d["choices"][0]["message"]["content"], (d.get("citations",[]) or [])[:8]
    except Exception as e:
        log.warning("ppx: %s", str(e)[:120]); return "", []

async def search_brave(client, query, cc=""):
    if not BRAVE_API_KEY: return []
    params = {"q":query,"count":10,"text_decorations":False}
    if cc and cc in {"AR","AU","BR","CA","CN","DE","FR","GB","HK","IN","ID","IT","JP","KR","MX","MY","NL","PH","RU","SG","TW","US","TH","NO"}:
        params["country"] = cc
    try:
        r = await client.get("https://api.search.brave.com/res/v1/web/search", params=params,
            headers={"Accept":"application/json","X-Subscription-Token":BRAVE_API_KEY}, timeout=30.0)
        r.raise_for_status()
        d = r.json()
        return [Evidence(source="brave", url=x.get("url",""), title=x.get("title",""),
                         snippet=(x.get("description") or "")[:300])
                for x in (d.get("web",{}).get("results") or [])[:10]]
    except Exception as e:
        log.warning("brave: %s", str(e)[:120]); return []

async def search_tavily(client, query, domains=None):
    if not TAVILY_API_KEY: return []
    body = {"api_key":TAVILY_API_KEY,"query":query,"search_depth":"advanced",
            "max_results":8,"include_answer":False,"include_raw_content":False}
    if domains: body["include_domains"] = domains[:20]
    try:
        r = await client.post("https://api.tavily.com/search", json=body, timeout=45.0)
        r.raise_for_status()
        d = r.json()
        return [Evidence(source="tavily", url=x.get("url",""), title=x.get("title",""),
                         snippet=(x.get("content") or "")[:400])
                for x in (d.get("results") or [])[:8]]
    except Exception as e:
        log.warning("tav: %s", str(e)[:120]); return []

async def llm_extract(client, ind, country, ppx_answer, web):
    if not DEEPSEEK_API_KEY:
        return {"value":(ppx_answer or "").strip()[:500],"source_url":"","caliber":"","confidence":"low","is_estimation":False}
    evid = "\n".join(f"[{i+1}] ({e.source}) {e.title}\nURL: {e.url}\n摘要: {e.snippet}"
                    for i,e in enumerate(web[:10]))
    prompt = f"""你是专门抽取事实的助手。请从下面的证据中为【{country}】在指标【{ind.name}】上给出回答。

## 指标
- 字段类型：{ind.ftype}
- 说明：{ind.desc}
- 优先级：{ind.priority}  分类：{ind.category}

## Perplexity
{ppx_answer or '(空)'}

## 网页
{evid or '(无)'}

返回 JSON:
- value: 回答（贴合字段类型）
- source_url: 首选出处
- caliber: 口径（年份/分子分母/统计范围）
- confidence: high|medium|low
- is_estimation: 布尔
- rationale: 30字
"""
    try:
        r = await client.post("https://api.deepseek.com/v1/chat/completions",
            json={"model":"deepseek-chat","messages":[{"role":"user","content":prompt}],
                  "temperature":0.1,"response_format":{"type":"json_object"}},
            headers={"Authorization":f"Bearer {DEEPSEEK_API_KEY}","Content-Type":"application/json"},
            timeout=60.0)
        r.raise_for_status()
        return json.loads(r.json()["choices"][0]["message"]["content"])
    except Exception as e:
        log.warning("ds %s#%d: %s", country, ind.id, str(e)[:120])
        return {"value":(ppx_answer or "")[:300],"source_url":web[0].url if web else "",
                "caliber":"","confidence":"low","is_estimation":False}

def build_ppx_q(country, ind):
    return (f"请提供关于【{country}】在指标【{ind.name}】上的最权威信息。\n"
            f"- 取值口径：{ind.ftype}\n- 说明：{ind.desc}\n"
            f"- 优先来源：官方政府/教育部/驻外使馆/双边声明/权威研究\n\n"
            f"给出：(1) 事实/数据；(2) 官方来源 URL；(3) 不同口径/年份差异。中文。")

def build_web_q(country, ind, lang="cn"):
    if lang == "en":
        return ind.hint_en.replace("国家名", country) + f" Chinese language education {country}"
    return ind.hint_cn.replace("国家名", country)

async def fill_one_cell(sem, country, ind, clients, ckpt):
    key = f"{country}::{ind.id}"
    if key in ckpt:
        cached = ckpt[key]
        r = CellResult(country=country, indicator_id=ind.id, indicator_name=ind.name,
                       value=cached.get("value",""), source_url=cached.get("source_url",""),
                       caliber=cached.get("caliber",""), confidence=cached.get("confidence","low"),
                       is_estimation=bool(cached.get("is_estimation")), error=cached.get("error",""))
        r.evidence = [Evidence(**e) for e in cached.get("evidence",[])]
        return r

    async with sem:
        ppx_c, brave_c, tavily_c, ds_c = clients
        result = CellResult(country=country, indicator_id=ind.id, indicator_name=ind.name)
        country_meta = COUNTRY_MAP[country]
        official = country_meta[3] + CHINA_SIDE_DOMAINS
        cc = {"马来西亚":"MY","泰国":"TH","印尼":"ID","老挝":"","缅甸":""}[country]

        ppx_a, ppx_c_urls = await ask_perplexity(ppx_c, build_ppx_q(country, ind))
        for u in ppx_c_urls:
            result.evidence.append(Evidence(source="perplexity", url=u))

        q_cn = build_web_q(country, ind, "cn")
        q_en = build_web_q(country, ind, "en")
        br_r, tav_o, tav_g = await asyncio.gather(
            search_brave(brave_c, q_en, cc),
            search_tavily(tavily_c, q_en, official),
            search_tavily(tavily_c, q_cn, None),
            return_exceptions=True,
        )
        for x in (br_r, tav_o, tav_g):
            if isinstance(x, list): result.evidence.extend(x[:5])

        if not ppx_a and not any(e.url for e in result.evidence):
            result.error = "no evidence"
            _save(ckpt, key, result); return result

        ext = await llm_extract(ds_c, ind, country, ppx_a, result.evidence)
        result.value = ext["value"]
        result.source_url = ext["source_url"] or (result.evidence[0].url if result.evidence else "")
        result.caliber = ext["caliber"]
        result.confidence = ext["confidence"]
        result.is_estimation = ext.get("is_estimation", False)
        _save(ckpt, key, result)
        return result

def _save(ckpt, key, r):
    ckpt[key] = {"value":r.value,"source_url":r.source_url,"caliber":r.caliber,
                 "confidence":r.confidence,"is_estimation":r.is_estimation,
                 "evidence":[asdict(e) for e in r.evidence[:12]],"error":r.error}
    with open(CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump(ckpt, f, ensure_ascii=False, indent=1)

def load_indicators():
    with open(FIELD_DICT, encoding='utf-8') as f:
        raw = json.load(f)
    inds = []
    for i, x in enumerate(raw):
        if x.get("指标ID") is None: continue
        inds.append(Indicator(
            id=int(x["指标ID"]), row=i+2,
            name=str(x.get("字段名","") or ""),
            ftype=str(x.get("字段类型","") or ""),
            category=str(x.get("分类","") or ""),
            priority=str(x.get("优先级","") or ""),
            desc=str(x.get("字段说明","") or ""),
            hint_cn=str(x.get("中文检索提示","") or ""),
            hint_en=str(x.get("英文检索提示","") or ""),
        ))
    return inds

def find_gaps(inds, only=None):
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True); ws = wb.active
    gaps = []
    for country, vcol, _, _, _ in COUNTRIES:
        if only and country not in only: continue
        for ind in inds:
            v = ws.cell(row=ind.row, column=vcol).value
            if v is None or (isinstance(v,str) and not v.strip()):
                gaps.append((country, ind))
    return gaps

def write_v2(inds, results_by_key):
    wb = openpyxl.load_workbook(SRC_XLSX); ws = wb.active
    est_fill = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")
    low_fill = PatternFill(start_color="FFEAEA", end_color="FFEAEA", fill_type="solid")
    for country, vcol, scol, _, _ in COUNTRIES:
        for ind in inds:
            r = results_by_key.get(f"{country}::{ind.id}")
            if not r or not r.value: continue
            existing = ws.cell(row=ind.row, column=vcol).value
            if existing is not None and str(existing).strip(): continue
            prefix = "【估算】" if r.is_estimation else ""
            suffix = f"（口径：{r.caliber}）" if r.caliber else ""
            c = ws.cell(row=ind.row, column=vcol, value=f"{prefix}{r.value}{suffix}")
            ws.cell(row=ind.row, column=scol, value=r.source_url)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if r.is_estimation: c.fill = est_fill
            elif r.confidence == "low": c.fill = low_fill
    wb.save(V2_XLSX)
    log.info("saved %s", V2_XLSX.name)

def write_evidence_md(inds, results_by_key):
    lines = [f"# 批次 2 - 5 国证据日志", f"时间：{datetime.now():%Y-%m-%d %H:%M}", ""]
    for country, _, _, _, _ in COUNTRIES:
        lines.append(f"\n## {country}")
        for ind in inds:
            r = results_by_key.get(f"{country}::{ind.id}")
            if not r: continue
            lines.append(f"\n### #{ind.id} {ind.name} [{ind.priority}]")
            if r.error: lines.append(f"> ⚠️ {r.error}"); continue
            flag = " 【估算】" if r.is_estimation else ""
            lines.append(f"- **值**{flag}：{r.value}")
            if r.caliber: lines.append(f"- **口径**：{r.caliber}")
            lines.append(f"- **置信度**：{r.confidence}")
            lines.append(f"- **首选出处**：{r.source_url}")
            if r.evidence:
                lines.append(f"- **证据链**：")
                for e in r.evidence[:6]:
                    t = (e.title or e.snippet[:60] or "").replace("\n"," ")
                    lines.append(f"  - ({e.source}) {t} — {e.url}")
    EVIDENCE_MD.write_text("\n".join(lines), encoding="utf-8")
    log.info("saved %s", EVIDENCE_MD.name)

def write_estimation_md(inds, results_by_key):
    lines = [f"# 批次 2 估算口径说明", f"时间：{datetime.now():%Y-%m-%d %H:%M}", ""]
    total = 0
    for country, _, _, _, _ in COUNTRIES:
        rows = []
        for ind in inds:
            r = results_by_key.get(f"{country}::{ind.id}")
            if not r or not r.is_estimation: continue
            total += 1
            rows.append(f"- #{ind.id} **{ind.name}** ({ind.priority})：{r.value}  \n  口径：{r.caliber or '未说明'}  \n  来源：{r.source_url}")
        if rows:
            lines.append(f"\n## {country}"); lines.extend(rows)
    lines.insert(2, f"合计估算：**{total}** 格\n")
    ESTIMATION_MD.write_text("\n".join(lines), encoding="utf-8")
    log.info("saved %s", ESTIMATION_MD.name)

async def run(only=None, sample=None, conc=5):
    inds = load_indicators()
    log.info("loaded %d indicators", len(inds))
    gaps = find_gaps(inds, only)
    if sample:
        keep = {}; g2 = []
        for c, i in gaps:
            if keep.get(c,0) < sample:
                g2.append((c,i)); keep[c]=keep.get(c,0)+1
        gaps = g2
    log.info("gaps: %d", len(gaps))

    ckpt = {}
    if CHECKPOINT.exists():
        try: ckpt = json.loads(CHECKPOINT.read_text(encoding="utf-8"))
        except Exception: pass
    log.info("ckpt: %d", len(ckpt))

    ppx_p = get_proxies_for_url("https://api.perplexity.ai", OVERSEAS_PROXY)
    brv_p = get_proxies_for_url("https://api.search.brave.com", OVERSEAS_PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", OVERSEAS_PROXY)

    ppx_c = build_httpx_client(ppx_p or "", 60.0)
    brv_c = build_httpx_client(brv_p or "", 30.0)
    tav_c = build_httpx_client(tav_p or "", 45.0)
    ds_c = build_httpx_client("", 60.0)

    sem = asyncio.Semaphore(conc)
    async with ppx_c, brv_c, tav_c, ds_c:
        t0 = time.monotonic()
        async def _task(country, ind, idx, total):
            r = await fill_one_cell(sem, country, ind, (ppx_c,brv_c,tav_c,ds_c), ckpt)
            dt = time.monotonic() - t0
            m = "✓" if r.value else "✗"
            e = "[EST]" if r.is_estimation else ""
            log.info("%s [%d/%d] %s#%d %s %s %s (%.1fs)",
                     m, idx+1, total, country, ind.id, r.indicator_name[:20],
                     r.confidence, e, dt)
            return r
        total = len(gaps)
        tasks = [_task(c, i, idx, total) for idx,(c,i) in enumerate(gaps)]
        results = await asyncio.gather(*tasks)

    by_key = {}
    for r in results: by_key[f"{r.country}::{r.indicator_id}"] = r
    for k, v in ckpt.items():
        if k in by_key: continue
        country, iid = k.split("::")
        ind = next((i for i in inds if i.id == int(iid)), None)
        if not ind: continue
        r = CellResult(country=country, indicator_id=ind.id, indicator_name=ind.name,
                       value=v.get("value",""), source_url=v.get("source_url",""),
                       caliber=v.get("caliber",""), confidence=v.get("confidence","low"),
                       is_estimation=bool(v.get("is_estimation")), error=v.get("error",""))
        r.evidence = [Evidence(**e) for e in v.get("evidence",[])]
        by_key[k] = r

    write_v2(inds, by_key)
    write_evidence_md(inds, by_key)
    write_estimation_md(inds, by_key)
    filled = sum(1 for r in by_key.values() if r.value and not r.error)
    est = sum(1 for r in by_key.values() if r.is_estimation)
    log.info("DONE: %d filled, %d est, %d total", filled, est, len(by_key))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sample", type=int, default=None)
    ap.add_argument("--country", type=str, default=None)
    ap.add_argument("--concurrency", type=int, default=5)
    ap.add_argument("--reset", action="store_true")
    args = ap.parse_args()
    if args.reset and CHECKPOINT.exists():
        CHECKPOINT.unlink(); log.info("ckpt cleared")
    only = [c.strip() for c in args.country.split(",")] if args.country else None
    asyncio.run(run(only, args.sample, args.concurrency))

if __name__ == "__main__":
    main()
