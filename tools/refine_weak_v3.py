"""V3 轮精修：对 241 低质量格做扩展口径重查，提升命中率。

扩展策略：
  - #9/#5   双边合作文件/MOU/一带一路框架文件也算
  - #18/#37 列出该国在中国留学获中文/汉学学位的返华学者（即便不担任编委/非博硕点）
  - #20     参与 HSK / CIEF / Hanban 标准制定也算
  - #27     国际通用 APP 的本地化版/使用版也算
  - #45/#46 含鲁班工坊、海丝学院、中文工坊等广义中外合作机构
  - #49     民间培训班、华人社团开办学校算偏远教学点
  - #59     国际场合中文使用、该国代表中文发言
  - 默认    近 5 年时间窗 + 代理变量 + 间接证据
  - 对于小国，额外说明"是否有学生在华攻读中文教育博硕士"

输入：output/bri_current_weak_targets_{DATE}.json（weak_cells 数组）
输出：xlsx 写回 *_refined447_{DATE}.xlsx 原文件（添加 【精修V3】 前缀）
      ckpt: output/bri_refine_v3_ckpt.json
"""
from __future__ import annotations
import argparse, asyncio, io, json, logging, os, re, shutil, sys, time
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "tools"))
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError: pass

import httpx, openpyxl
from openpyxl.styles import Alignment, PatternFill
from news_monitor.proxy import build_httpx_client, get_proxies_for_url
from academic_search import academic_search, format_results

TOPICS = ROOT / "topics"
OUTPUT = ROOT / "output"
DATE = datetime.now().strftime("%Y%m%d")
TARGETS = OUTPUT / f"bri_current_weak_targets_{DATE}.json"
CKPT = OUTPUT / f"bri_refine_v3_ckpt.json"
RUN_LOG = OUTPUT / f"bri_refine_v3_run.log"

PPX = os.environ.get("PERPLEXITY_API_KEY", "")
TAV = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
DS = os.environ.get("DEEPSEEK_API_KEY", "")
PROXY = os.environ.get("LLM_PROXY", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s",
                    datefmt="%H:%M:%S",
                    handlers=[logging.FileHandler(RUN_LOG, encoding='utf-8'),
                              logging.StreamHandler()])
log = logging.getLogger("v3")

# 复用 refine_421_cells.py 的国家元数据
from refine_421_cells import COUNTRY_META, CHINA_SIDE, ppx_ask, tav, ds_call, _persist_ckpt


# ─── 指标级扩展口径 ───────────────────────────────────────────────────────
EXPANSION_HINTS = {
    4: """【扩展口径·政府主导中文教育项目】
  - 教育部主办的师资培训班、奖学金、学生交流项目
  - 双边政府协议框架下的中文教学项目
  - 该国驻华使馆文化处/教育处组织的项目
  - 任何有政府资助或政府部门挂名的中文活动""",

    5: """【扩展口径·中文教育政策文件】
  - 中国与该国签署的双边教育合作备忘录、文化交流协定、一带一路人文合作文件
  - 该国教育部发布的外语教育政策中涉及中文的条款
  - Hanban/国际中文教育基金会与该国机构的合作协议
  - 该国政府承认的中文教学机构的授权文件
  - 中国驻外使馆发布的中文教育声明、公报""",

    9: """【扩展口径·中文教育法律法规】
  - 不局限于专门的"中文教育法"，凡是教育法/外语教育法/国民教育法/文化交流协定中提及"中文/汉语/普通话/孔子学院"的条款均算
  - 双边 MOU、教育合作协议中的中文条款
  - 该国教育部课程纲要中包含中文的部分
  - 该国政府关于一带一路 / 中外文化交流的法律文件""",

    10: """【扩展口径·教育政策环境风险】
  - 不只是直接"中文教育风险"，可参考：该国对中国机构的态度、对孔子学院的立法、对外来文化的限制
  - 该国教育部对私立外语学校的监管变化
  - 近 3 年该国媒体报道的中文教学争议事件""",

    12: """【扩展口径·国别研究专著数】
  - 不仅限中国学者出版的专著，该国学者撰写的中国研究专著也算
  - 涉及该国与中国文化/教育交流的专著
  - 学术系列丛书中的该国专题卷""",

    14: """【扩展口径·中文教育论文质量】
  - 基于近 5 年该国学者发表的中文教育论文的期刊等级加权
  - 考虑：SSCI/CSSCI/Scopus 收录情况、引用数
  - 若数据不足，用该国学者在中国举办的国际中文教育会议发表数量作代理""",

    15: """【扩展口径·学术会议数】
  - 该国主办或承办的中文教育相关国际会议
  - 该国学者参与的中国中文教育国际会议（如国际汉语教学学会 ICLTE）
  - 一带一路学术论坛中涉及中文教育的分会""",

    18: """【扩展口径·当地学者参与中文期刊编委】※ 小国重点
  关键扩展：
  1. 直接担任编委/副编委/审稿委员的学者（若有，优先列出）
  2. **该国有哪些学者曾在中国高校获得中文/华文教育/汉语国际教育专业的博士或硕士学位**（留华归国学者）——请尽量列出姓名、获得学位的中国高校、回国后所在机构。这是重要补充信息，即便他们未担任编委。
  3. 受邀担任《世界汉语教学》《华文教学与研究》《国际中文教育》等期刊审稿人的
  4. 在 Hanban/国家语委/CIEF 等机构担任顾问的
  5. 该国学者主编的中文相关图书/教材

  返回值格式："0（无直接编委，但 N 位留华学者：张三-北语博士/李四-北外硕士 ...）"
""",

    20: """【扩展口径·参与中文国际标准】
  1. HSK / HSKK 考试标准制定参与
  2. 国际中文教师专业能力标准
  3. CIEF（国际中文教育基金会）理事会、咨询委员会成员
  4. UNESCO / ISO 下的中文相关标准
  5. 该国本土制定的中文教学大纲（作为代理变量）""",

    26: """【扩展口径·本土化中文学习网站/平台】
  - 该国本土开发的网站/平台
  - 国际通用平台（LingQ、Duolingo、HelloChinese）在该国的本地化版本
  - 孔子学院官方学习平台（eChineseLearning 等）
  - 该国中文教师建立的 WordPress/博客教学平台""",

    27: """【扩展口径·针对性中文学习 APP】
  - 本土开发 APP
  - 国际 APP（HelloChinese、HSK Online、Duolingo Chinese、ChineseSkill、Pleco 等）在该国的使用量/本地化
  - 中国公司开发的针对海外市场 APP（HelloChinese、Chinese Manhua）
  - 小程序、微信端中文学习工具""",

    37: """【扩展口径·中文教育博硕点】※ 小国重点
  关键扩展：
  1. 该国高校开设的博硕点（若有，优先列出）
  2. **该国学生在中国高校攻读中文教育博硕的数量**（在华留学生数据）
  3. 中外合作办学项目中的中文硕博点
  4. 该国高校与中国高校联合培养项目
  5. 涉及中文/汉学/东亚研究方向的非专门博硕点（作为代理）

  返回值格式："0 本土博硕点，但 N 位学生在华攻读中文硕博（北语、北外、华东师大等）"
""",

    39: """【扩展口径·本土中文教师培养项目】
  - 本土高校开设的中文教学法课程/培训
  - 孔子学院的本土教师培训班
  - Hanban 海外教师培训项目派驻该国的项目
  - 中外联合的暑期/寒期培训班
  - 网络教师培训课程（免费/付费）""",

    41: """【扩展口径·世汉理事会会员】
  - CIEF（国际中文教育基金会）理事/顾问/咨询委员
  - 世汉学会、国际汉语教师协会等相关国际组织的该国代表
  - 各地区中文教师协会的该国分部/理事
  - 参加 Hanban 高级研修班的该国教师（作为代理）""",

    43: """【扩展口径·开设中文专业高校占比】
  - 本科/专科层次开设中文专业的高校
  - 中文作为第二外语选修课的高校（作为代理）
  - 中国学/东亚研究专业中包含中文课程的高校
  - 孔子学院所在高校（这些高校至少开中文选修）""",

    45: """【扩展口径·"中文工坊"或孔子学院数】
  - 孔子学院（正式）
  - 孔子课堂
  - 中文工坊
  - **鲁班工坊（若包含中文教学模块）**
  - **海丝学院**（涉外中文教学）
  - 中外合作办学中的中文教学机构
  - 驻外使馆直接开设的中文班/中文学校""",

    46: """【扩展口径·孔子课堂数】
  - 正式注册的孔子课堂
  - 孔子学院下属的教学点
  - Hanban/CIEF 备案的海外中小学教学点
  - 合作中学的中文教学示范点""",

    49: """【扩展口径·偏远地区中文教学点】
  - 首都/一线城市以外的任何中文教学机构
  - 民间中文培训班、华人社团学校
  - 偏远省份高校的中文课程
  - 农村地区的中文教师派驻""",

    51: """【扩展口径·本土化教育机构自主运营能力】
  - 不依赖汉办直接资助可独立运营的中文机构
  - 已实现本土师资 > 50% 的机构
  - 具备独立招生 / 独立认证能力
  - 该国民间中文教育产业的成熟度""",

    59: """【扩展口径·国际组织中文使用次数】
  - 联合国、WTO、WHO、APEC、上合、一带一路 国际场合使用中文
  - 该国代表在国际场合使用中文发言
  - 国际协议/条约中文版的使用""",

    62: """【扩展口径·中华文化产品影响/评价/传播面】
  - 该国 Netflix / 本土流媒体平台上华语影视的占比
  - 该国书店/图书馆中文图书销量
  - 该国对《西游记》《三国》等经典的接受度
  - 该国中国春节 / 中秋等节日庆祝规模""",

    64: """【扩展口径·中文教师职业资格认证】
  - 国际中文教师证书（CTCSOL）在该国的持证人数
  - 本国教师资格体系中的中文专项认证
  - 与中国教育部联合认证的教师资格""",

    # 默认兜底
    0: """【扩展口径·通用放宽原则】
  - 近 5 年（2020-2025）而非仅 2024 的数据均可
  - 相关代理变量的数值（如用孔院学生数代理"学习者规模"）
  - 间接证据：新闻报道、学术论文中的提及
  - 媒体估算：只要有可核验来源即可
  - 若彻底无数据，列出已检索的文件/来源名称，说明为何无数据"""
}


def get_hint(iid: int) -> str:
    return EXPANSION_HINTS.get(iid, EXPANSION_HINTS[0])


async def refine_one_v3(ppx_c, tav_c, ds_c, g, ckpt, ckpt_lock):
    key = f"{g['country']}::{g['id']}"
    async with ckpt_lock:
        if key in ckpt: return ckpt[key]

    t0 = time.time()
    meta = COUNTRY_META.get(g['country'], ([], "", g['country']))
    official = meta[0] + CHINA_SIDE
    en = meta[2]
    hint = get_hint(g['id'])

    # DS 生成扩展搜索计划
    plan_prompt = f"""你是国际中文教育研究员。需要为该"低质量/0 值"格子重查，尽力挖掘数据。

国家：{g['country']}（{en}）
字段ID：{g['id']}
字段名：{g['name']}
字段类型：{g['type']}
分类：{g['category']}
优先级：{g['priority']}
当前可疑值：{g['current'][:150]}
问题类型：{g['kind']}

扩展口径指示：
{hint}

要求：
1. extended_definition: 根据上述扩展指示，写一个针对性的扩展口径（4-6 行）
2. queries: 5 路检索 query（更广谱）
   - perplexity: 综合性中文提问（含扩展口径内的关键词）
   - tavily_zh: 中文窄关键词
   - tavily_en: 英文关键词+国家英文名
   - academic: 学术检索
   - bilateral: 双边文件/MOU 专查（若适用，用中文 + 英文混合）
3. priority_domains: 5-8 个优先域名

返回 JSON:
{{"extended_definition":"", "queries":{{"perplexity":"", "tavily_zh":"", "tavily_en":"", "academic":"", "bilateral":""}}, "priority_domains":[]}}"""

    plan = await ds_call(ds_c, plan_prompt)
    if not plan.get('queries'):
        result = {"value": "V3 无可用策略", "confidence": "low", "_elapsed": time.time()-t0}
        async with ckpt_lock:
            ckpt[key] = result
            _persist_ckpt(ckpt, CKPT)
        return result

    q = plan['queries']
    doms = plan.get('priority_domains', [])

    # 并发跑 5 路
    results = await asyncio.gather(
        ppx_ask(ppx_c, q.get('perplexity', '')),
        tav(tav_c, q.get('tavily_zh', ''), doms),
        tav(tav_c, q.get('tavily_en', ''), official),
        tav(tav_c, q.get('bilateral', ''), None),
        academic_search(q.get('academic', ''), max_per_source=3, proxy_url=PROXY, include_scholar=False),
        return_exceptions=True)
    ppx_r = results[0] if isinstance(results[0], tuple) else ("", [])
    tav_zh = results[1] if isinstance(results[1], tuple) and len(results[1]) == 2 else ([], "")
    tav_en = results[2] if isinstance(results[2], tuple) and len(results[2]) == 2 else ([], "")
    tav_bi = results[3] if isinstance(results[3], tuple) and len(results[3]) == 2 else ([], "")
    acad = results[4] if isinstance(results[4], dict) else {}

    evid = f"""## Perplexity
{ppx_r[0][:2500]}
Citations: {chr(10).join(ppx_r[1][:6])}

## Tavily-zh: {tav_zh[1][:500]}
{chr(10).join(f'  - {x["title"]} | {x["url"]}' for x in tav_zh[0][:4])}

## Tavily-en: {tav_en[1][:500]}
{chr(10).join(f'  - {x["title"]} | {x["url"]}' for x in tav_en[0][:4])}

## Tavily-双边文件: {tav_bi[1][:500]}
{chr(10).join(f'  - {x["title"]} | {x["url"]}' for x in tav_bi[0][:4])}

## Academic
{format_results(acad)[:3000]}"""[:14000]

    synth_prompt = f"""根据扩展口径 + 五路证据，给【{g['country']}】【{g['name']}】最终值。

扩展口径：
{plan.get('extended_definition', '')}

原始扩展指示：
{hint}

取值口径：{g['type']}
字段说明：{g['desc'] if 'desc' in g else ''}
当前可疑值：{g['current'][:150]}

证据：
{evid}

硬约束（非常重要）：
1. 禁止直接回 0/无/有限/少量/若干 除非已用扩展口径全部检索仍无果
2. 若最终值仍为 0，derivation 必须逐条说明：
   - 检索了哪些文件/来源
   - 是否找到扩展口径下的相关数据（如留华学者、双边文件）
   - 补充信息必须写进 additional_notes 字段
3. 若扩展口径找到间接数据（如"无 X，但有 N 个 Y"），value 字段应反映这个复合值
4. 对小国（不丹/黑山/也门/北马其顿/文莱 等），derivation 必须引用具体文件或人名

返回 JSON：
{{
  "value": "最终值，优先数字，其次区间，最后定性分级",
  "proxy_used": "使用的代理变量/扩展口径",
  "derivation": "3-5 行推导过程，引用具体来源",
  "source_urls": ["url1", "url2", ...],
  "confidence": "high|medium|low",
  "is_estimation": true|false,
  "rationale": "2 行判断依据",
  "additional_notes": "扩展口径下找到的相关信息（如留华学者姓名、双边文件条款等），没有则填空字符串"
}}"""

    synth = await ds_call(ds_c, synth_prompt)
    synth["_elapsed"] = time.time() - t0
    async with ckpt_lock:
        ckpt[key] = synth
        _persist_ckpt(ckpt, CKPT)

    log.info("[%s #%d] %.1fs value=%s notes=%s",
             g['country'], g['id'], synth["_elapsed"],
             str(synth.get('value', ''))[:30],
             str(synth.get('additional_notes', ''))[:40])
    return synth


async def run_refine_v3(targets, concurrency=5):
    ckpt = {}
    if CKPT.exists():
        try: ckpt = json.loads(CKPT.read_text(encoding='utf-8'))
        except: pass
    done = len([t for t in targets if f"{t['country']}::{t['id']}" in ckpt])
    log.info("ckpt 已有 %d 格，待处理 %d", done, len(targets) - done)

    ppx_p = get_proxies_for_url("https://api.perplexity.ai", PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", PROXY)
    async with (build_httpx_client(ppx_p or "", 90) as ppx_c,
                build_httpx_client(tav_p or "", 60) as tav_c,
                build_httpx_client("", 120) as ds_c):
        sem = asyncio.Semaphore(concurrency)
        lock = asyncio.Lock()

        async def _task(g):
            async with sem:
                try:
                    return await refine_one_v3(ppx_c, tav_c, ds_c, g, ckpt, lock)
                except Exception as e:
                    log.error("[%s #%d] err: %s", g['country'], g['id'], str(e)[:100])
                    return {"value":"V3 异常","error":str(e)[:200]}

        await asyncio.gather(*[_task(g) for g in targets], return_exceptions=True)
    return ckpt


# ─── 回写到 13 份批次 xlsx ───────────────────────────────────────────────
BATCH_FILES = {
    1: "“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v11_20260422__refined447_20260424.xlsx",
    2: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次2_五国_v4_20260422__refined447_20260424.xlsx",
    3: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次3_南亚I_五国_v_final_20260423__refined447_20260424.xlsx",
    4: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次4_南亚II_五国_v_final_20260423__refined447_20260424.xlsx",
    5: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次5_中亚_五国_v_final_20260423__refined447_20260424.xlsx",
    6: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次6_西亚海湾_五国_v_final_20260423__refined447_20260424.xlsx",
    7: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次7_西亚阿拉伯_五国_v_final_20260423__refined447_20260424.xlsx",
    8: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次8_地中海东岸_五国_v_final_20260423__refined447_20260424.xlsx",
    9: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次9_高加索北非_五国_v_final_20260423__refined447_20260424.xlsx",
    10: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次10_欧洲中部I_五国_v_final_20260423__refined447_20260424.xlsx",
    11: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次11_巴尔干I_五国_v_final_20260423__refined447_20260424.xlsx",
    12: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次12_欧洲中部II_五国_v_final_20260423__refined447_20260424.xlsx",
    13: "“一带一路”沿线国家中文教育发展指标体系数据总表_批次13_东欧_五国_v_final_20260423__refined447_20260424.xlsx",
}
BATCH_COUNTRIES = {
    1: ["蒙古","越南","柬埔寨","菲律宾","新加坡"],
    2: ["马来西亚","泰国","印尼","老挝","缅甸"],
    3: ["文莱","东帝汶","不丹","孟加拉国","尼泊尔"],
    4: ["斯里兰卡","印度","巴基斯坦","阿富汗","马尔代夫"],
    5: ["土库曼斯坦","乌兹别克斯坦","哈萨克斯坦","塔吉克斯坦","吉尔吉斯斯坦"],
    6: ["阿曼","伊朗","阿联酋","卡塔尔","巴林"],
    7: ["沙特阿拉伯","科威特","也门","土耳其","伊拉克"],
    8: ["叙利亚","约旦","黎巴嫩","以色列","格鲁吉亚"],
    9: ["亚美尼亚","阿塞拜疆","埃及","爱沙尼亚","拉脱维亚"],
    10: ["立陶宛","德国","波兰","捷克","斯洛文尼亚"],
    11: ["克罗地亚","塞尔维亚","波黑","黑山","阿尔巴尼亚"],
    12: ["匈牙利","斯洛伐克","北马其顿","罗马尼亚","保加利亚"],
    13: ["摩尔多瓦","乌克兰","白俄罗斯","俄罗斯"],
}

def country_batch(c):
    for b, cs in BATCH_COUNTRIES.items():
        if c in cs: return b, cs.index(c)
    return None, None


def writeback_v3(ckpt, targets):
    """把 V3 结果写回批次 xlsx（原地覆盖 refined447 文件，保留备份一次）。"""
    updated_per_batch = {}
    # 按批次分组
    by_batch = {}
    for t in targets:
        b, idx = country_batch(t['country'])
        if b is None: continue
        by_batch.setdefault(b, []).append((idx, t))

    for bid, items in by_batch.items():
        fp = TOPICS / BATCH_FILES[bid]
        # 先备份
        bak = fp.with_suffix(fp.suffix + ".pre_v3.bak")
        if not bak.exists():
            shutil.copy2(fp, bak)

        wb = openpyxl.load_workbook(fp)
        ws = wb.active
        fill = PatternFill(start_color="FFCCBC", fill_type="solid")
        updated = skipped = 0
        for idx, t in items:
            vcol = 5 + 2 * idx
            key = f"{t['country']}::{t['id']}"
            r = ckpt.get(key)
            if not r:
                skipped += 1; continue
            val = str(r.get('value', '')).strip()
            if not val or val in ('V3 无可用策略', 'V3 异常', '合成失败'):
                skipped += 1; continue

            # 查找该 iid 所在行
            row = None
            for rr in range(2, ws.max_row + 1):
                if ws.cell(rr, 1).value == t['id']:
                    row = rr; break
            if row is None:
                skipped += 1; continue

            proxy = r.get('proxy_used', '')
            deriv = r.get('derivation', '')
            notes = r.get('additional_notes', '')
            rationale = r.get('rationale', '')
            conf = r.get('confidence', '')

            txt = f"【精修V3】{val}"
            if proxy: txt += f"（代理：{proxy[:80]}）"
            if deriv: txt += f"\n推导：{deriv[:300]}"
            if notes: txt += f"\n扩展补充：{notes[:300]}"
            if rationale: txt += f"\n判断：{rationale[:150]} [置信:{conf}]"

            ws.cell(row, vcol, txt).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, vcol).fill = fill
            urls = r.get('source_urls', [])
            if urls:
                if isinstance(urls, str): urls = [urls]
                existing_src = ws.cell(row, vcol + 1).value or ''
                new_src = "; ".join(str(u) for u in urls[:3])[:500]
                combined = (existing_src + " | V3: " + new_src)[:1000] if existing_src else new_src
                ws.cell(row, vcol + 1, combined)
            updated += 1
        wb.save(fp)
        updated_per_batch[bid] = (updated, skipped, len(items))
        log.info("批次 %d: %d 更新 / %d 跳过 / 共 %d", bid, updated, skipped, len(items))
    return updated_per_batch


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--concurrency", type=int, default=5)
    ap.add_argument("--writeback", action="store_true")
    ap.add_argument("--indicator", type=int, default=0, help="只跑指定指标 ID")
    args = ap.parse_args()

    d = json.loads(TARGETS.read_text(encoding='utf-8'))
    targets = d['weak_cells']
    if args.indicator:
        targets = [t for t in targets if t['id'] == args.indicator]
    if args.limit:
        targets = targets[:args.limit]
    # 加 desc 字段（目标 JSON 缺，补上）
    with open(TOPICS / "bri_field_dict.json", encoding='utf-8') as f:
        fd = {int(x['指标ID']): x for x in json.load(f) if x.get('指标ID')}
    for t in targets:
        t['desc'] = (fd.get(t['id'], {}).get('字段说明', '') or '')[:300]

    log.info("V3 目标数: %d", len(targets))

    if args.writeback:
        ckpt = json.loads(CKPT.read_text(encoding='utf-8')) if CKPT.exists() else {}
        with open(TARGETS, encoding='utf-8') as f:
            all_targets = json.load(f)['weak_cells']
        for t in all_targets:
            t['desc'] = ''
        writeback_v3(ckpt, all_targets)
        return

    ckpt = asyncio.run(run_refine_v3(targets, concurrency=args.concurrency))
    writeback_v3(ckpt, targets)
    log.info("V3 完成 ckpt=%d", len(ckpt))


if __name__ == "__main__":
    main()
