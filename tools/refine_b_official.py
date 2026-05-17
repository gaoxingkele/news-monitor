"""B 档 6 格官方名录直查（第三轮）。

每格定制：Tavily raw_content 抓官方名录 + Perplexity 深问 + DeepSeek 合成。

B1 蒙古#43: 蒙古教育部高校登记 (mecss.gov.mn)
B2 蒙古#47: 侨务办华校名录 + 汉办蒙古
B3 越南#43: 越南教育部高校登记 (moet.gov.vn)
B4 柬埔寨#32: 华裔/非华裔学习者结构直查
B5 菲律宾#41: CIEF 官网理事名录 (clef.org.cn)
B6 新加坡#31: MOE 年度统计摘要 (moe.gov.sg)
"""
from __future__ import annotations
import asyncio, io, json, os, sys, logging
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError: pass

import httpx, openpyxl
from openpyxl.styles import PatternFill, Alignment
from news_monitor.proxy import build_httpx_client, get_proxies_for_url

TOPICS = ROOT / "topics"
SRC = TOPICS / "“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v5_20260420.xlsx"
DATE = datetime.now().strftime("%Y%m%d")
V6 = TOPICS / f"“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v6_{DATE}.xlsx"
LOG_MD = TOPICS / f"bri_b_official_{DATE}.md"

PPX = os.environ.get("PERPLEXITY_API_KEY", "")
TAV = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
DS = os.environ.get("DEEPSEEK_API_KEY", "")
PROXY = os.environ.get("LLM_PROXY", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("b")

COL = {"蒙古":(5,6), "越南":(7,8), "柬埔寨":(9,10), "菲律宾":(11,12), "新加坡":(13,14)}

# ── 每格定制配置 ────────────────────────────────────────────────────────────
TASKS = {
    ("蒙古", 43): {
        "name": "开设中文专业高校占比",
        "tavily_queries": [
            ("Mongolia universities total number 2024 list Ministry of Education",
             ["mecss.gov.mn", "edu.mn", "worldbank.org", "uis.unesco.org"]),
            ("Mongolia Chinese language major Chinese studies university program 汉语师范 蒙古",
             ["mecss.gov.mn", "chinese.cn", "ci.cn", "edu.mn"]),
            ("蒙古国 高等教育 高校名录 大学 2024",
             ["chinese.cn", "ci.cn", "mecss.gov.mn", "xinhuanet.com"]),
        ],
        "ppx_q": "蒙古国 2024 年共有多少所高校？其中多少所开设中文专业（汉语/中国研究）？请给出官方/权威来源 URL，列出具体高校名单。",
        "schema": "返回 JSON: {value:'X.XX%', numerator:'开设中文专业高校数 X (来源)', denominator:'全国高校总数 Y (来源)', computation:'X/Y=%', chinese_unis:['高校1','高校2'], source_url:'', confidence:'', rationale:''}",
    },
    ("蒙古", 47): {
        "name": "华校数",
        "tavily_queries": [
            ("蒙古 华校 华文学校 名录 侨办 中华人民共和国国务院侨务办公室",
             ["qwgzyj.gqb.gov.cn", "gqb.gov.cn", "chinese.cn", "ci.cn", "chinaqw.com"]),
            ("Mongolia Chinese school list huawen school Confucius Institute classroom",
             ["ci.cn", "chinese.cn", "hanban.org", "confucius.edu.cn"]),
            ("蒙古国 中文学校 汉语学校 数量 统计",
             ["chinanews.com.cn", "xinhuanet.com", "people.com.cn", "chinese.cn"]),
        ],
        "ppx_q": "蒙古国目前有多少所华校/华文学校（不含孔子学院/孔子课堂）？请给出具体名单和官方来源 URL（如中国侨务办公室、中国驻蒙古使馆、华文教育联合会）。",
        "schema": "返回 JSON: {value:'N所', school_list:['校名1','校名2',...], source_url:'', confidence:'', rationale:''}",
    },
    ("越南", 43): {
        "name": "开设中文专业高校占比",
        "tavily_queries": [
            ("Vietnam universities total number 2024 Ministry of Education MOET list",
             ["moet.gov.vn", "gov.vn", "vnu.edu.vn", "worldbank.org"]),
            ("Vietnam Chinese language major university department 中文系 越南 高校",
             ["moet.gov.vn", "chinese.cn", "ci.cn"]),
            ("越南 高校 总数 2024 中文系 汉语专业 名单",
             ["chinese.cn", "people.com.cn", "xinhuanet.com", "vietnamplus.vn"]),
        ],
        "ppx_q": "越南 2024 年共有多少所高校（大学+学院）？其中多少所开设中文专业或中文系？请列出具体高校清单和官方来源 URL。",
        "schema": "返回 JSON: {value:'X.XX%', numerator:'开设中文专业高校数 X', denominator:'越南高校总数 Y', computation:'X/Y=%', chinese_unis:['高校1'], source_url:'', confidence:'', rationale:''}",
    },
    ("柬埔寨", 32): {
        "name": "非华裔学习者比例",
        "tavily_queries": [
            ("Cambodia Chinese language learners ethnic Chinese non-Chinese ratio demographics 华裔 非华裔",
             ["moeys.gov.kh", "rupp.edu.kh", "chinese.cn", "ci.cn"]),
            ("柬埔寨 端华学校 中文学习者 华裔 非华裔 结构",
             ["chinese.cn", "qwgzyj.gqb.gov.cn", "chinanews.com.cn", "xinhuanet.com"]),
            ("Cambodia Confucius Institute RUPP students ethnic composition",
             ["ci.cn", "rupp.edu.kh", "ci.rupp.edu.kh"]),
        ],
        "ppx_q": "柬埔寨中文学习者中，非华裔（高棉族等非华人血统）占比是多少？请给出官方/学术统计来源。若为倒推（100%-华裔），请同时给出华裔占比的直接统计来源。",
        "schema": "返回 JSON: {value:'X%', numerator:'非华裔学习者 X 人 (或比例)', denominator:'总学习者 Y 人', computation:'X/Y=%', method:'direct(直接统计)/derivation(倒推)', source_url:'', confidence:'', rationale:''}",
    },
    ("菲律宾", 41): {
        "name": "世汉理事会会员数（CIEF理事会会员数）",
        "tavily_queries": [
            ("CIEF 世汉理事会 理事 名单 Philippines 菲律宾",
             ["clef.org.cn", "chinese.cn", "ci.cn", "hanban.org"]),
            ("World Chinese Language Education Federation council members Philippines",
             ["clef.org.cn", "chinese.cn"]),
            ("CIEF council members list 世界汉语教育联合会 理事会",
             ["clef.org.cn", "chinese.cn"]),
        ],
        "ppx_q": "世界汉语教育联合会（CIEF/CLEF）理事会中，来自菲律宾的理事/会员有多少人？请给出 CIEF 官方理事名录的直接 URL。",
        "schema": "返回 JSON: {value:'N', members:['姓名1 (机构)','姓名2'], source_url:'', confidence:'', rationale:''}",
    },
    ("新加坡", 31): {
        "name": "基础教育阶段中文学习者比例",
        "tavily_queries": [
            ("Singapore MOE education statistics digest primary secondary mother tongue Chinese enrollment",
             ["moe.gov.sg", "data.gov.sg", "singstat.gov.sg"]),
            ("Singapore Chinese mother tongue language CL1 CL2 primary school enrollment percentage",
             ["moe.gov.sg", "straitstimes.com"]),
            ("新加坡 基础教育 中文 华文 母语 在校生 人数 占比 MOE",
             ["moe.gov.sg", "chinese.cn", "straitstimes.com"]),
        ],
        "ppx_q": "新加坡基础教育（小学+中学）阶段修读华文（中文母语课程）的学生占全体学生比例是多少？请引用 MOE Education Statistics Digest 最新年度数据，给出分子（华文学习者）+分母（总学生数）+来源 URL。",
        "schema": "返回 JSON: {value:'X%', numerator:'华文学习者 X 人 (年份)', denominator:'总学生数 Y 人 (年份)', computation:'X/Y=%', source_url:'', confidence:'', rationale:''}",
    },
}

async def tav_raw(client, query, domains):
    """Tavily with raw_content to get full page text."""
    if not TAV: return []
    try:
        resp = await client.post(
            "https://api.tavily.com/search",
            json={"api_key": TAV, "query": query, "search_depth": "advanced",
                  "max_results": 8, "include_domains": domains[:20],
                  "include_answer": True, "include_raw_content": True},
            timeout=90.0,
        )
        resp.raise_for_status()
        d = resp.json()
        out = [{"url": r.get("url", ""), "title": r.get("title", ""),
                "snippet": (r.get("content") or "")[:500],
                "raw": (r.get("raw_content") or "")[:3000]}
               for r in (d.get("results") or [])[:8]]
        return out, d.get("answer", "")
    except Exception as e:
        log.warning("tav err: %s", str(e)[:120])
        return [], ""

async def ppx_ask(client, question):
    if not PPX: return "", []
    try:
        resp = await client.post(
            "https://api.perplexity.ai/chat/completions",
            json={"model": "sonar",
                  "messages": [
                      {"role": "system", "content": "你是教育统计专家。只给出来源可验证的数字。若无确切数据，明确说'无确切数据'。"},
                      {"role": "user", "content": question},
                  ],
                  "temperature": 0.05, "return_citations": True},
            headers={"Authorization": f"Bearer {PPX}", "Content-Type": "application/json"},
            timeout=120.0,
        )
        resp.raise_for_status()
        d = resp.json()
        return d["choices"][0]["message"]["content"], d.get("citations", [])[:10]
    except Exception as e:
        log.warning("ppx err: %s", str(e)[:120]); return "", []

async def ds_synth(client, country, iid, cfg, ppx_ans, ppx_cites, tavily_bundles):
    """tavily_bundles: list of (answer, results) from each query."""
    evid = []
    tav_ans_block = []
    for i, (results, tav_ans) in enumerate(tavily_bundles):
        if tav_ans: tav_ans_block.append(f"[Tavily Q{i+1} 答案] {tav_ans}")
        for r in results[:4]:
            evid.append(f"[{r.get('title','')}]\nURL: {r.get('url','')}\n摘要: {r.get('snippet','')}\n正文片段: {r.get('raw','')[:1500]}")
    evid_block = "\n\n---\n\n".join(evid[:20])
    tav_ans_str = "\n".join(tav_ans_block)
    ppx_cite_str = "\n".join(f"- {c}" for c in ppx_cites)

    prompt = f"""你是官方数据审核员。为【{country}】指标【{cfg['name']}】（#{iid}）合成最终值。

## Perplexity 深问回答
{ppx_ans}
Perplexity citations:
{ppx_cite_str}

## Tavily 官方名录 answer 汇总
{tav_ans_str}

## Tavily 官方名录 原始证据
{evid_block[:12000]}

## 输出 Schema
{cfg['schema']}

## 规则
- 必须优先采用官方名录/官方统计的直接数据（MOE、CIEF、侨办等）
- 若多源冲突，取最权威（官方 > 新华社/人民网 > 学术 > 百科）
- 所有数字必须能在证据中找到锚点，否则 confidence=low
- 若真的无法确定，value 给最接近的估算，标注 is_still_estimation=true
"""
    if not DS:
        return {"value": ppx_ans[:300], "confidence": "low"}
    try:
        resp = await client.post(
            "https://api.deepseek.com/v1/chat/completions",
            json={"model": "deepseek-chat", "messages": [{"role": "user", "content": prompt}],
                  "temperature": 0.05, "response_format": {"type": "json_object"}},
            headers={"Authorization": f"Bearer {DS}", "Content-Type": "application/json"},
            timeout=120.0,
        )
        resp.raise_for_status()
        return json.loads(resp.json()["choices"][0]["message"]["content"])
    except Exception as e:
        log.warning("ds err: %s", str(e)[:120])
        return {"value": "合成失败", "confidence": "low"}

async def solve_cell(country, iid, cfg, ppx_c, tav_c, ds_c):
    t0 = asyncio.get_event_loop().time()
    # Tavily 3 轮
    tav_tasks = [tav_raw(tav_c, q, doms) for q, doms in cfg["tavily_queries"]]
    tav_results = await asyncio.gather(*tav_tasks, return_exceptions=True)
    tav_bundles = [r for r in tav_results if isinstance(r, tuple)]

    # Perplexity 深问
    ppx_ans, ppx_cites = await ppx_ask(ppx_c, cfg["ppx_q"])

    # DS 合成
    result = await ds_synth(ds_c, country, iid, cfg, ppx_ans, ppx_cites, tav_bundles)
    dt = asyncio.get_event_loop().time() - t0
    log.info("%s#%d %.1fs value=%s conf=%s",
             country, iid, dt, str(result.get("value",""))[:40], result.get("confidence","?"))
    return result

async def run():
    ppx_p = get_proxies_for_url("https://api.perplexity.ai", PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", PROXY)
    results = {}
    async with (build_httpx_client(ppx_p or "", 120) as ppx_c,
                build_httpx_client(tav_p or "", 90) as tav_c,
                build_httpx_client("", 120) as ds_c):
        tasks = [solve_cell(c, i, cfg, ppx_c, tav_c, ds_c) for (c, i), cfg in TASKS.items()]
        rs = await asyncio.gather(*tasks)
        for (c, i), r in zip(TASKS.keys(), rs):
            results[f"{c}::{i}"] = r

    with open(ROOT / "output" / f"bri_b_official_{DATE}.json", "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    # 写 v6 xlsx
    wb = openpyxl.load_workbook(SRC)
    ws = wb.active
    est_fill = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")
    refined_fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")

    log_md = [f"# B 档官方名录直查 (v6)", f"时间：{datetime.now():%Y-%m-%d %H:%M}\n"]

    for (country, iid), cfg in TASKS.items():
        r = results[f"{country}::{iid}"]
        vcol, scol = COL[country]
        row = next(rr for rr in range(2, ws.max_row+1) if ws.cell(rr,1).value == iid)
        val = str(r.get("value", "")).strip()
        if not val: continue
        is_est = str(r.get("is_still_estimation", False)).lower() == "true"
        comp = r.get("computation", "")
        suffix = ""
        if comp and comp != "N/A": suffix = f"（计算：{comp}）"
        elif r.get("school_list"):
            schools = r["school_list"]
            if isinstance(schools, list) and len(schools) <= 8:
                suffix = f"（名录：{', '.join(str(s) for s in schools)}）"
        prefix = "【估算】" if is_est else ""
        cell_txt = f"{prefix}{val}{suffix}"
        ws.cell(row, vcol, cell_txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, vcol).fill = est_fill if is_est else refined_fill
        if r.get("source_url"):
            ws.cell(row, scol, r["source_url"])

        log_md.append(f"\n## {country}#{iid} {cfg['name']}")
        log_md.append(f"- **值**：{val}")
        if r.get("numerator"): log_md.append(f"- 分子：{r['numerator']}")
        if r.get("denominator"): log_md.append(f"- 分母：{r['denominator']}")
        if comp: log_md.append(f"- 计算：{comp}")
        if r.get("school_list"):
            log_md.append(f"- 名录：")
            for s in (r['school_list'] or [])[:15]:
                log_md.append(f"  - {s}")
        if r.get("members"):
            log_md.append(f"- 会员：")
            for m in (r['members'] or [])[:10]:
                log_md.append(f"  - {m}")
        if r.get("chinese_unis"):
            log_md.append(f"- 中文专业高校：")
            for u in (r['chinese_unis'] or [])[:10]:
                log_md.append(f"  - {u}")
        if r.get("method"): log_md.append(f"- 方法：{r['method']}")
        log_md.append(f"- 置信度：{r.get('confidence','?')}")
        log_md.append(f"- 主要来源：{r.get('source_url','')}")
        log_md.append(f"- 理由：{r.get('rationale','')}")

    wb.save(V6)
    LOG_MD.write_text("\n".join(log_md), encoding="utf-8")
    log.info("saved %s", V6.name)
    log.info("saved %s", LOG_MD.name)

if __name__ == "__main__":
    asyncio.run(run())
