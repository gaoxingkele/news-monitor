from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROWS = [
    {
        "platform": "X/Twitter",
        "date": "2026-04-07",
        "account_or_thread": "ABC",
        "region": "美国/澳大利亚",
        "orientation": "中性",
        "heat_signal": "331 likes / 86 reposts / 73 replies / 101k views",
        "narrative": "和平之旅/历史性访问",
        "summary": "中性报道，强调郑丽文应习近平邀请展开“和平之旅”，但背景是北京推进统一叙事。",
        "link": "https://x.com/ABC/status/2041473564742684966",
        "notes": "主流媒体高传播样本",
    },
    {
        "platform": "X/Twitter",
        "date": "2026-04-07",
        "account_or_thread": "Carl Zha",
        "region": "美国/亲中传播圈",
        "orientation": "正向",
        "heat_signal": "8k likes / 1.3k reposts / 408k views",
        "narrative": "和平之旅/历史性访问",
        "summary": "将此次访问定义为十年来首次KMT主席访中，突出历史性与对话价值。",
        "link": "https://x.com/CarlZha/status/2041481953766203445",
        "notes": "本轮最强传播帖之一",
    },
    {
        "platform": "X/Twitter",
        "date": "2026-04-08",
        "account_or_thread": "Carl Zha（二次传播）",
        "region": "美国/亲中传播圈",
        "orientation": "正向",
        "heat_signal": "3.4k likes / 465 reposts / 84k views",
        "narrative": "历史性象征/KMT路线再定义",
        "summary": "放大南京中山陵、赴京会习的仪式化意义，继续强调历史时刻。",
        "link": "https://x.com/CarlZha/status/2041788238319513732",
        "notes": "二次传播仍高热",
    },
    {
        "platform": "X/Twitter",
        "date": "2026-04-08",
        "account_or_thread": "Ian Miles Cheong",
        "region": "国际英文圈",
        "orientation": "正向/亲中",
        "heat_signal": "888 likes / 277 reposts / 120k views",
        "narrative": "历史性象征/KMT路线再定义",
        "summary": "将其解释为台湾进一步靠近中国、统一路线升温。",
        "link": "https://x.com/ianmiles/status/2041904586387312717",
        "notes": "高传播评论账号",
    },
    {
        "platform": "X/Twitter",
        "date": "2026-04-07",
        "account_or_thread": "denisewu",
        "region": "台湾/海外评论圈",
        "orientation": "负向",
        "heat_signal": "104 likes / 14 reposts / 7k views",
        "narrative": "亲中/卖台/统战样板",
        "summary": "明显负面，把郑丽文比附为牺牲台湾利益、为北京服务的政治人物。",
        "link": "https://x.com/denisewu/status/2041369740359377294",
        "notes": "代表反对派评论口径",
    },
    {
        "platform": "X/Twitter",
        "date": "2026-04-07",
        "account_or_thread": "clashreport",
        "region": "国际地缘政治圈",
        "orientation": "中性",
        "heat_signal": "404 likes / 52 reposts / 36k views",
        "narrative": "和平之旅/历史性访问",
        "summary": "中性放大“十年来首次KMT领袖访中”的新闻性和地缘政治意味。",
        "link": "https://x.com/clashreport/status/2041402814908268663",
        "notes": "国际二级扩散号",
    },
    {
        "platform": "Reddit",
        "date": "2026-04-07~2026-04-09",
        "account_or_thread": "r/worldnews 热帖 1",
        "region": "英文国际圈",
        "orientation": "偏负向/怀疑",
        "heat_signal": "约 +450",
        "narrative": "影响军购与台美关系",
        "summary": "评论区聚焦访问是否削弱台湾安全、KMT是否更靠近北京，以及香港作为反例。",
        "link": "https://www.reddit.com/r/worldnews/comments/1sdnt9j/taiwan_opposition_leader_to_visit_china_as/",
        "notes": "英文圈代表性讨论",
    },
    {
        "platform": "Reddit",
        "date": "2026-04-07",
        "account_or_thread": "r/worldnews 热帖 2",
        "region": "英文国际圈",
        "orientation": "负向",
        "heat_signal": "约 +67",
        "narrative": "亲中/卖台/统战样板",
        "summary": "评论区更情绪化，典型表达包括对KMT的羞辱性评价与对统一路线的否定。",
        "link": "https://www.reddit.com/r/worldnews/comments/1sexjak/taiwans_opposition_leader_arrives_in_china_for_a/",
        "notes": "情绪强度高",
    },
    {
        "platform": "Reddit",
        "date": "2026-04-07",
        "account_or_thread": "r/worldnews 热帖 3",
        "region": "英文国际圈",
        "orientation": "负向/讽刺",
        "heat_signal": "约 +35",
        "narrative": "历史性象征/KMT路线再定义",
        "summary": "更多从历史反差角度批评KMT从反共到更公开靠近北京。",
        "link": "https://www.reddit.com/r/worldnews/comments/1serrde/cheng_liwun_taiwan_opposition_kuomintang_leader/",
        "notes": "路线讽刺类样本",
    },
    {
        "platform": "Reddit",
        "date": "背景串，仍有参考价值",
        "account_or_thread": "r/China",
        "region": "英文国际圈",
        "orientation": "偏负向/怀疑",
        "heat_signal": "约 +142",
        "narrative": "影响军购与台美关系",
        "summary": "系统讨论阻军购预算、对北京示好、绕开赖政府等问题。",
        "link": "https://www.reddit.com/r/China/comments/1s80s72/taiwan_opposition_leader_accepts_xis_invite_to/",
        "notes": "背景参考，不属于48小时新帖",
    },
    {
        "platform": "Dcard",
        "date": "2026-04-08",
        "account_or_thread": "郑丽文参访大陆",
        "region": "台湾年轻用户",
        "orientation": "负向",
        "heat_signal": "热度可见，互动量公开页可见",
        "narrative": "亲中/卖台/统战样板",
        "summary": "偏负面，代表年轻用户对访中时机与路线风险的警惕。",
        "link": "https://www.dcard.tw/f/talk/p/261254296",
        "notes": "年轻用户情绪池",
    },
    {
        "platform": "Dcard",
        "date": "2026-04-09",
        "account_or_thread": "真的假的，郑丽文直接公开高喊中华民国",
        "region": "台湾年轻用户",
        "orientation": "正向/辩护",
        "heat_signal": "热度可见，互动量公开页可见",
        "narrative": "和平之旅/对话降温",
        "summary": "少见的正向/辩护型帖子，试图证明其并非完全向北京让步。",
        "link": "https://www.dcard.tw/f/trending/p/261259450",
        "notes": "逆风支持样本",
    },
    {
        "platform": "Dcard",
        "date": "2026-04-08",
        "account_or_thread": "郑丽文打脸赖政府",
        "region": "台湾年轻用户",
        "orientation": "正向/反绿",
        "heat_signal": "热度可见，互动量公开页可见",
        "narrative": "岛内选战化/蓝绿路线战",
        "summary": "代表反绿、反赖政府操作的逆向声量。",
        "link": "https://www.dcard.tw/f/trending/p/261251055",
        "notes": "路线战样本",
    },
    {
        "platform": "PTT",
        "date": "持续讨论入口",
        "account_or_thread": "HatePolitics",
        "region": "台湾政治论坛",
        "orientation": "分裂/情绪化",
        "heat_signal": "高讨论度",
        "narrative": "岛内选战化/蓝绿路线战",
        "summary": "大量讨论转向‘卖台/抹红/加分减分/2028路线战’，情绪强于事实。",
        "link": "https://www.ptt.cc/bbs/HatePolitics/M.1774838712.A.083.html",
        "notes": "岛内即时情绪最强平台之一",
    },
    {
        "platform": "Mobile01",
        "date": "2026-04-08前后持续可见",
        "account_or_thread": "国共交流北京主导",
        "region": "台湾综合论坛",
        "orientation": "负向/警戒",
        "heat_signal": "高浏览讨论串",
        "narrative": "亲中/卖台/统战样板",
        "summary": "把此行解释为北京设定议程、统战推进。",
        "link": "https://www.mobile01.com/topicdetail.php?f=638&t=7248718",
        "notes": "警戒口径样本",
    },
    {
        "platform": "Mobile01",
        "date": "2026-04-08前后持续可见",
        "account_or_thread": "中网灌爆“臭乞丐”",
        "region": "台湾综合论坛",
        "orientation": "复杂/围观",
        "heat_signal": "高浏览讨论串",
        "narrative": "大陆网民反应/跨平台转述",
        "summary": "补到大陆网民嘲讽与敌意反应，但仍属媒体/论坛转述。",
        "link": "https://www.mobile01.com/topicdetail.php?f=638&p=1&t=7244919",
        "notes": "补充侧证",
    },
    {
        "platform": "Mobile01",
        "date": "2026-04-08前后持续可见",
        "account_or_thread": "简舒培批“投名状”",
        "region": "台湾综合论坛",
        "orientation": "负向",
        "heat_signal": "高浏览讨论串",
        "narrative": "岛内选战化/蓝绿路线战",
        "summary": "代表绿营支持者攻击框架在论坛继续扩散。",
        "link": "https://www.mobile01.com/topicdetail.php?f=638&t=7245404",
        "notes": "攻击框架扩散样本",
    },
    {
        "platform": "LIHKG",
        "date": "背景串",
        "account_or_thread": "郑丽文相关旧串",
        "region": "香港论坛",
        "orientation": "负向",
        "heat_signal": "非当期热帖",
        "narrative": "长期标签化认知",
        "summary": "提供郑丽文长期身份认同争议的背景样本，不属于本轮48小时主战场。",
        "link": "https://lihkg.com/thread/3979309/page/1",
        "notes": "仅作背景参考",
    },
]


def autosize(ws) -> None:
    widths = {
        "A": 12, "B": 16, "C": 24, "D": 16, "E": 12,
        "F": 34, "G": 22, "H": 48, "I": 58, "J": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_workbook(output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "热帖清单"
    headers = ["平台", "日期", "账号/帖子", "地区", "倾向", "热度信号", "叙事分类", "观点摘要", "链接", "备注"]
    ws.append(headers)
    style_header(ws[1])
    ws.freeze_panes = "A2"
    for row in ROWS:
        ws.append([
            row["platform"],
            row["date"],
            row["account_or_thread"],
            row["region"],
            row["orientation"],
            row["heat_signal"],
            row["narrative"],
            row["summary"],
            row["link"],
            row["notes"],
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws)

    ws2 = wb.create_sheet("平台汇总")
    ws2.append(["平台", "样本数", "主要特征", "建议用途"])
    style_header(ws2[1])
    platform_counter = Counter(row["platform"] for row in ROWS)
    platform_notes = {
        "X/Twitter": ("传播量最大、国际扩散快、分裂明显", "看大V、媒体、国际传播方向"),
        "Reddit": ("英文圈路线与安全争论集中", "看英文评论区与国际受众理解"),
        "Dcard": ("年轻用户即时政治情绪明显", "看年轻群体态度与身份认同争论"),
        "PTT": ("情绪最强、选战化最明显", "看岛内即时攻击点与口号"),
        "Mobile01": ("逆向声量较多、反绿话语明显", "看与PTT不同的岛内样本"),
        "LIHKG": ("当期样本弱、背景参考为主", "仅作背景观察"),
    }
    for platform, count in platform_counter.items():
        notes = platform_notes.get(platform, ("", ""))
        ws2.append([platform, count, notes[0], notes[1]])
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 34

    ws3 = wb.create_sheet("叙事汇总")
    ws3.append(["叙事分类", "样本数", "代表平台", "说明"])
    style_header(ws3[1])
    narrative_counter = Counter(row["narrative"] for row in ROWS)
    narrative_examples = {
        "和平之旅/历史性访问": "X/Twitter",
        "历史性象征/KMT路线再定义": "X/Twitter, Reddit",
        "亲中/卖台/统战样板": "X/Twitter, Reddit, Dcard, Mobile01",
        "影响军购与台美关系": "Reddit",
        "岛内选战化/蓝绿路线战": "Dcard, PTT, Mobile01",
        "大陆网民反应/跨平台转述": "Mobile01",
        "长期标签化认知": "LIHKG",
    }
    narrative_desc = {
        "和平之旅/历史性访问": "国际传播层面声量强，多由中性媒体或正向包装账号放大。",
        "历史性象征/KMT路线再定义": "强调十年来首次KMT主席访中、路线信号大于单次访问。",
        "亲中/卖台/统战样板": "中文社媒与评论区最强的负面批评框架。",
        "影响军购与台美关系": "英文圈更关注此行是否削弱台湾安全与对美关系。",
        "岛内选战化/蓝绿路线战": "台湾论坛高频把事件转为政党对冲与路线站队。",
        "大陆网民反应/跨平台转述": "用于补侧大陆网民舆情，但需注意多为二手转述。",
        "长期标签化认知": "提供人物背景认知，不作为48小时主证据。",
    }
    for narrative, count in narrative_counter.items():
        ws3.append([narrative, count, narrative_examples.get(narrative, ""), narrative_desc.get(narrative, "")])
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 24
    ws3.column_dimensions["D"].width = 52

    wb.save(output_path)


def main() -> None:
    out_dir = Path(r"D:\BaiduSyncdisk\2026-04\台情")
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "郑丽文来访大陆过去48小时社交热帖归类_20260409_表格版.xlsx"
    build_workbook(path)
    print(path)
    print(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")


if __name__ == "__main__":
    main()
