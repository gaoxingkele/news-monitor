from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


SRC = Path(r"D:\BaiduSyncdisk\aicoding\news-monitor\topics\一带一路_数据采集任务清单_拆解版.xlsx")
OUT = Path(r"D:\BaiduSyncdisk\aicoding\news-monitor\topics\一带一路_数据采集任务清单_拆解版_已回填五国.xlsx")

DATA = {
    "新加坡": [
        (1, "已确认", "全方位高质量的前瞻性伙伴关系", "https://www.mfa.gov.sg/newsroom/press-statements-transcripts-and-photos/exchange-of-congratulatory-messages-between-sg-and-china", "中新关系层级高且稳定。"),
        (2, "已确认", "基础教育、高中前、高等教育、成人教育", "https://www.moe.gov.sg/primary/curriculum/mother-tongue-languages/learning-in-school", "官方母语课程体系完善。"),
        (3, "已确认", "官方语言之一；母语教育的重要组成部分", "https://www.tech.gov.sg/technews/how-singpass-learnt-three-languages", "中文具官方语言地位。"),
        (10, "部分确认", "总体低风险，但存在英语主导导致中文使用环境弱化的结构性压力", "https://www.sg101.gov.sg/resources/connexionsg/everythingsg-billingual-policy/", "属于评估性结论。"),
        (17, "已确认", "有", "https://www.sccl.sg/en/about-sccl", "SCCL 为代表性机构。"),
        (37, "部分确认", "至少1个成熟研究生培养点", "https://fass.nus.edu.sg/cs/discovergrad/", "研究生培养已可核验。"),
        (39, "部分确认", "已确认多条培养通道", "https://www.moe.gov.sg/careers/teaching-scholarships-sponsorships/moe-teaching-sponsorships/chinese-language-teacher", "项目数待统一口径。"),
        (40, "部分确认", "已确认培训规模较大，精确年度值待补", "https://www.moe.gov.sg/news/press-releases/20230406-new-teaching-resource-to-help-primary-students-build-strong-foundations-in-chinese-language", "已有700多名教师受训样本。"),
        (42, "已确认", "至少3所", "https://www.suss.edu.sg/programmes/detail/Bachelor-of-Arts-in-Chinese-Studies", "NUS/NTU/SUSS 可确认。"),
        (45, "已确认", "1", "https://www.ntu.edu.sg/ci/about-us", "CI-NTU。"),
        (48, "已确认", "至少1个", "https://www.ntuc.org.sg/uportal/about-us/affiliated-unions/singapore-chinese-teachers-union", "已确认教师工会。"),
        (60, "部分确认", "存在持续稳定举办机制，精确次数待补", "https://www.sccl.sg/en/", "活动存在，待量化。"),
        (63, "部分确认", "存在商务中文、BCT等课程，比例待补", "https://www.ntu.edu.sg/ci/bct", "职业导向课程存在。"),
        (64, "已确认", "有", "https://www.moe.gov.sg/careers/become-teachers/pri-sec-jc-ci/chinese-language-teaching/applicants-with-teaching-qualifications", "有明确教师资格路径。"),
        (65, "部分确认", "已确认存在，精确项目数待补", "https://www.ntu.edu.sg/ci/programmes", "存在定制化项目。"),
    ],
    "马来西亚": [
        (1, "已确认", "全面战略伙伴关系，并持续推进命运共同体建设", "https://www.kln.gov.my/web/guest/home?_101_assetEntryId=9861626&_101_struts_action=%2Fasset_publisher%2Fview_content&_101_type=content&_101_urlTitle=joint-statement-between-the-people-s-republic-of-china-and-malaysia-on-deepening-the-comprehensive-strategic-partnership-towards-china-malaysia-commun&inheritRedirect=false&p_p_id=101&p_p_lifecycle=0&p_p_mode=view&p_p_state=maximized", "2024年官方联合声明。"),
        (2, "已确认", "基础教育明确纳入；中学、高教、社区教育持续存在", "https://www.moe.gov.my/storage/files/shares/Dasar/Dasar%20Pendidikan%20Kebangsaan/Dasar%20Pendidikan%20Kebangsaan%20Edisi%20Keempat.pdf", "SJKC 以中文为教学媒介。"),
        (3, "已确认", "非官方语言；在华小中具有法定教学媒介地位", "https://www.malaysia.gov.my/en/government/kenali-malaysia/bahasa-rasmi", "中文不属官方语言。"),
        (10, "部分确认", "中等；制度存在但长期有政策与舆论争议", "https://www.moe.gov.my/sistem-pendidikan", "多源流教育议题常被政治化。"),
        (17, "已确认", "有", "https://fcs.utar.edu.my/", "UTAR、UM 等均可支撑。"),
        (37, "已确认", "有", "https://study.um.edu.my/master-of-chinese-studies-coursework", "UTAR 和 UM 可确认硕博培养。"),
        (39, "部分确认", "存在成熟社群与学校系统培养机制，量化待补", "https://www.dongzong.my/en/about-us/", "董教总网络是核心支撑。"),
        (42, "已确认", "至少2所", "https://study.utar.edu.my/chinese-studies.php", "UTAR 与 UM。"),
        (45, "已确认", "至少1个", "https://www.um.edu.my/global-engagement", "UM 页面列出 Kong Zi Institute。"),
        (48, "已确认", "有且网络化程度高", "https://jiaozong.org.my/", "教师协会与董事会网络明显。"),
        (63, "部分确认", "存在并在高校中强化", "https://fcs.utar.edu.my/", "应用中文导向明显。"),
    ],
    "泰国": [
        (1, "已确认", "全面战略合作伙伴关系", "https://mfa.go.th/en/content/cscp-2024-en?cate=5d5bcb4e15e39c306000683e", "外交部官方可核验。"),
        (2, "部分确认", "学校项目与高校体系已存在", "https://www.moe.go.th/%E0%B8%AB%E0%B9%89%E0%B8%AD%E0%B8%87%E0%B9%80%E0%B8%A3%E0%B8%B5%E0%B8%A2%E0%B8%99%E0%B8%A0%E0%B8%B2%E0%B8%A9%E0%B8%B2%E0%B8%88%E0%B8%B5%E0%B8%99/", "中文课堂项目存在。"),
        (3, "已确认", "高热度外语；非官方语言", "https://thailand.go.th/page/thai-language", "泰语为官方语言。"),
        (10, "部分确认", "低到中", "https://www.moe.go.th/%E0%B8%AB%E0%B9%89%E0%B8%AD%E0%B8%87%E0%B9%80%E0%B8%A3%E0%B8%B5%E0%B8%A2%E0%B8%99%E0%B8%A0%E0%B8%B2%E0%B8%A9%E0%B8%B2%E0%B8%88%E0%B8%B5%E0%B8%99/", "总体政策友好。"),
        (17, "已确认", "有", "https://hcuchinamonitor.hcu.ac.th/en/home-en", "研究与智库支撑存在。"),
        (37, "已确认", "有", "https://www.arts.chula.ac.th/east/chinese/Curriculum.html", "Chula 可确认 BA/MA/PhD。"),
        (39, "部分确认", "存在，且与高校体系绑定", "https://www.arts.chula.ac.th/east/chinese/history_en.html", "教师培养路径已存在。"),
        (42, "已确认", "至少1所成熟样本", "https://www.arts.chula.ac.th/east/chinese/index_en.html", "先确认 Chula。"),
        (63, "部分确认", "存在明显商务与产业导向", "https://hcuchinamonitor.hcu.ac.th/en/home-en", "应用导向较强。"),
    ],
    "印尼": [
        (1, "已确认", "全面战略伙伴关系，并持续向命运共同体方向深化", "https://www.mfa.gov.cn/eng/zy/jj/2022/cxesgjtytjhtg/202211/t20221117_10976784.html", "联合声明可核验。"),
        (2, "已确认", "中学课程标准与评测体系已进入；高校体系存在", "https://guru.kemdikbud.go.id/kurikulum/referensi-penerapan/capaian-pembelajaran/sd-sma/bahasa-mandarin/", "Bahasa Mandarin 已进入 CP。"),
        (3, "已确认", "外语科目；非官方语言", "https://indonesia.go.id//kategori/sosial-budaya/10256/bahasa-indonesia-resmi-digunakan-sebagai-bahasa-kerja-di-sidang-umum-unesco?lang=1%3Flang%3D1", "印尼语为国家语言。"),
        (10, "部分确认", "中等", "https://guru.kemdikbud.go.id/kurikulum/referensi-penerapan/capaian-pembelajaran/sd-sma/bahasa-mandarin/", "制度化深度仍在上升。"),
        (17, "部分确认", "有高校与孔院支撑", "https://international.ui.ac.id/humanities-bachelors/", "国家级智库证据待补。"),
        (39, "部分确认", "存在教师认证/专业化路径", "https://pusatinformasi.guru.kemdikbud.go.id/hc/en-us/articles/35466120859161-Informasi-Tentang-Rumpun-Bidang-Studi-Pada-Sertifikasi-Pendidik", "Bahasa Mandarin 已入认证体系。"),
        (42, "已确认", "至少2所", "https://chinese.binus.ac.id/", "UI 与 BINUS。"),
        (45, "已确认", "至少1个", "https://jurnal.uns.ac.id/maobi/about/contact", "UNS Confucius Institute。"),
        (64, "已确认", "有", "https://pusatinformasi.guru.kemdikbud.go.id/hc/en-us/articles/35466120859161-Informasi-Tentang-Rumpun-Bidang-Studi-Pada-Sertifikasi-Pendidik", "教师认证可核验。"),
    ],
    "越南": [
        (1, "已确认", "全面战略合作伙伴关系，并推进具有战略意义的命运共同体", "https://mofa.gov.vn/web/ministry-of-foreign-affairs/detail/chi-tiet/developing-ties-with-china-a-strategic-priority-in-vietnam-s-foreign-policy-deputy-pm-58405-591.html", "越方官方表述明确。"),
        (2, "已确认", "中学阶段已进入外语1课程框架；高校体系成熟", "https://moet.gov.vn/giaoducquocdan/day-va-hoc-ngoai-ngu/Pages/Default.aspx?ItemID=7461", "MOET 外语1中文课程。"),
        (3, "已确认", "正式外语课程选项；非国家共同语言", "https://mofa.gov.vn/web/guest/dan-toc-ngon-ngu", "越语为共同语言。"),
        (10, "部分确认", "中低", "https://moet.gov.vn/giaoducquocdan/day-va-hoc-ngoai-ngu/Pages/Default.aspx?ItemID=7461", "主要是竞争语种压力。"),
        (17, "已确认", "有", "https://ussh.vnu.edu.vn/en/organizational-structure/centers-institutes-companies/center-for-chinese-studies-13067.html", "USSH 与 ULIS 可支撑。"),
        (37, "已确认", "有", "https://ulis.vnu.edu.vn/co-cau-to-chuc/khoa-nnvh-trung/", "ULIS 明确含硕博培养。"),
        (39, "已确认", "有", "https://ulis.vnu.edu.vn/khai-mac-khoa-boi-duong-giao-vien-tieng-trung-tai-viet-nam/", "教师培训班已可核验。"),
        (42, "已确认", "至少2所", "https://chinese.ftu.edu.vn/index.php/gi%E1%BB%9Bi-thi%E1%BB%87u/l%E1%BB%8Bch-s%E1%BB%AD-h%C3%ACnh-th%C3%A0nh", "ULIS 与 FTU。"),
        (45, "已确认", "至少1个", "https://en.ulis.vnu.edu.vn/blog/archives/ulis-students-won-prizes-at-20th-chinese-bridge-competition/", "ULIS 页面提及 Confucius Institute – Hanoi University。"),
        (64, "部分确认", "存在较清晰培养与认证路径", "https://en.ulis.vnu.edu.vn/aboutulis-organizational-structure/faculty-of-chinese-language-and-culture/ulis-sunwah-chinese-teaching-and-research-centre-ulis-sunwah/", "认证支持路径较清晰。"),
    ],
}

def style_header(ws):
    fill = PatternFill('solid', fgColor='D9EAF7')
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

wb = load_workbook(SRC)
task_ws = wb['任务清单']
headers = [c.value for c in task_ws[1]]
if '当前值' not in headers:
    task_ws.cell(row=1, column=len(headers)+1, value='当前值')
    task_ws.cell(row=1, column=len(headers)+2, value='来源链接')
    task_ws.cell(row=1, column=len(headers)+3, value='进度备注')
    headers = [c.value for c in task_ws[1]]
style_header(task_ws)
value_col = headers.index('当前值') + 1
source_col = headers.index('来源链接') + 1
note_col = headers.index('进度备注') + 1
status_col = headers.index('状态') + 1
country_col = headers.index('国家') + 1
metric_col = headers.index('指标ID') + 1

filled = defaultdict(lambda: {'已确认': 0, '部分确认': 0})
for row in range(2, task_ws.max_row + 1):
    country = task_ws.cell(row=row, column=country_col).value
    metric_id = task_ws.cell(row=row, column=metric_col).value
    if country not in DATA:
        continue
    for mid, status, value, source, note in DATA[country]:
        if metric_id == mid:
            task_ws.cell(row=row, column=status_col, value=status)
            task_ws.cell(row=row, column=value_col, value=value)
            task_ws.cell(row=row, column=source_col, value=source)
            task_ws.cell(row=row, column=note_col, value=note)
            filled[country][status] += 1
            break
for row in task_ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=True)
for col, width in {'Q': 28, 'R': 72, 'S': 40}.items():
    task_ws.column_dimensions[col].width = width

if '五国进度汇总' in wb.sheetnames:
    del wb['五国进度汇总']
sum_ws = wb.create_sheet('五国进度汇总')
sum_ws.append(['国家','已确认','部分确认','合计已启动','总体判断'])
style_header(sum_ws)
judgement = {
    '新加坡': '高制度化样本',
    '马来西亚': '社会基础最强样本',
    '泰国': '外语扩张型样本',
    '印尼': '成长型样本',
    '越南': '增长型样本',
}
for country in ['新加坡','马来西亚','泰国','印尼','越南']:
    yes = filled[country]['已确认']
    partial = filled[country]['部分确认']
    sum_ws.append([country, yes, partial, yes + partial, judgement[country]])
for row in sum_ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=True)
for col, width in {'A': 12, 'B': 10, 'C': 10, 'D': 12, 'E': 24}.items():
    sum_ws.column_dimensions[col].width = width

if '五国已填样本' in wb.sheetnames:
    del wb['五国已填样本']
sample_ws = wb.create_sheet('五国已填样本')
sample_ws.append(['国家','指标ID','字段','状态','当前值','来源链接','备注'])
style_header(sample_ws)
metric_name_map = {}
for row in range(2, task_ws.max_row + 1):
    metric_name_map[(task_ws.cell(row=row, column=country_col).value, task_ws.cell(row=row, column=metric_col).value)] = task_ws.cell(row=row, column=headers.index('字段名') + 1).value
for country, rows in DATA.items():
    for mid, status, value, source, note in rows:
        sample_ws.append([country, mid, metric_name_map.get((country, mid), ''), status, value, source, note])
for row in sample_ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=True)
for col, width in {'A': 12, 'B': 8, 'C': 24, 'D': 12, 'E': 30, 'F': 72, 'G': 40}.items():
    sample_ws.column_dimensions[col].width = width

wb.save(OUT)
print(OUT)
