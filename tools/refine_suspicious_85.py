"""精修 85 格可疑 0/无：扩展口径 + 广义定义 + 2024-2025 最新数据检索。

策略：
1. 用户指定 4 格 → 硬编码扩展口径提示
2. 其余 81 格 → LLM 自动生成扩展口径（"广义定义、任何相关内容都算、最新年份"）
3. 每格 4 轮搜索（Perplexity + 3 路 Tavily：zh/en/2025 新闻）+ 5 源学术
4. DS 严格合成：绝不回退到 0 或 '无'
"""
from __future__ import annotations
import asyncio, io, json, os, sys, logging, time
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "tools"))
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError: pass

import httpx, openpyxl
from openpyxl.styles import PatternFill, Alignment
from news_monitor.proxy import build_httpx_client, get_proxies_for_url
from academic_search import academic_search, format_results

TOPICS = ROOT / "topics"
DATE = datetime.now().strftime("%Y%m%d")
SRC = TOPICS / "“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v10_20260421.xlsx"
V11 = TOPICS / f"“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v11_{DATE}.xlsx"
LOG_MD = TOPICS / f"bri_suspicious85_{DATE}.md"
CKPT = ROOT / "output" / "bri_suspicious85_ckpt.json"
TARGETS_FILE = ROOT / "output" / "bri_suspicious_targets.json"

PPX = os.environ.get("PERPLEXITY_API_KEY", "")
TAV = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
DS = os.environ.get("DEEPSEEK_API_KEY", "")
PROXY = os.environ.get("LLM_PROXY", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("refine85")

COL = {"蒙古":(5,6),"越南":(7,8),"柬埔寨":(9,10),"菲律宾":(11,12),"新加坡":(13,14)}
COUNTRY_EN = {"蒙古":"Mongolia","越南":"Vietnam","柬埔寨":"Cambodia","菲律宾":"Philippines","新加坡":"Singapore"}

# ── 用户手动标注的 4 格定制扩展口径 ─────────────────────────────────────────
USER_HINTS = {
    ("新加坡", 47): {
        "extended_definition": (
            "【扩展口径】新加坡华校不仅指传统华文学校，还应包括：\n"
            "1. 新型华文教育机构：新跃社科大学中华学术中心（SUSS CSC）、SCCL、华文教研中心\n"
            "2. 高校华文系/中国研究系：NUS Chinese Studies、NTU Chinese Programme、SUSS BA Chinese\n"
            "3. 双语幼儿园/学前中心：如 MindChamps Chinese、Tien Hsia、Berries World of Learning 等连锁品牌\n"
            "4. 在线华文教育服务（总部在新加坡的企业）：Sprout、Hello Chinese、ChineseLingua、"
            "Tientian Chinese、SgHaoXue 等\n"
            "5. 补习中心（Tuition Centre）开设华文专业：Learning Point、The Learning Lab 等\n"
            "合计这几类，估算 20-50 家。"
        ),
        "queries": {
            "perplexity": "新加坡以下几类华文教育机构各有多少家：(1) 高校华文系/中国研究系；(2) 双语幼儿园连锁；(3) 新加坡本土的在线中文学习平台/APP 企业（如 Hello Chinese、Sprout）；(4) 华文补习中心？列出具体机构名。",
            "tavily_zh": "新加坡 华文教育 机构 双语幼儿园 在线中文 学习平台 MindChamps",
            "tavily_en": "Singapore Chinese language education center bilingual kindergarten online learning platform",
            "academic": "Singapore Chinese language education institutions platforms",
        },
        "domains": ["moe.gov.sg", "suss.edu.sg", "sccl.sg", "ntu.edu.sg", "nus.edu.sg", "mindchamps.org"],
    },
    ("菲律宾", 43): {
        "extended_definition": (
            "【扩展口径】不仅看 2019 年老报告，要查 2024-2025 年：\n"
            "TOP50 菲律宾高校中有哪些开设中文专业/中文选修课/汉语+专业双学位？\n"
            "已知：Ateneo de Manila 有 Chinese Studies 程序、UP Diliman 有 Linguistics 含中文、"
            "DLSU、UST 等部分有。另有约 4 所孔子学院合作高校：布拉干州立大学、Angeles 大学基金会、"
            "Bulacan State University 等。\n"
            "分母：菲律宾 CHED 承认的大学约 1,800 所（广义），TOP50 更合理作为分母。"
        ),
        "queries": {
            "perplexity": "截至 2025 年，菲律宾 TOP50 大学中有多少所开设中文专业或中文系（含中文选修、中国研究）？列出具体高校名单（如 Ateneo、UP、DLSU、UST 等）。",
            "tavily_zh": "菲律宾 高校 中文专业 中国研究 Ateneo UP DLSU 2025",
            "tavily_en": "Philippines universities Chinese Studies program 2024 2025 Ateneo UP DLSU",
            "academic": "Philippines Chinese language program university 2024",
        },
        "domains": ["ched.gov.ph", "ateneo.edu", "up.edu.ph", "dlsu.edu.ph", "ust.edu.ph", "ci.cn"],
    },
    ("柬埔寨", 31): {
        "extended_definition": (
            "【扩展口径】不用 2019 老数据，查 2023-2025 最新：\n"
            "(a) 2021 柬埔寨 MOEYS 正式将中文纳入高中第二外语，2023-2024 全国覆盖数扩大；\n"
            "(b) 柬华理事总会 2024 报告：全国华文学校 ~80 所、学生 ~5 万人；\n"
            "(c) 柬埔寨 2023 基础教育在校生约 3,300,000；\n"
            "比例 ≈ 50,000 / 3,300,000 ≈ 1.52%。\n"
            "也可取近 5 年年增长率（5%-15% 区间）。"
        ),
        "queries": {
            "perplexity": "2023-2025 年柬埔寨中文教育最新数据：华文学校/公立中学中文学习者总数、占基础教育在校生比例、近 5 年年增长率。",
            "tavily_zh": "柬埔寨 2024 2025 中文教育 华文学校 学生人数 MOEYS 最新",
            "tavily_en": "Cambodia Chinese education students 2024 2025 MOEYS enrollment latest",
            "academic": "Cambodia Chinese language education enrollment 2024 2025",
        },
        "domains": ["moeys.gov.kh", "chinese.cn", "ci.cn", "rupp.edu.kh"],
    },
    ("越南", 49): {
        "extended_definition": (
            "【扩展口径】偏远地区不仅指孔子学院，应包括：\n"
            "(a) 偏远省份/城市的华文学校（如广南、广宁、老街、奠边等省）；\n"
            "(b) 私立中文补习中心/培训班（Trung tâm tiếng Trung）；\n"
            "(c) 高校外语系的中文选修课程（在非河内/胡志明的地方大学）；\n"
            "(d) 在线中文课程在偏远省的覆盖率；\n"
            "(e) Facebook 上越南偏远地区的中文培训班广告/招生宣传。\n"
            "从 0% 扩展到 15-30% 区间。"
        ),
        "queries": {
            "perplexity": "越南偏远省份（广南、广宁、老街、奠边等非河内/胡志明地区）有哪些华文学校、中文培训中心、高校中文选修课？包括宣传招生信息。",
            "tavily_zh": "越南 偏远省份 中文培训 Trung tâm tiếng Trung 广南 广宁 老街",
            "tavily_en": "Vietnam rural provinces Chinese language training center private",
            "academic": "Vietnam rural Chinese language education provinces",
        },
        "domains": ["moet.gov.vn", "chinese.cn", "ci.cn", "facebook.com"],
    },
}

# ── API ─────────────────────────────────────────────────────────────────────
async def ppx_ask(client, q, system=None):
    if not PPX: return "", []
    sys_msg = system or (
        "你是代理研究员。对定量指标，绝不返回 0 或'无'，至少给出定性量级（'有限'/'若干'/'多'）。"
        "对比例指标，给区间估值。所有答案必须带 URL。"
    )
    try:
        r = await client.post(
            "https://api.perplexity.ai/chat/completions",
            json={"model":"sonar","messages":[
                {"role":"system","content":sys_msg},
                {"role":"user","content":q}],
                "temperature":0.1,"return_citations":True},
            headers={"Authorization":f"Bearer {PPX}","Content-Type":"application/json"},
            timeout=90.0,
        )
        r.raise_for_status()
        d = r.json()
        return d["choices"][0]["message"]["content"], d.get("citations",[])[:10]
    except Exception as e:
        log.warning("ppx err: %s", str(e)[:100])
        return "", []

async def tav(client, q, domains=None):
    if not TAV: return ([], "")
    try:
        body = {"api_key":TAV,"query":q,"search_depth":"advanced","max_results":8,
                "include_answer":True,"include_raw_content":False}
        if domains: body["include_domains"] = domains[:15]
        r = await client.post("https://api.tavily.com/search", json=body, timeout=60.0)
        r.raise_for_status()
        d = r.json()
        items = [{"url":x.get("url",""),"title":x.get("title",""),
                 "snippet":(x.get("content") or "")[:400]}
                for x in (d.get("results") or [])[:8]]
        return (items, d.get("answer","") or "")
    except Exception as e:
        log.warning("tav err: %s", str(e)[:100])
        return ([], "")

# ── Auto-generate extended definition for 81 non-user-flagged cells ────────
async def gen_extended_hint(client, t) -> dict:
    """用 DS 生成扩展口径提示 + 重新搜索查询."""
    prompt = f"""你是研究方法专家。一个定量指标被错误地填成了 0 或'无'，因为之前用了过度狭义的定义/过时数据。
请设计【扩展口径】，让这个指标能用合理的广义定义拿到非零值。

## 指标
- 国家：{t['country']}
- 字段：{t['name']}
- 原取值口径：{t['type']}
- 字段说明：{t['desc']}
- 当前错误值：{t['current']}

## 任务
1. 诊断为什么会搜不到数据（定义太严？年份太老？只看单一渠道？）
2. 给出扩展后的定义（广义、含多种形式、近年最新）
3. 写 3 个检索查询（中文、英文、2025 最新）

返回 JSON:
{{
  "diagnosis": "30 字：为什么之前搜不到",
  "extended_definition": "扩展口径说明，80 字内，列出应包括的所有形式",
  "queries": {{
    "perplexity": "给 Perplexity 的具体提问",
    "tavily_zh": "中文 Tavily 查询",
    "tavily_en": "英文 Tavily 查询",
    "academic": "学术数据库查询"
  }},
  "priority_domains": ["官方/权威域名1","域名2"]
}}
"""
    try:
        r = await client.post(
            "https://api.deepseek.com/v1/chat/completions",
            json={"model":"deepseek-chat","messages":[{"role":"user","content":prompt}],
                  "temperature":0.1,"response_format":{"type":"json_object"}},
            headers={"Authorization":f"Bearer {DS}","Content-Type":"application/json"},
            timeout=60.0,
        )
        r.raise_for_status()
        return json.loads(r.json()["choices"][0]["message"]["content"])
    except Exception as e:
        log.warning("gen_hint err %s#%d: %s", t['country'], t['id'], str(e)[:100])
        return {}

async def ds_synth(client, t, hint_info, evidence):
    prompt = f"""你是数据研究员。在【扩展口径】下重新给【{t['country']}】【{t['name']}】合成值。
当前被错误填成的值：{t['current']}

## 扩展口径
{hint_info['extended_definition']}

## 取值口径
{t['type']}

## 诊断
{hint_info.get('diagnosis','')}

## 证据
{evidence}

## 任务
绝对不要再返回 0 / '无' / '合成失败'。
- 如果找到了具体数字/清单，按扩展口径给出数字（带分子分母）
- 如果仍无精确数据，给区间估值（如 "5-15" / "10-30%"）
- 如果真无量化证据，给定性分级（"少量 / 若干 / 多"）+ 推理依据

返回 JSON:
{{
  "value": "最终答案（贴合口径）",
  "extended_scope": "本次扩展了的定义边界，30 字",
  "derivation": "推导过程",
  "source_urls": ["URL1","URL2"],
  "confidence": "high/medium/low",
  "rationale": "30 字内"
}}
"""
    try:
        r = await client.post(
            "https://api.deepseek.com/v1/chat/completions",
            json={"model":"deepseek-chat","messages":[{"role":"user","content":prompt}],
                  "temperature":0.1,"response_format":{"type":"json_object"}},
            headers={"Authorization":f"Bearer {DS}","Content-Type":"application/json"},
            timeout=90.0,
        )
        r.raise_for_status()
        return json.loads(r.json()["choices"][0]["message"]["content"])
    except Exception as e:
        log.warning("synth err %s#%d: %s", t['country'], t['id'], str(e)[:100])
        return {"value":"合成失败","confidence":"low"}

async def solve(sem, t, clients, ckpt):
    key = f"{t['country']}::{t['id']}"
    if key in ckpt: return ckpt[key]
    async with sem:
        t0 = time.time()
        ppx_c, tav_c, ds_c = clients

        # Step 1: get extended hint
        uh = USER_HINTS.get((t['country'], t['id']))
        if uh:
            hint_info = {
                'extended_definition': uh['extended_definition'],
                'queries': uh['queries'],
                'priority_domains': uh.get('domains', []),
                'diagnosis': '用户手动标注需扩展口径',
            }
        else:
            hint_info = await gen_extended_hint(ds_c, t)
            if not hint_info.get('queries'):
                ckpt[key] = {"value":"扩展策略失败","confidence":"low"}
                _save(ckpt)
                return ckpt[key]

        queries = hint_info['queries']
        domains = hint_info.get('priority_domains', [])

        # Step 2: parallel search
        results = await asyncio.gather(
            ppx_ask(ppx_c, queries.get('perplexity','')),
            tav(tav_c, queries.get('tavily_zh',''), domains),
            tav(tav_c, queries.get('tavily_en',''), None),
            academic_search(queries.get('academic',''), max_per_source=3,
                            proxy_url=PROXY, include_scholar=True),
            return_exceptions=True,
        )
        ppx_res = results[0] if isinstance(results[0], tuple) else ("", [])
        tav_zh = results[1] if isinstance(results[1], tuple) and len(results[1])==2 else ([],"")
        tav_en = results[2] if isinstance(results[2], tuple) and len(results[2])==2 else ([],"")
        acad = results[3] if isinstance(results[3], dict) else {}

        evidence_parts = [f"## Perplexity\n{ppx_res[0]}\nCitations:\n" + "\n".join(ppx_res[1][:8])]
        evidence_parts.append(f"## Tavily zh answer: {tav_zh[1]}")
        for x in tav_zh[0][:5]:
            evidence_parts.append(f"  - {x['title']} | {x['url']}\n    {x['snippet'][:200]}")
        evidence_parts.append(f"## Tavily en answer: {tav_en[1]}")
        for x in tav_en[0][:5]:
            evidence_parts.append(f"  - {x['title']} | {x['url']}\n    {x['snippet'][:200]}")
        evidence_parts.append(f"## Academic\n{format_results(acad)[:4000]}")
        evidence = "\n\n".join(evidence_parts)[:15000]

        # Step 3: synth
        result = await ds_synth(ds_c, t, hint_info, evidence)
        result['extended_definition_applied'] = hint_info['extended_definition'][:200]
        result['diagnosis'] = hint_info.get('diagnosis','')
        ckpt[key] = result
        _save(ckpt)
        log.info("%s#%d %.1fs value=%s conf=%s",
                 t['country'], t['id'], time.time()-t0,
                 str(result.get('value',''))[:40], result.get('confidence','?'))
        return result

def _save(ckpt):
    with open(CKPT, "w", encoding="utf-8") as f:
        json.dump(ckpt, f, ensure_ascii=False, indent=1)

async def run():
    with open(TARGETS_FILE, encoding='utf-8') as f:
        targets = json.load(f)
    log.info("精修目标: %d 格", len(targets))

    ckpt = {}
    if CKPT.exists():
        try: ckpt = json.loads(CKPT.read_text(encoding='utf-8'))
        except Exception: pass
    log.info("ckpt: %d", len(ckpt))

    ppx_p = get_proxies_for_url("https://api.perplexity.ai", PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", PROXY)
    async with (build_httpx_client(ppx_p or "", 90) as ppx_c,
                build_httpx_client(tav_p or "", 60) as tav_c,
                build_httpx_client("", 90) as ds_c):
        sem = asyncio.Semaphore(4)
        tasks = [solve(sem, t, (ppx_c, tav_c, ds_c), ckpt) for t in targets]
        results = await asyncio.gather(*tasks)

    # write v11
    wb = openpyxl.load_workbook(SRC); ws = wb.active
    refine_fill = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid")
    log_md = [f"# 85 格可疑 0/无 精修 (v11)", f"时间：{datetime.now():%Y-%m-%d %H:%M}\n"]

    updated = 0
    for t, r in zip(targets, results):
        val = str(r.get('value','')).strip()
        if not val or val in ('合成失败','扩展策略失败',''): continue
        # 如果还是 0/无，说明真的无法救回，跳过
        import re
        if re.match(r'^(0|无|没有|0%|0人|0所|0项|0个|0次)$', val): continue
        updated += 1
        prefix = "【精修-广义】"
        scope = r.get('extended_scope','')
        deriv = r.get('derivation','')
        txt = f"{prefix}{val}"
        if scope: txt += f"（扩展：{scope}）"
        if deriv: txt += f"\n推导：{deriv}"
        vcol, scol = COL[t['country']]
        row = t['row']
        ws.cell(row, vcol, txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, vcol).fill = refine_fill
        urls = r.get('source_urls', [])
        if urls: ws.cell(row, scol, "; ".join(urls[:3]))

        log_md.append(f"\n## {t['country']}#{t['id']} {t['name']} ({t['priority']})")
        log_md.append(f"- **旧值**：{t['current'][:60]}")
        log_md.append(f"- **新值**：{val}")
        log_md.append(f"- **扩展**：{scope}")
        log_md.append(f"- **推导**：{deriv}")
        log_md.append(f"- **置信度**：{r.get('confidence','?')}")
        log_md.append(f"- **理由**：{r.get('rationale','')}")
        for u in (urls or [])[:3]: log_md.append(f"- URL: {u}")

    wb.save(V11)
    LOG_MD.write_text("\n".join(log_md), encoding="utf-8")
    log.info("saved %s, %s", V11.name, LOG_MD.name)
    log.info("DONE: %d/%d 成功精修", updated, len(targets))


if __name__ == "__main__":
    asyncio.run(run())
