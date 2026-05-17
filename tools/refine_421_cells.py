"""对 bri_421_suspicious_targets_{DATE}.json 中所有可疑格做统一精修。

流程（每格）：
  1. DeepSeek 生成扩展口径 + 4 路检索 query
  2. 并发跑 Perplexity + Tavily(zh/en) + academic(5 源)
  3. DeepSeek 合成 {value, proxy_used, derivation, source_urls, confidence, rationale}
  4. 按批次写回 xlsx 为 *_refined447_{DATE}.xlsx

断点续跑：output/bri_refine447_ckpt.json 按 {country}::{id} 索引

用法：
  python tools/refine_421_cells.py                # 跑全部
  python tools/refine_421_cells.py --limit 10     # 只跑前 10 格（测试）
  python tools/refine_421_cells.py --batch 3      # 只跑某批次
  python tools/refine_421_cells.py --writeback    # 不跑新任务，只把现有 ckpt 写回 xlsx
"""
from __future__ import annotations
import argparse, asyncio, io, json, logging, os, re, sys, time, shutil
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "tools"))
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError: pass

import httpx, openpyxl
from openpyxl.styles import Alignment, PatternFill
from news_monitor.proxy import build_httpx_client, get_proxies_for_url
from academic_search import academic_search, format_results

TOPICS = ROOT / "topics"
OUTPUT = ROOT / "output"
DATE = datetime.now().strftime("%Y%m%d")
TARGETS_JSON = OUTPUT / f"bri_421_suspicious_targets_{DATE}.json"
CKPT_FILE = OUTPUT / f"bri_refine447_ckpt.json"
RUN_LOG = OUTPUT / f"bri_refine447_run.log"
SUMMARY_MD = TOPICS / f"bri_refine447_summary_{DATE}.md"

PPX = os.environ.get("PERPLEXITY_API_KEY", "")
TAV = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
DS = os.environ.get("DEEPSEEK_API_KEY", "")
PROXY = os.environ.get("LLM_PROXY", "")

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(message)s",
                    datefmt="%H:%M:%S",
                    handlers=[logging.FileHandler(RUN_LOG, encoding='utf-8'),
                              logging.StreamHandler()])
log = logging.getLogger("r447")

# 国家 → 英文 / 官方域 / Brave CC
COUNTRY_META = {
    "蒙古": (["president.mn","mfa.gov.mn","moed.gov.mn"], "", "Mongolia"),
    "越南": (["mofa.gov.vn","moet.gov.vn","chinhphu.vn","vnu.edu.vn"], "", "Vietnam"),
    "柬埔寨": (["mfaic.gov.kh","moeys.gov.kh","rupp.edu.kh"], "", "Cambodia"),
    "菲律宾": (["deped.gov.ph","dfa.gov.ph","up.edu.ph"], "PH", "Philippines"),
    "新加坡": (["moe.gov.sg","mfa.gov.sg","gov.sg","nus.edu.sg"], "SG", "Singapore"),
    "马来西亚": (["moe.gov.my","kln.gov.my","gov.my","um.edu.my"], "MY", "Malaysia"),
    "泰国": (["moe.go.th","mfa.go.th","go.th","chula.ac.th"], "TH", "Thailand"),
    "印尼": (["kemdikbud.go.id","kemlu.go.id","go.id","ui.ac.id"], "ID", "Indonesia"),
    "老挝": (["moes.edu.la","mofa.gov.la","la","nuol.edu.la"], "", "Laos"),
    "缅甸": (["moe.gov.mm","mofa.gov.mm","mm","uy.edu.mm"], "", "Myanmar"),
    "文莱": (["mfa.gov.bn","moe.gov.bn","gov.bn","ubd.edu.bn"], "", "Brunei"),
    "东帝汶": (["mfac.gov.tl","gov.tl","unp.edu.tl"], "", "Timor-Leste"),
    "不丹": (["moe.gov.bt","mfa.gov.bt","gov.bt","rub.edu.bt"], "", "Bhutan"),
    "孟加拉国": (["moedu.gov.bd","mofa.gov.bd","gov.bd","du.ac.bd"], "", "Bangladesh"),
    "尼泊尔": (["moe.gov.np","mofa.gov.np","gov.np","tribhuvan-university.edu.np"], "", "Nepal"),
    "斯里兰卡": (["moe.gov.lk","mfa.gov.lk","gov.lk","cmb.ac.lk"], "", "Sri Lanka"),
    "印度": (["education.gov.in","mea.gov.in","gov.in","du.ac.in"], "IN", "India"),
    "巴基斯坦": (["mofept.gov.pk","mofa.gov.pk","gov.pk","qau.edu.pk"], "", "Pakistan"),
    "阿富汗": (["moe.gov.af","mfa.gov.af","gov.af"], "", "Afghanistan"),
    "马尔代夫": (["moe.gov.mv","foreign.gov.mv","gov.mv","mnu.edu.mv"], "", "Maldives"),
    "土库曼斯坦": (["edu.gov.tm","mfa.gov.tm","gov.tm"], "", "Turkmenistan"),
    "乌兹别克斯坦": (["edu.uz","mfa.uz","gov.uz","nuu.uz"], "", "Uzbekistan"),
    "哈萨克斯坦": (["edu.gov.kz","mfa.gov.kz","gov.kz","kaznu.kz"], "", "Kazakhstan"),
    "塔吉克斯坦": (["moa.tj","mfa.tj","gov.tj","tnu.tj"], "", "Tajikistan"),
    "吉尔吉斯斯坦": (["edu.gov.kg","mfa.gov.kg","gov.kg","kyrgyz-natgeo.edu.kg"], "", "Kyrgyzstan"),
    "阿曼": (["moe.gov.om","fm.gov.om","omanportal.gov.om","squ.edu.om"], "", "Oman"),
    "伊朗": (["medu.ir","mfa.ir","ir","ut.ac.ir"], "", "Iran"),
    "阿联酋": (["moe.gov.ae","mofaic.gov.ae","government.ae","uaeu.ac.ae"], "", "UAE"),
    "卡塔尔": (["edu.gov.qa","mofa.gov.qa","qu.edu.qa"], "", "Qatar"),
    "巴林": (["moe.gov.bh","mofa.gov.bh","uob.edu.bh"], "", "Bahrain"),
    "沙特阿拉伯": (["moe.gov.sa","mofa.gov.sa","gov.sa","ksu.edu.sa"], "SA", "Saudi Arabia"),
    "科威特": (["moe.edu.kw","mofa.gov.kw","kuniv.edu.kw"], "", "Kuwait"),
    "也门": (["yemen.gov.ye","mofa.gov.ye","ye"], "", "Yemen"),
    "土耳其": (["meb.gov.tr","mfa.gov.tr","turkiye.gov.tr","istanbul.edu.tr"], "TR", "Turkey"),
    "伊拉克": (["moedu.gov.iq","mofa.gov.iq","iq","uobaghdad.edu.iq"], "", "Iraq"),
    "叙利亚": (["syrianeducation.org.sy","mofaex.gov.sy","damascus-university.edu.sy"], "", "Syria"),
    "约旦": (["moe.gov.jo","mfa.gov.jo","jo","ju.edu.jo"], "", "Jordan"),
    "黎巴嫩": (["mehe.gov.lb","general-security.gov.lb","lb","ul.edu.lb"], "", "Lebanon"),
    "以色列": (["edu.gov.il","mfa.gov.il","gov.il","huji.ac.il"], "IL", "Israel"),
    "格鲁吉亚": (["mes.gov.ge","mfa.gov.ge","tsu.ge"], "", "Georgia"),
    "亚美尼亚": (["edu.am","mfa.am","ysu.am"], "", "Armenia"),
    "阿塞拜疆": (["edu.gov.az","mfa.gov.az","gov.az","bsu.edu.az"], "", "Azerbaijan"),
    "埃及": (["moe.gov.eg","mfa.gov.eg","egypt.gov.eg","cu.edu.eg"], "EG", "Egypt"),
    "爱沙尼亚": (["hm.ee","vm.ee","ee","ut.ee"], "EE", "Estonia"),
    "拉脱维亚": (["izm.gov.lv","mfa.gov.lv","lv","lu.lv"], "LV", "Latvia"),
    "立陶宛": (["smm.lrv.lt","urm.lt","lt","vu.lt"], "LT", "Lithuania"),
    "德国": (["bmbf.de","auswaertiges-amt.de","de","uni-heidelberg.de"], "DE", "Germany"),
    "波兰": (["gov.pl","msz.gov.pl","pl","uw.edu.pl"], "PL", "Poland"),
    "捷克": (["msmt.cz","mzv.cz","cz","cuni.cz"], "CZ", "Czech Republic"),
    "斯洛文尼亚": (["mvzi.gov.si","mzz.gov.si","si","uni-lj.si"], "SI", "Slovenia"),
    "克罗地亚": (["mzo.gov.hr","mvep.gov.hr","hr","unizg.hr"], "HR", "Croatia"),
    "塞尔维亚": (["mpn.gov.rs","mfa.gov.rs","srbija.gov.rs","bg.ac.rs"], "RS", "Serbia"),
    "波黑": (["mcp.gov.ba","mvp.gov.ba","ba","unsa.ba"], "BA", "Bosnia and Herzegovina"),
    "黑山": (["mpin.gov.me","mvp.gov.me","me","ucg.ac.me"], "ME", "Montenegro"),
    "阿尔巴尼亚": (["arsimi.gov.al","punetejashtme.gov.al","al","unitir.edu.al"], "AL", "Albania"),
    "匈牙利": (["kormany.hu","mfa.gov.hu","hu","elte.hu"], "HU", "Hungary"),
    "斯洛伐克": (["minedu.sk","mzv.sk","slovensko.sk","uniba.sk"], "SK", "Slovakia"),
    "北马其顿": (["mon.gov.mk","mfa.gov.mk","mk","ukim.edu.mk"], "MK", "North Macedonia"),
    "罗马尼亚": (["edu.ro","mae.ro","gov.ro","unibuc.ro"], "RO", "Romania"),
    "保加利亚": (["mon.bg","mfa.bg","bg","uni-sofia.bg"], "BG", "Bulgaria"),
    "摩尔多瓦": (["mecc.gov.md","mfa.gov.md","gov.md","usm.md"], "MD", "Moldova"),
    "乌克兰": (["mon.gov.ua","mfa.gov.ua","gov.ua","knu.ua"], "UA", "Ukraine"),
    "白俄罗斯": (["edu.gov.by","mfa.gov.by","belarus.by","bsu.by"], "BY", "Belarus"),
    "俄罗斯": (["edu.gov.ru","mid.ru","gov.ru","msu.ru","spbu.ru"], "RU", "Russia"),
}
CHINA_SIDE = ["hanban.org","chinese.cn","ci.cn","mfa.gov.cn","fmprc.gov.cn",
              "moe.gov.cn","edu.cn","xinhuanet.com","chinanews.com","people.com.cn"]


# ── API helpers ─────────────────────────────────────────────────────────────
async def ppx_ask(client, q, sys_msg=None):
    if not PPX or not q.strip(): return "", []
    msg = sys_msg or "你是研究员。给事实+URL。若无直接数据给代理推导值，绝不返回'无可核验数据'。"
    try:
        r = await client.post("https://api.perplexity.ai/chat/completions",
            json={"model":"sonar","messages":[
                {"role":"system","content":msg},
                {"role":"user","content":q}],
                "temperature":0.1,"return_citations":True},
            headers={"Authorization":f"Bearer {PPX}","Content-Type":"application/json"},
            timeout=90.0)
        r.raise_for_status()
        d = r.json()
        return d["choices"][0]["message"]["content"], (d.get("citations",[]) or [])[:10]
    except Exception as e:
        log.warning("ppx: %s", str(e)[:80]); return "", []


async def tav(client, q, domains=None):
    if not TAV or not q.strip(): return ([], "")
    try:
        body = {"api_key":TAV,"query":q,"search_depth":"advanced","max_results":8,
                "include_answer":True,"include_raw_content":False}
        if domains: body["include_domains"] = domains[:15]
        r = await client.post("https://api.tavily.com/search", json=body, timeout=60.0)
        r.raise_for_status()
        d = r.json()
        items = [{"url":x.get("url",""),"title":x.get("title",""),
                 "snippet":(x.get("content") or "")[:400]}
                for x in (d.get("results") or [])[:8]]
        return (items, d.get("answer","") or "")
    except Exception as e:
        log.warning("tav: %s", str(e)[:80]); return ([], "")


async def ds_call(client, prompt, json_mode=True):
    if not DS: return {}
    try:
        payload = {"model":"deepseek-chat","messages":[{"role":"user","content":prompt}],
                   "temperature":0.1}
        if json_mode: payload["response_format"] = {"type":"json_object"}
        r = await client.post("https://api.deepseek.com/v1/chat/completions",
            json=payload,
            headers={"Authorization":f"Bearer {DS}","Content-Type":"application/json"},
            timeout=120.0)
        r.raise_for_status()
        content = r.json()["choices"][0]["message"]["content"]
        return json.loads(content) if json_mode else content
    except Exception as e:
        log.warning("ds: %s", str(e)[:80]); return {}


def _persist_ckpt(ckpt, path):
    path.parent.mkdir(exist_ok=True)
    tmp = path.with_suffix(path.suffix + '.tmp')
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(ckpt, f, ensure_ascii=False)
    tmp.replace(path)


async def refine_one(ppx_c, tav_c, ds_c, g, ckpt, ckpt_lock):
    """处理单格。返回 synth dict。"""
    key = f"{g['country']}::{g['id']}"
    async with ckpt_lock:
        if key in ckpt: return ckpt[key]

    t0 = time.time()
    meta = COUNTRY_META.get(g['country'], ([], "", g['country']))
    official = meta[0] + CHINA_SIDE
    en = meta[2]

    plan = await ds_call(ds_c, f"""为该"搜不到确切数据"的中文教育指标设计扩展代理方案。
国家：{g['country']}（{en}）
字段：{g['name']}
取值口径：{g['type']}
字段说明：{g['desc']}
字段分类：{g['category']}
优先级：{g['priority']}
当前可疑值（需替换）：{g['current'][:120]}
问题类型：{g['kind']}（empty=空 / zero=值0 / wu=值"无" / soft=少量若干有限 / bad=无可核验）

要求：
1. extended_definition: 3-5 行扩展口径，指示把相关内容都算在内（类似概念、广义定义、近 3 年最新数据）
2. queries: 4 路具体检索 query
   - perplexity: 综合性中文提问
   - tavily_zh: 中文关键词+国家名，精准检索
   - tavily_en: 英文关键词+国家英文名
   - academic: 学术检索 query（中英文均可）
3. priority_domains: 3-5 个优先域名（.gov / 教育部 / 学术库）

返回 JSON: {{"extended_definition":"", "queries": {{"perplexity":"", "tavily_zh":"", "tavily_en":"", "academic":""}}, "priority_domains": []}}""")

    if not plan.get('queries'):
        result = {"value": "无可用策略", "confidence": "low", "_elapsed": time.time()-t0}
        async with ckpt_lock:
            ckpt[key] = result
            _persist_ckpt(ckpt, CKPT_FILE)
        return result

    q = plan['queries']
    doms = plan.get('priority_domains', [])

    results = await asyncio.gather(
        ppx_ask(ppx_c, q.get('perplexity','')),
        tav(tav_c, q.get('tavily_zh',''), doms),
        tav(tav_c, q.get('tavily_en',''), official),
        academic_search(q.get('academic',''), max_per_source=3, proxy_url=PROXY, include_scholar=False),
        return_exceptions=True)
    ppx_r = results[0] if isinstance(results[0], tuple) else ("", [])
    tav_zh = results[1] if isinstance(results[1], tuple) and len(results[1])==2 else ([],"")
    tav_en = results[2] if isinstance(results[2], tuple) and len(results[2])==2 else ([],"")
    acad = results[3] if isinstance(results[3], dict) else {}

    evid = f"""## Perplexity
{ppx_r[0]}
Citations: {chr(10).join(ppx_r[1][:6])}
## Tavily-zh answer: {tav_zh[1]}
{chr(10).join(f'  - {x["title"]} | {x["url"]}' for x in tav_zh[0][:4])}
## Tavily-en answer: {tav_en[1]}
{chr(10).join(f'  - {x["title"]} | {x["url"]}' for x in tav_en[0][:4])}
## Academic
{format_results(acad)[:3500]}"""[:12000]

    synth = await ds_call(ds_c, f"""根据扩展口径+证据给【{g['country']}】【{g['name']}】最终值。
扩展口径: {plan.get('extended_definition','')}
取值口径: {g['type']}
当前可疑值: {g['current'][:120]}

证据:
{evid}

硬约束：
- 绝不返回 0/无/合成失败/少量/若干/有限/无可核验
- 有数字给数字，无数字给区间，都没有给定性分级+依据
- 若为分类字段（如是/否、风险高/中/低），明确给出分类标签
- derivation 简述推导逻辑（代理变量→目标值映射）

返回 JSON: {{"value":"", "proxy_used":"", "derivation":"", "source_urls":[], "confidence":"high|medium|low", "is_estimation": true|false, "rationale":""}}""")

    synth["_elapsed"] = time.time() - t0
    async with ckpt_lock:
        ckpt[key] = synth
        _persist_ckpt(ckpt, CKPT_FILE)

    log.info("[%s #%d] %.1fs value=%s", g['country'], g['id'],
             synth["_elapsed"], str(synth.get('value',''))[:40])
    return synth


async def run_refine(targets, concurrency=4):
    ckpt = {}
    if CKPT_FILE.exists():
        try: ckpt = json.loads(CKPT_FILE.read_text(encoding='utf-8'))
        except: pass
    log.info("ckpt 已有 %d 格，待处理 %d 格", len(ckpt), len(targets) - len([t for t in targets if f"{t['country']}::{t['id']}" in ckpt]))

    ppx_p = get_proxies_for_url("https://api.perplexity.ai", PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", PROXY)
    async with (build_httpx_client(ppx_p or "", 90) as ppx_c,
                build_httpx_client(tav_p or "", 60) as tav_c,
                build_httpx_client("", 120) as ds_c):
        sem = asyncio.Semaphore(concurrency)
        ckpt_lock = asyncio.Lock()

        async def _task(g):
            async with sem:
                try:
                    return await refine_one(ppx_c, tav_c, ds_c, g, ckpt, ckpt_lock)
                except Exception as e:
                    log.error("[%s #%d] err: %s", g['country'], g['id'], e)
                    return {"value":"合成失败","error":str(e)[:200]}

        tasks = [_task(g) for g in targets]
        await asyncio.gather(*tasks, return_exceptions=True)
    return ckpt


def writeback(targets, ckpt):
    """把 ckpt 中的值写回各批次 xlsx。"""
    # 按文件分组
    by_file = {}
    for t in targets:
        by_file.setdefault(t['file'], []).append(t)

    summary = []
    for fname, ts in by_file.items():
        src = TOPICS / fname
        # 命名规则: {基名}__refined447_{DATE}.xlsx
        stem = src.stem
        dst = TOPICS / f"{stem}__refined447_{DATE}.xlsx"
        if not dst.exists():
            shutil.copy(src, dst)

        wb = openpyxl.load_workbook(dst)
        ws = wb.active
        fill = PatternFill(start_color="B3E5FC", fill_type="solid")  # 浅蓝
        updated = 0; skipped = 0
        for t in ts:
            key = f"{t['country']}::{t['id']}"
            r = ckpt.get(key)
            if not r: skipped += 1; continue
            val = str(r.get('value','')).strip()
            if not val: skipped += 1; continue
            if val in ('合成失败','无可用策略'): skipped += 1; continue
            # 0/无/有限 等若带扎实 derivation，仍写回（这也是经过精修验证的结果）
            if re.match(r'^(0|无|没有|0%|少量|若干|有限|无可核验|无可靠数据)$', val):
                if not (r.get('derivation','').strip() and r.get('rationale','').strip()):
                    skipped += 1; continue

            proxy = r.get('proxy_used','')
            deriv = r.get('derivation','')
            prefix = "【精修447】"
            txt = f"{prefix}{val}"
            if proxy: txt += f"（代理：{proxy}）"
            if deriv: txt += f"\n推导：{deriv[:200]}"
            cell = ws.cell(t['row'], t['vcol'], txt)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = fill
            urls = r.get('source_urls', [])
            if urls:
                if isinstance(urls, str): urls = [urls]
                ws.cell(t['row'], t['vcol']+1,
                        "; ".join(str(u) for u in urls[:3])[:500])
            updated += 1
        wb.save(dst)
        log.info("%s: %d 更新 / %d 跳过 → %s", fname[:40], updated, skipped, dst.name)
        summary.append({'file': fname, 'total': len(ts), 'updated': updated, 'skipped': skipped, 'out': dst.name})

    # 写 summary MD
    lines = [f"# BRI 447 可疑格精修回写结果",
             f"",
             f"**完成时间**：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
             f"**涉及批次**：{len(summary)}",
             f"",
             f"| 批次文件 | 目标 | 已更新 | 跳过 | 产出 |",
             f"|---|---|---|---|---|"]
    for s in summary:
        lines.append(f"| {s['file'][:40]}... | {s['total']} | {s['updated']} | {s['skipped']} | {s['out']} |")
    SUMMARY_MD.write_text('\n'.join(lines), encoding='utf-8')
    log.info("summary written to %s", SUMMARY_MD)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0, help="只跑前 N 格（测试用）")
    ap.add_argument("--batch", type=int, default=0, help="只跑指定批次")
    ap.add_argument("--concurrency", type=int, default=4)
    ap.add_argument("--writeback", action="store_true", help="不跑新任务，只把 ckpt 写回")
    args = ap.parse_args()

    with open(TARGETS_JSON, encoding='utf-8') as f:
        targets = json.load(f)
    if args.batch:
        targets = [t for t in targets if t['batch'] == args.batch]
    if args.limit:
        targets = targets[:args.limit]
    log.info("目标总数: %d", len(targets))

    if args.writeback:
        ckpt = {}
        if CKPT_FILE.exists():
            ckpt = json.loads(CKPT_FILE.read_text(encoding='utf-8'))
        with open(TARGETS_JSON, encoding='utf-8') as f:
            all_targets = json.load(f)
        writeback(all_targets, ckpt)
        return

    ckpt = asyncio.run(run_refine(targets, concurrency=args.concurrency))
    # 自动写回
    with open(TARGETS_JSON, encoding='utf-8') as f:
        all_targets = json.load(f)
    writeback(all_targets, ckpt)


if __name__ == "__main__":
    main()
