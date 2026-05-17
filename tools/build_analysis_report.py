"""生成 BRI 64 国对比分析报告（MD + 图表）。

产出：
  topics/bri_analysis_report_{DATE}.md
  topics/bri_charts/*.png (≈6 张)
"""
from __future__ import annotations
import io, json, re, sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
TOPICS = ROOT / "topics"
DATE = datetime.now().strftime("%Y%m%d")
MERGED = TOPICS / f"bri_64country_merged_{DATE}.xlsx"
CKPT = ROOT / "output" / "bri_refine447_ckpt.json"
FIELD_DICT = TOPICS / "bri_field_dict.json"
CHART_DIR = TOPICS / "bri_charts"
CHART_DIR.mkdir(exist_ok=True)
OUT_MD = TOPICS / f"bri_analysis_report_{DATE}.md"

import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm

# 中文字体
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei']
plt.rcParams['axes.unicode_minus'] = False


def classify_status(val: str) -> str:
    if not val: return 'empty'
    clean = re.sub(r'^(【[^】]+】\s*)+', '', val).strip()
    first = re.split(r'[（\(\n]', clean, maxsplit=1)[0].strip()
    first = re.split(r'[，,；;：:]', first, maxsplit=1)[0].strip()
    if first in ('0','无','没有','0%','少量','若干','有限','无可核验','无可靠数据'):
        return 'weak'
    if any(c.isdigit() for c in first):
        return 'numeric'
    if len(first) >= 2:
        return 'qualitative'
    return 'weak'


def main():
    with open(FIELD_DICT, encoding='utf-8') as f:
        fields = [x for x in json.load(f) if x.get('指标ID')]
    fld_by_id = {int(f['指标ID']): f for f in fields}

    ckpt = json.loads(CKPT.read_text(encoding='utf-8')) if CKPT.exists() else {}

    wb = openpyxl.load_workbook(MERGED, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    country_cols = {}
    for c_idx, h in enumerate(headers, 1):
        if h and not str(h).endswith('_出处') and h not in ('序号','搜集字段','字段类型','字段说明','分类','优先级'):
            country_cols[h] = c_idx

    indicator_rows = {}
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, 1).value
        if iid is None: continue
        try: iid = int(iid)
        except: continue
        indicator_rows[iid] = r

    # 按国家统计
    country_stats = {}
    for country, vcol in country_cols.items():
        stat = {'numeric':0,'qualitative':0,'weak':0,'empty':0}
        for iid, r in indicator_rows.items():
            v = ws.cell(r, vcol).value
            stat[classify_status(str(v) if v else '')] += 1
        stat['total'] = sum(stat.values())
        stat['solid'] = stat['numeric'] + stat['qualitative']
        stat['solid_pct'] = stat['solid'] * 100 / stat['total'] if stat['total'] else 0
        country_stats[country] = stat

    # 按指标统计
    ind_stats = {}
    for iid, r in indicator_rows.items():
        stat = {'numeric':0,'qualitative':0,'weak':0,'empty':0}
        for country, vcol in country_cols.items():
            v = ws.cell(r, vcol).value
            stat[classify_status(str(v) if v else '')] += 1
        stat['total'] = sum(stat.values())
        stat['solid_pct'] = (stat['numeric']+stat['qualitative'])*100/stat['total'] if stat['total'] else 0
        ind_stats[iid] = stat

    # 精修 ckpt 分析
    ref_conf = Counter()
    ref_numeric = 0
    for k, v in ckpt.items():
        ref_conf[v.get('confidence','')] += 1
        val = str(v.get('value','')).strip()
        if val not in ('0','无','没有','0%','少量','若干','有限','无可核验','无可靠数据','合成失败','无可用策略') and any(c.isdigit() for c in val):
            ref_numeric += 1

    # ── 图表 1: Top20 国数据质量排名（堆叠柱状） ────────────────────────────
    top20 = sorted(country_stats.items(), key=lambda x: -x[1]['solid'])[:20]
    names = [x[0] for x in top20]
    numeric = [x[1]['numeric'] for x in top20]
    qualit = [x[1]['qualitative'] for x in top20]
    weak = [x[1]['weak'] for x in top20]

    fig, ax = plt.subplots(figsize=(14, 7))
    x = range(len(names))
    ax.bar(x, numeric, label='数字值', color='#2E7D32')
    ax.bar(x, qualit, bottom=numeric, label='定性值', color='#1976D2')
    ax.bar(x, weak, bottom=[a+b for a,b in zip(numeric, qualit)],
           label='弱数据(0/无/有限)', color='#FFA000')
    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=45, ha='right')
    ax.set_ylabel('指标数（68 总）')
    ax.set_title('Top 20 国数据质量排名 · 堆叠构成', fontsize=14)
    ax.legend()
    ax.grid(True, alpha=0.3, axis='y')
    plt.tight_layout()
    f1 = CHART_DIR / 'top20_quality_stack.png'
    plt.savefig(f1, dpi=110); plt.close()
    print(f"图 1: {f1.name}")

    # ── 图表 2: 64 国数据完整度排名（数字+定性 占比） ─────────────────────
    all_ranked = sorted(country_stats.items(), key=lambda x: -x[1]['solid_pct'])
    names_all = [x[0] for x in all_ranked]
    pcts = [x[1]['solid_pct'] for x in all_ranked]

    fig, ax = plt.subplots(figsize=(15, 12))
    colors = ['#2E7D32' if p >= 90 else '#66BB6A' if p >= 80 else
              '#FFA000' if p >= 70 else '#E53935' for p in pcts]
    ax.barh(range(len(names_all)), pcts, color=colors)
    ax.set_yticks(range(len(names_all)))
    ax.set_yticklabels(names_all, fontsize=8)
    ax.invert_yaxis()
    ax.set_xlabel('扎实数据占比 (%)')
    ax.set_title('64 国数据扎实度排名（数字 + 定性 占 68 指标之比）', fontsize=13)
    ax.axvline(90, color='green', linestyle='--', alpha=0.5, label='90% 线')
    ax.axvline(80, color='orange', linestyle='--', alpha=0.5, label='80% 线')
    ax.grid(True, alpha=0.3, axis='x')
    ax.legend()
    plt.tight_layout()
    f2 = CHART_DIR / 'all64_solid_rank.png'
    plt.savefig(f2, dpi=110); plt.close()
    print(f"图 2: {f2.name}")

    # ── 图表 3: 指标分类覆盖率（按 A-G 分组）─────────────────────────────
    cat_stats = defaultdict(lambda: {'numeric':0,'qualitative':0,'weak':0,'empty':0,'total':0})
    for iid, stat in ind_stats.items():
        cat = fld_by_id[iid].get('分类','?')
        for k in ('numeric','qualitative','weak','empty','total'):
            cat_stats[cat][k] += stat.get(k, 0)

    cats = sorted(cat_stats.keys())
    c_num = [cat_stats[c]['numeric'] for c in cats]
    c_qual = [cat_stats[c]['qualitative'] for c in cats]
    c_weak = [cat_stats[c]['weak'] for c in cats]
    c_emp = [cat_stats[c]['empty'] for c in cats]

    fig, ax = plt.subplots(figsize=(13, 6))
    x = range(len(cats))
    ax.bar(x, c_num, label='数字', color='#2E7D32')
    ax.bar(x, c_qual, bottom=c_num, label='定性', color='#1976D2')
    ax.bar(x, c_weak, bottom=[a+b for a,b in zip(c_num,c_qual)], label='弱', color='#FFA000')
    ax.bar(x, c_emp, bottom=[a+b+c for a,b,c in zip(c_num,c_qual,c_weak)], label='空', color='#E53935')
    ax.set_xticks(x)
    ax.set_xticklabels([c[:18] for c in cats], rotation=15, ha='right', fontsize=9)
    ax.set_ylabel('格子数（64 国）')
    ax.set_title('各分类下 64 国指标数据构成', fontsize=14)
    ax.legend()
    ax.grid(True, alpha=0.3, axis='y')
    plt.tight_layout()
    f3 = CHART_DIR / 'category_coverage.png'
    plt.savefig(f3, dpi=110); plt.close()
    print(f"图 3: {f3.name}")

    # ── 图表 4: 优先级覆盖率 ─────────────────────────────────────────
    prio_stats = defaultdict(lambda: {'numeric':0,'qualitative':0,'weak':0,'empty':0,'total':0})
    for iid, stat in ind_stats.items():
        p = fld_by_id[iid].get('优先级','?')
        for k in ('numeric','qualitative','weak','empty','total'):
            prio_stats[p][k] += stat.get(k, 0)

    prios = sorted(prio_stats.keys())
    p_num = [prio_stats[p]['numeric'] for p in prios]
    p_qual = [prio_stats[p]['qualitative'] for p in prios]
    p_weak = [prio_stats[p]['weak'] for p in prios]
    p_emp = [prio_stats[p]['empty'] for p in prios]

    fig, ax = plt.subplots(figsize=(8, 5.5))
    x = range(len(prios))
    ax.bar(x, p_num, label='数字', color='#2E7D32')
    ax.bar(x, p_qual, bottom=p_num, label='定性', color='#1976D2')
    ax.bar(x, p_weak, bottom=[a+b for a,b in zip(p_num,p_qual)], label='弱', color='#FFA000')
    ax.bar(x, p_emp, bottom=[a+b+c for a,b,c in zip(p_num,p_qual,p_weak)], label='空', color='#E53935')
    ax.set_xticks(x)
    ax.set_xticklabels(prios)
    ax.set_ylabel('格子数')
    ax.set_title('P1/P2/P3 优先级数据构成（64 国合计）', fontsize=13)
    ax.legend()
    ax.grid(True, alpha=0.3, axis='y')
    plt.tight_layout()
    f4 = CHART_DIR / 'priority_coverage.png'
    plt.savefig(f4, dpi=110); plt.close()
    print(f"图 4: {f4.name}")

    # ── 图表 5: 精修 447 置信度分布 ────────────────────────────────────
    conf_order = ['high','medium','low']
    conf_cnt = [ref_conf.get(c, 0) for c in conf_order]
    fig, ax = plt.subplots(figsize=(6, 5))
    colors5 = ['#2E7D32','#1976D2','#E53935']
    bars = ax.bar(conf_order, conf_cnt, color=colors5)
    for b, v in zip(bars, conf_cnt):
        ax.text(b.get_x()+b.get_width()/2, v+2, str(v), ha='center', fontsize=11)
    ax.set_ylabel('格数')
    ax.set_title(f'精修 447 格置信度分布（总 {sum(conf_cnt)}）', fontsize=13)
    ax.grid(True, alpha=0.3, axis='y')
    plt.tight_layout()
    f5 = CHART_DIR / 'refine447_confidence.png'
    plt.savefig(f5, dpi=110); plt.close()
    print(f"图 5: {f5.name}")

    # ── 图表 6: Top 弱数据指标（最难填） ───────────────────────────────
    hard_ind = sorted(ind_stats.items(), key=lambda x: -x[1]['weak'])[:15]
    hard_names = [f"#{iid} {fld_by_id[iid]['字段名'][:18]}" for iid, _ in hard_ind]
    hard_weak = [s['weak'] for _, s in hard_ind]

    fig, ax = plt.subplots(figsize=(12, 7))
    ax.barh(range(len(hard_names)), hard_weak, color='#E53935')
    ax.set_yticks(range(len(hard_names)))
    ax.set_yticklabels(hard_names, fontsize=9)
    ax.invert_yaxis()
    ax.set_xlabel('弱数据国家数（0/无/有限）')
    ax.set_title('Top 15 最难填指标（64 国中弱数据最多者）', fontsize=13)
    ax.grid(True, alpha=0.3, axis='x')
    plt.tight_layout()
    f6 = CHART_DIR / 'hardest_indicators.png'
    plt.savefig(f6, dpi=110); plt.close()
    print(f"图 6: {f6.name}")

    # ── MD 报告 ──────────────────────────────────────────────────────
    lines = []
    lines.append(f"# BRI 64 国中文教育指标体系 · 对比分析报告")
    lines.append("")
    lines.append(f"**生成时间**：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"**数据底本**：13 份 refined447 xlsx → 合并表 `bri_64country_merged_{DATE}.xlsx`")
    lines.append(f"**国家数**：64 | **指标数**：68（#36 空字段已跳过）| **总格数**：4,352")
    lines.append("")
    lines.append("## 一、总体数据质量")
    lines.append("")
    # 全量统计
    grand = {'numeric':0,'qualitative':0,'weak':0,'empty':0}
    for s in country_stats.values():
        for k in grand: grand[k] += s[k]
    gt = sum(grand.values())
    lines.append(f"| 类型 | 数量 | 占比 |")
    lines.append(f"|---|---|---|")
    lines.append(f"| 数字值 | {grand['numeric']} | {grand['numeric']*100/gt:.1f}% |")
    lines.append(f"| 定性值 | {grand['qualitative']} | {grand['qualitative']*100/gt:.1f}% |")
    lines.append(f"| 弱数据（0/无/有限）| {grand['weak']} | {grand['weak']*100/gt:.1f}% |")
    lines.append(f"| 空 | {grand['empty']} | {grand['empty']*100/gt:.1f}% |")
    lines.append(f"| **扎实（数字+定性）** | **{grand['numeric']+grand['qualitative']}** | **{(grand['numeric']+grand['qualitative'])*100/gt:.1f}%** |")
    lines.append("")

    lines.append("## 二、Top 20 国数据质量排名")
    lines.append("")
    lines.append("![Top20](bri_charts/top20_quality_stack.png)")
    lines.append("")
    lines.append(f"| 排名 | 国家 | 数字 | 定性 | 弱 | 空 | 扎实率 |")
    lines.append(f"|---|---|---|---|---|---|---|")
    for i, (c, s) in enumerate(top20, 1):
        lines.append(f"| {i} | {c} | {s['numeric']} | {s['qualitative']} | {s['weak']} | {s['empty']} | {s['solid_pct']:.1f}% |")
    lines.append("")

    lines.append("## 三、64 国数据扎实度完整排名")
    lines.append("")
    lines.append("![All64](bri_charts/all64_solid_rank.png)")
    lines.append("")
    lines.append(f"| 排名 | 国家 | 扎实率 | 数字 | 定性 | 弱 | 空 |")
    lines.append(f"|---|---|---|---|---|---|---|")
    for i, (c, s) in enumerate(all_ranked, 1):
        lines.append(f"| {i} | {c} | {s['solid_pct']:.1f}% | {s['numeric']} | {s['qualitative']} | {s['weak']} | {s['empty']} |")
    lines.append("")

    lines.append("## 四、指标分类维度（A-G 七大类）")
    lines.append("")
    lines.append("![Category](bri_charts/category_coverage.png)")
    lines.append("")
    lines.append(f"| 分类 | 总格 | 数字 | 定性 | 弱 | 空 | 扎实率 |")
    lines.append(f"|---|---|---|---|---|---|---|")
    for c in cats:
        s = cat_stats[c]
        solid_pct = (s['numeric']+s['qualitative'])*100/s['total'] if s['total'] else 0
        lines.append(f"| {c} | {s['total']} | {s['numeric']} | {s['qualitative']} | {s['weak']} | {s['empty']} | {solid_pct:.1f}% |")
    lines.append("")

    lines.append("## 五、优先级维度（P1/P2/P3）")
    lines.append("")
    lines.append("![Priority](bri_charts/priority_coverage.png)")
    lines.append("")
    lines.append(f"| 优先级 | 总格 | 数字 | 定性 | 弱 | 空 | 扎实率 |")
    lines.append(f"|---|---|---|---|---|---|---|")
    for p in prios:
        s = prio_stats[p]
        solid_pct = (s['numeric']+s['qualitative'])*100/s['total'] if s['total'] else 0
        lines.append(f"| {p} | {s['total']} | {s['numeric']} | {s['qualitative']} | {s['weak']} | {s['empty']} | {solid_pct:.1f}% |")
    lines.append("")

    lines.append("## 六、精修 447 轮置信度")
    lines.append("")
    lines.append("![Confidence](bri_charts/refine447_confidence.png)")
    lines.append("")
    lines.append(f"- 总格数：{sum(conf_cnt)}")
    lines.append(f"- 高置信：{ref_conf.get('high',0)}（多源均证确认）")
    lines.append(f"- 中置信：{ref_conf.get('medium',0)}")
    lines.append(f"- 低置信：{ref_conf.get('low',0)}（真无数据，derivation 仍保留）")
    lines.append(f"- 含数字值产出：{ref_numeric}（真精修成功）")
    lines.append("")

    lines.append("## 七、最难填的指标 Top 15")
    lines.append("")
    lines.append("![Hardest](bri_charts/hardest_indicators.png)")
    lines.append("")
    lines.append(f"| # | 指标 | 分类 | 优先级 | 弱数据国 |")
    lines.append(f"|---|---|---|---|---|")
    for iid, s in hard_ind:
        fld = fld_by_id[iid]
        lines.append(f"| {iid} | {fld['字段名']} | {fld.get('分类','')} | {fld.get('优先级','')} | {s['weak']}/64 |")
    lines.append("")

    lines.append("## 八、0 可疑格满分国")
    lines.append("")
    zero_weak = [c for c, s in country_stats.items() if s['weak'] == 0 and s['empty'] == 0]
    lines.append(f"{len(zero_weak)} 国：" + "、".join(zero_weak) if zero_weak else "（无）")
    lines.append("")

    lines.append("## 九、关键发现")
    lines.append("")
    # 跨洲区域对比
    lines.append(f"1. **总扎实率 {(grand['numeric']+grand['qualitative'])*100/gt:.1f}%**，即 4,352 格中 {grand['numeric']+grand['qualitative']} 格（数字+定性）")
    lines.append(f"2. **最扎实国**：{all_ranked[0][0]}（{all_ranked[0][1]['solid_pct']:.1f}%），**最薄弱国**：{all_ranked[-1][0]}（{all_ranked[-1][1]['solid_pct']:.1f}%）")
    lines.append(f"3. **最难填指标**：#{hard_ind[0][0]} {fld_by_id[hard_ind[0][0]]['字段名']}（弱数据覆盖 {hard_ind[0][1]['weak']}/64 国）")
    # 分类分析
    best_cat = max(cat_stats.items(), key=lambda x: (x[1]['numeric']+x[1]['qualitative'])*100/x[1]['total'] if x[1]['total'] else 0)
    worst_cat = min(cat_stats.items(), key=lambda x: (x[1]['numeric']+x[1]['qualitative'])*100/x[1]['total'] if x[1]['total'] else 0)
    lines.append(f"4. **分类扎实率最高**：{best_cat[0]}（{(best_cat[1]['numeric']+best_cat[1]['qualitative'])*100/best_cat[1]['total']:.1f}%），**最弱**：{worst_cat[0]}（{(worst_cat[1]['numeric']+worst_cat[1]['qualitative'])*100/worst_cat[1]['total']:.1f}%）")
    lines.append(f"5. **精修 447 成功率**：{ref_numeric}/{sum(conf_cnt)} 获得数字值（{ref_numeric*100//sum(conf_cnt) if sum(conf_cnt) else 0}%）")
    lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## 相关产出")
    lines.append("")
    lines.append(f"- 合并总表：`topics/bri_64country_merged_{DATE}.xlsx`")
    lines.append(f"- 每国详细指标：`topics/per_country/*_完整指标数据_{DATE}.md`（64 份）")
    lines.append(f"- 精修 447 详单：`topics/bri_421_suspicious_detail_{DATE}.md`")
    lines.append(f"- 精修摘要：`topics/bri_refine447_summary_{DATE}.md`")
    lines.append(f"- 精修推导 JSON：`output/bri_refine447_ckpt.json`")
    lines.append(f"- 13 份批次 refined xlsx：`topics/*_refined447_{DATE}.xlsx`")
    lines.append("")

    OUT_MD.write_text('\n'.join(lines), encoding='utf-8')
    print(f"\n报告：{OUT_MD}")


if __name__ == '__main__':
    main()
