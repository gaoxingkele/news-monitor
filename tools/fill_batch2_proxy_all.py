"""批次 2 统一代理填充：对所有空格 + 语义空 + 可疑 0/无 一次性做扩展代理推导。

流程：
1. 扫 v2，找 gaps（空格、无确切、0/无 非分类）
2. DS 生成扩展代理策略
3. 并发 Perplexity + Tavily(zh/en) + 学术 5 源
4. DS 合成，允许"少量/若干/有限"定性兜底但不允许"无可核验数据"
"""
from __future__ import annotations
import asyncio, io, json, os, sys, logging, time, re
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "tools"))
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError: pass

import httpx, openpyxl
from openpyxl.styles import PatternFill, Alignment
from news_monitor.proxy import build_httpx_client, get_proxies_for_url
from academic_search import academic_search, format_results

TOPICS = ROOT / "topics"
DATE = datetime.now().strftime("%Y%m%d")
SRC = TOPICS / "“一带一路”沿线国家中文教育发展指标体系数据总表_批次2_五国_v2_20260422.xlsx"
V3 = TOPICS / f"“一带一路”沿线国家中文教育发展指标体系数据总表_批次2_五国_v3_{DATE}.xlsx"
LOG_MD = TOPICS / f"bri_batch2_proxy_{DATE}.md"
CKPT = ROOT / "output" / "bri_batch2_proxy_ckpt.json"
FIELD_DICT = TOPICS / "bri_field_dict.json"

PPX = os.environ.get("PERPLEXITY_API_KEY", "")
TAV = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
DS = os.environ.get("DEEPSEEK_API_KEY", "")
PROXY = os.environ.get("LLM_PROXY", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("b2px")

COL = {"马来西亚":(5,6),"泰国":(7,8),"印尼":(9,10),"老挝":(11,12),"缅甸":(13,14)}
EN = {"马来西亚":"Malaysia","泰国":"Thailand","印尼":"Indonesia","老挝":"Laos","缅甸":"Myanmar"}
QUALITATIVE = {2,3,17,51,52,64}

BAD_TERMS = ['无可核验','无确切','无具体','无官方','无数据','无统计','无公开','无法','未发布',
             '未公布','未披露','未见','未找到','未统计','未知','暂无','尚无','不公开','不详',
             '待核','N/A','无法确定','无法计算','无从查证','缺乏','无直接','无明确','无系统']

async def ppx_ask(client, q):
    if not PPX: return "", []
    try:
        r = await client.post("https://api.perplexity.ai/chat/completions",
            json={"model":"sonar","messages":[
                {"role":"system","content":"你是代理研究员。绝不返回 0 或'无'。给定性量级+URL。"},
                {"role":"user","content":q}],
                "temperature":0.1,"return_citations":True},
            headers={"Authorization":f"Bearer {PPX}","Content-Type":"application/json"},
            timeout=90.0)
        r.raise_for_status()
        d = r.json()
        return d["choices"][0]["message"]["content"], (d.get("citations",[]) or [])[:10]
    except Exception as e:
        log.warning("ppx: %s", str(e)[:100]); return "", []

async def tav(client, q, domains=None):
    if not TAV: return ([], "")
    try:
        body = {"api_key":TAV,"query":q,"search_depth":"advanced","max_results":8,
                "include_answer":True,"include_raw_content":False}
        if domains: body["include_domains"] = domains[:15]
        r = await client.post("https://api.tavily.com/search", json=body, timeout=60.0)
        r.raise_for_status()
        d = r.json()
        items = [{"url":x.get("url",""),"title":x.get("title",""),
                 "snippet":(x.get("content") or "")[:400]}
                for x in (d.get("results") or [])[:8]]
        return (items, d.get("answer","") or "")
    except Exception as e:
        log.warning("tav: %s", str(e)[:100]); return ([], "")

async def plan_proxy(client, gap):
    prompt = f"""你是研究方法专家。该指标没有直接数据，请设计扩展口径+代理推导方案。

国家：{gap['country']}
字段：{gap['name']}
取值口径：{gap['type']}
说明：{gap['desc']}
当前：{gap.get('current','(空)')[:100]}

返回 JSON:
{{
  "extended_definition": "扩展口径说明（广义、含多形式、最新年份）",
  "queries": {{
    "perplexity": "给Perplexity的提问",
    "tavily_zh": "中文查询",
    "tavily_en": "英文查询",
    "academic": "学术库查询"
  }},
  "priority_domains": ["官方域名1","域名2"]
}}
"""
    try:
        r = await client.post("https://api.deepseek.com/v1/chat/completions",
            json={"model":"deepseek-chat","messages":[{"role":"user","content":prompt}],
                  "temperature":0.1,"response_format":{"type":"json_object"}},
            headers={"Authorization":f"Bearer {DS}","Content-Type":"application/json"},
            timeout=60.0)
        r.raise_for_status()
        return json.loads(r.json()["choices"][0]["message"]["content"])
    except Exception as e:
        log.warning("plan err: %s", str(e)[:100]); return {}

async def ds_synth(client, gap, hint, evidence):
    prompt = f"""你是数据研究员。为【{gap['country']}】【{gap['name']}】给出最终值。

## 扩展口径
{hint.get('extended_definition','')}

## 取值口径
{gap['type']}

## 证据
{evidence[:12000]}

绝不返回 0/无/合成失败。
- 有数字给数字
- 无数字给区间估值
- 都没有给定性分级（少量/若干/多）+ 依据

返回 JSON:
{{
  "value": "最终答案",
  "proxy_used": "使用的代理",
  "derivation": "推导过程",
  "source_urls": ["URL1","URL2"],
  "confidence": "high/medium/low",
  "is_estimation": true/false,
  "rationale": "30字"
}}
"""
    try:
        r = await client.post("https://api.deepseek.com/v1/chat/completions",
            json={"model":"deepseek-chat","messages":[{"role":"user","content":prompt}],
                  "temperature":0.1,"response_format":{"type":"json_object"}},
            headers={"Authorization":f"Bearer {DS}","Content-Type":"application/json"},
            timeout=90.0)
        r.raise_for_status()
        return json.loads(r.json()["choices"][0]["message"]["content"])
    except Exception as e:
        log.warning("synth err %s#%d: %s", gap['country'], gap['id'], str(e)[:100])
        return {"value":"合成失败","confidence":"low"}

async def solve(sem, gap, clients, ckpt):
    key = f"{gap['country']}::{gap['id']}"
    if key in ckpt: return ckpt[key]
    async with sem:
        t0 = time.time()
        ppx_c, tav_c, ds_c = clients

        plan = await plan_proxy(ds_c, gap)
        if not plan.get('queries'):
            ckpt[key] = {"value":"无可用策略","confidence":"low"}
            _save(ckpt); return ckpt[key]

        q = plan['queries']; doms = plan.get('priority_domains', [])
        results = await asyncio.gather(
            ppx_ask(ppx_c, q.get('perplexity','')),
            tav(tav_c, q.get('tavily_zh',''), doms),
            tav(tav_c, q.get('tavily_en',''), None),
            academic_search(q.get('academic',''), max_per_source=3,
                            proxy_url=PROXY, include_scholar=True),
            return_exceptions=True,
        )
        ppx_res = results[0] if isinstance(results[0],tuple) else ("", [])
        tav_zh = results[1] if isinstance(results[1],tuple) and len(results[1])==2 else ([],"")
        tav_en = results[2] if isinstance(results[2],tuple) and len(results[2])==2 else ([],"")
        acad = results[3] if isinstance(results[3],dict) else {}

        evid_parts = [f"## Perplexity\n{ppx_res[0]}\nCitations:\n" + "\n".join(ppx_res[1][:8])]
        evid_parts.append(f"## Tavily-zh answer: {tav_zh[1]}")
        for x in tav_zh[0][:5]: evid_parts.append(f"  - {x['title']} | {x['url']}\n    {x['snippet'][:200]}")
        evid_parts.append(f"## Tavily-en answer: {tav_en[1]}")
        for x in tav_en[0][:5]: evid_parts.append(f"  - {x['title']} | {x['url']}\n    {x['snippet'][:200]}")
        evid_parts.append(f"## Academic\n{format_results(acad)[:3500]}")
        evidence = "\n\n".join(evid_parts)

        result = await ds_synth(ds_c, gap, plan, evidence)
        result['plan'] = plan
        ckpt[key] = result
        _save(ckpt)
        log.info("%s#%d %.1fs value=%s conf=%s",
                 gap['country'], gap['id'], time.time()-t0,
                 str(result.get('value',''))[:40], result.get('confidence','?'))
        return result

def _save(ckpt):
    with open(CKPT, "w", encoding="utf-8") as f:
        json.dump(ckpt, f, ensure_ascii=False, indent=1)

def identify_gaps():
    with open(FIELD_DICT, encoding='utf-8') as f:
        raw = json.load(f)
    fd = {int(x['指标ID']):x for x in raw if x.get('指标ID')}
    wb = openpyxl.load_workbook(SRC, data_only=True); ws = wb.active
    gaps = []
    for cname, (vcol, _) in COL.items():
        for r in range(2, ws.max_row+1):
            v_raw = ws.cell(r, vcol).value
            iid = ws.cell(r, 1).value
            if iid is None or iid == 36: continue
            fld = fd.get(int(iid));
            if not fld: continue
            v = str(v_raw or '').strip()
            # 空
            is_empty = not v
            # 语义空
            is_soft = any(b in v for b in BAD_TERMS)
            # 0/无（非分类）
            clean = re.sub(r'^(【估算】|【代理推导】)\s*', '', v)
            first = re.split(r'[（\(]', clean)[0].strip()
            is_zero = bool(re.match(r'^0(\.\d+)?(%|人|所|项|个|次|篇)?$', first)) or first in ('0','无','没有')
            if iid in QUALITATIVE and is_zero and not is_empty and not is_soft:
                continue  # 分类类真 0/无，跳过
            if is_empty or is_soft or is_zero:
                gaps.append({
                    'country': cname, 'id': int(iid), 'row': r, 'vcol': vcol,
                    'name': fld.get('字段名',''), 'type': fld.get('字段类型',''),
                    'desc': (fld.get('字段说明','') or '')[:200],
                    'category': fld.get('分类',''),
                    'priority': fld.get('优先级','P?'),
                    'current': v[:100] if v else '(空)',
                })
    return gaps

async def run():
    gaps = identify_gaps()
    log.info("gaps: %d", len(gaps))
    by_p = {}
    for g in gaps: by_p[g['priority']] = by_p.get(g['priority'], 0) + 1
    log.info("by P: %s", by_p)

    ckpt = {}
    if CKPT.exists():
        try: ckpt = json.loads(CKPT.read_text(encoding='utf-8'))
        except Exception: pass
    log.info("ckpt: %d", len(ckpt))

    ppx_p = get_proxies_for_url("https://api.perplexity.ai", PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", PROXY)
    async with (build_httpx_client(ppx_p or "", 90) as ppx_c,
                build_httpx_client(tav_p or "", 60) as tav_c,
                build_httpx_client("", 90) as ds_c):
        sem = asyncio.Semaphore(4)
        tasks = [solve(sem, g, (ppx_c, tav_c, ds_c), ckpt) for g in gaps]
        await asyncio.gather(*tasks)

    # Write v3
    wb = openpyxl.load_workbook(SRC); ws = wb.active
    proxy_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    log_md = [f"# 批次 2 代理推导日志 (v3)", f"时间：{datetime.now():%Y-%m-%d %H:%M}",
              f"对 {len(gaps)} 空/可疑格做扩展代理推导。\n"]

    updated = 0
    for g in gaps:
        r = ckpt.get(f"{g['country']}::{g['id']}", {})
        val = str(r.get('value','')).strip()
        if not val or val in ('合成失败','无可用策略'): continue
        # 仍为 0/无 则不覆盖
        if re.match(r'^(0|无|没有|0%)$', val): continue
        updated += 1
        prefix = "【代理推导】" if r.get('is_estimation') else "【代理】"
        proxy = r.get('proxy_used',''); deriv = r.get('derivation','')
        txt = f"{prefix}{val}"
        if proxy: txt += f"（代理：{proxy}）"
        if deriv: txt += f"\n推导：{deriv}"
        ws.cell(g['row'], g['vcol'], txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(g['row'], g['vcol']).fill = proxy_fill
        urls = r.get('source_urls', [])
        if urls: ws.cell(g['row'], g['vcol']+1, "; ".join(urls[:3]))

        log_md.append(f"\n## {g['country']}#{g['id']} {g['name']} ({g['priority']})")
        log_md.append(f"- 旧值：{g['current'][:60]}")
        log_md.append(f"- 新值：{val}")
        log_md.append(f"- 代理：{proxy}")
        log_md.append(f"- 推导：{deriv}")
        log_md.append(f"- 置信度：{r.get('confidence','?')}")
        log_md.append(f"- 理由：{r.get('rationale','')}")
        for u in (urls or [])[:3]: log_md.append(f"- URL: {u}")

    wb.save(V3)
    LOG_MD.write_text("\n".join(log_md), encoding="utf-8")
    log.info("saved %s", V3.name)
    log.info("saved %s", LOG_MD.name)
    log.info("DONE: %d/%d 成功更新", updated, len(gaps))


if __name__ == "__main__":
    asyncio.run(run())
