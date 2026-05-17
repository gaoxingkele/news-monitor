"""扫描 13 份 v_final xlsx，生成 421 可疑格详细 MD 清单。

可疑判定（与 run_batch_pipeline.py refine 阶段一致）：
  - empty: 空
  - zero: 0 / 无 / 没有（分类字段 2/3/17/51/52/64 除外）
  - bad: 含 BAD_TERMS（无可核验/未发布/N/A 等）
  - soft: 以 少量/若干/有限 开头

输出：topics/bri_421_suspicious_detail_{DATE}.md
"""
from __future__ import annotations
import io, json, os, re, sys
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

ROOT = Path(r"D:\aicoding\news-monitor")
TOPICS = ROOT / "topics"
FIELD_DICT = TOPICS / "bri_field_dict.json"
DATE = datetime.now().strftime("%Y%m%d")
OUT_MD = TOPICS / f"bri_421_suspicious_detail_{DATE}.md"
OUT_JSON = ROOT / "output" / f"bri_421_suspicious_targets_{DATE}.json"

import openpyxl

QUALITATIVE_IDS = {2, 3, 17, 51, 52, 64}
BAD_TERMS = ['无可核验','无确切','无具体','无官方','无数据','无统计','无公开','无法','未发布',
             '未公布','未披露','未见','未找到','未统计','未知','暂无','尚无','不公开','不详',
             '待核','N/A','无法确定','无法计算','无从查证','缺乏','无直接','无明确','无系统']
SOFT_TERMS = ['少量', '若干', '有限']

# 批次 → (文件名, [国家顺序])
BATCHES = [
    (1, "“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v11_20260422.xlsx",
        ["蒙古", "越南", "柬埔寨", "菲律宾", "新加坡"]),
    (2, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次2_五国_v4_20260422.xlsx",
        ["马来西亚", "泰国", "印尼", "老挝", "缅甸"]),
    (3, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次3_南亚I_五国_v_final_20260423.xlsx",
        ["文莱", "东帝汶", "不丹", "孟加拉国", "尼泊尔"]),
    (4, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次4_南亚II_五国_v_final_20260423.xlsx",
        ["斯里兰卡", "印度", "巴基斯坦", "阿富汗", "马尔代夫"]),
    (5, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次5_中亚_五国_v_final_20260423.xlsx",
        ["土库曼斯坦", "乌兹别克斯坦", "哈萨克斯坦", "塔吉克斯坦", "吉尔吉斯斯坦"]),
    (6, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次6_西亚海湾_五国_v_final_20260423.xlsx",
        ["阿曼", "伊朗", "阿联酋", "卡塔尔", "巴林"]),
    (7, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次7_西亚阿拉伯_五国_v_final_20260423.xlsx",
        ["沙特阿拉伯", "科威特", "也门", "土耳其", "伊拉克"]),
    (8, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次8_地中海东岸_五国_v_final_20260423.xlsx",
        ["叙利亚", "约旦", "黎巴嫩", "以色列", "格鲁吉亚"]),
    (9, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次9_高加索北非_五国_v_final_20260423.xlsx",
        ["亚美尼亚", "阿塞拜疆", "埃及", "爱沙尼亚", "拉脱维亚"]),
    (10, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次10_欧洲中部I_五国_v_final_20260423.xlsx",
        ["立陶宛", "德国", "波兰", "捷克", "斯洛文尼亚"]),
    (11, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次11_巴尔干I_五国_v_final_20260423.xlsx",
        ["克罗地亚", "塞尔维亚", "波黑", "黑山", "阿尔巴尼亚"]),
    (12, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次12_欧洲中部II_五国_v_final_20260423.xlsx",
        ["匈牙利", "斯洛伐克", "北马其顿", "罗马尼亚", "保加利亚"]),
    (13, "“一带一路”沿线国家中文教育发展指标体系数据总表_批次13_东欧_五国_v_final_20260423.xlsx",
        ["摩尔多瓦", "乌克兰", "白俄罗斯", "俄罗斯"]),
]


def classify(v_raw, iid: int) -> str:
    """与 audit JSON 对齐的分类器：

    - empty: 空
    - zero: 实际值为 0（分类字段除外）
    - wu: 实际值为 无/没有
    - soft: 实际值以 少量/若干/有限 开头
    - bad: 实际值以 BAD_TERMS（无可核验/未发布 等）开头
    - ok: 其他

    "实际值" = 剥掉 【代理推导】/【估算】 前缀 + 取首个圆括号 / 行前部分
    """
    v = str(v_raw or '').strip()
    if not v:
        return 'empty'
    # 剥所有前置的【…】标签
    clean = re.sub(r'^(【[^】]+】\s*)+', '', v).strip()
    if not clean:
        return 'bad'
    # 再检查剥下来的标签中是否含 BAD_TERM（比如 【未公开】）
    prefix_tags = re.findall(r'【([^】]+)】', v[:len(v) - len(clean)])
    for tag in prefix_tags:
        if any(b in tag for b in BAD_TERMS):
            return 'bad'
    # 取首段（圆括号 / 换行之前）
    first_seg = re.split(r'[（\(\n]', clean, maxsplit=1)[0].strip()
    # 再按逗号/分号截取主值
    first = re.split(r'[，,；;：:]', first_seg, maxsplit=1)[0].strip()

    # zero
    if re.match(r'^0(\.\d+)?(%|人|所|项|个|次|篇|家|门|部|册|台|万|亿|元|美元)?$', first) or first in ('0', '0.0'):
        if iid in QUALITATIVE_IDS:
            return 'ok'
        return 'zero'

    # wu: 只在分类字段以外，且不是长推导内容
    if first in ('无', '没有', '无数据', '无明确数据', '无相关数据'):
        if iid in QUALITATIVE_IDS:
            return 'ok'
        # 若整段文本很长（>80 字），通常是带推导的合法答案，不算可疑
        if len(clean) > 80:
            return 'ok'
        return 'wu'

    # soft
    for s in SOFT_TERMS:
        if first.startswith(s):
            return 'soft'

    # bad — 实际值以 BAD_TERMS 开头
    for b in BAD_TERMS:
        if first.startswith(b) or first == b:
            return 'bad'

    return 'ok'


def main():
    with open(FIELD_DICT, encoding='utf-8') as f:
        fields = {int(x['指标ID']): x for x in json.load(f) if x.get('指标ID')}

    targets = []
    batch_stats = {}
    for batch_id, fname, countries in BATCHES:
        fp = TOPICS / fname
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        cnt = 0
        for i, cname in enumerate(countries):
            vcol = 5 + 2 * i
            for r in range(2, ws.max_row + 1):
                iid_raw = ws.cell(r, 1).value
                if iid_raw is None:
                    continue
                try:
                    iid = int(iid_raw)
                except (TypeError, ValueError):
                    continue
                if iid == 36:
                    continue
                fld = fields.get(iid)
                if not fld:
                    continue
                v_raw = ws.cell(r, vcol).value
                kind = classify(v_raw, iid)
                if kind == 'ok':
                    continue
                targets.append({
                    'batch': batch_id,
                    'file': fname,
                    'country': cname,
                    'vcol': vcol,
                    'row': r,
                    'id': iid,
                    'name': fld.get('字段名', ''),
                    'type': fld.get('字段类型', ''),
                    'category': fld.get('分类', ''),
                    'priority': fld.get('优先级', '?'),
                    'desc': (fld.get('字段说明', '') or '')[:300],
                    'current': str(v_raw) if v_raw is not None else '',
                    'kind': kind,
                })
                cnt += 1
        batch_stats[batch_id] = cnt
        print(f"批次 {batch_id} {fname[:30]}... {cnt} 格可疑")

    # 写 JSON
    OUT_JSON.parent.mkdir(exist_ok=True)
    OUT_JSON.write_text(json.dumps(targets, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"\n写入 {OUT_JSON} — 共 {len(targets)} 格")

    # 统计
    by_country = {}
    by_kind = {'empty': 0, 'zero': 0, 'wu': 0, 'bad': 0, 'soft': 0}
    by_priority = {}
    for t in targets:
        by_country.setdefault(t['country'], {'total': 0, 'empty': 0, 'zero': 0, 'wu': 0, 'bad': 0, 'soft': 0})
        by_country[t['country']]['total'] += 1
        by_country[t['country']][t['kind']] += 1
        by_kind[t['kind']] += 1
        by_priority[t['priority']] = by_priority.get(t['priority'], 0) + 1

    # 写 MD
    lines = []
    lines.append(f"# BRI 64 国 · 可疑格详细清单")
    lines.append("")
    lines.append(f"**生成时间**：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"**数据来源**：13 份 v_final xlsx")
    lines.append(f"**总可疑格**：{len(targets)}")
    lines.append("")
    lines.append(f"## 按类型分布")
    lines.append(f"| 类型 | 数量 | 说明 |")
    lines.append(f"|---|---|---|")
    lines.append(f"| empty | {by_kind['empty']} | 单元格为空 |")
    lines.append(f"| zero | {by_kind['zero']} | 值为 0（非分类字段）|")
    lines.append(f"| wu | {by_kind['wu']} | 值为 无/没有 |")
    lines.append(f"| bad | {by_kind['bad']} | 含『无可核验/未发布/N/A』等失败标记 |")
    lines.append(f"| soft | {by_kind['soft']} | 以『少量/若干/有限』开头的软定性 |")
    lines.append("")
    lines.append(f"## 按批次分布")
    lines.append(f"| 批次 | 可疑数 |")
    lines.append(f"|---|---|")
    for bid in sorted(batch_stats):
        lines.append(f"| {bid} | {batch_stats[bid]} |")
    lines.append("")
    lines.append(f"## 按优先级分布")
    lines.append(f"| 优先级 | 可疑数 |")
    lines.append(f"|---|---|")
    for p in sorted(by_priority):
        lines.append(f"| {p} | {by_priority[p]} |")
    lines.append("")
    lines.append(f"## 按国家分布")
    lines.append(f"| 国家 | 总数 | empty | zero | wu | bad | soft |")
    lines.append(f"|---|---|---|---|---|---|---|")
    for c in sorted(by_country, key=lambda k: -by_country[k]['total']):
        s = by_country[c]
        lines.append(f"| {c} | {s['total']} | {s['empty']} | {s['zero']} | {s['wu']} | {s['bad']} | {s['soft']} |")
    lines.append("")
    lines.append(f"---")
    lines.append("")
    lines.append(f"## 详细清单（按批次→国家→指标ID 排序）")
    lines.append("")
    cur_batch = None
    cur_country = None
    targets_sorted = sorted(targets, key=lambda t: (t['batch'], t['country'], t['id']))
    for t in targets_sorted:
        if t['batch'] != cur_batch:
            cur_batch = t['batch']
            cur_country = None
            lines.append(f"### 批次 {cur_batch}")
            lines.append("")
        if t['country'] != cur_country:
            cur_country = t['country']
            lines.append(f"#### {cur_country}（批次 {cur_batch}）")
            lines.append("")
            lines.append("| ID | 字段名 | 分类 | 优先级 | 类型 | 当前值 | 问题 |")
            lines.append("|---|---|---|---|---|---|---|")
        cur_val = t['current'].replace('|', '\\|').replace('\n', ' ')
        if len(cur_val) > 60:
            cur_val = cur_val[:60] + '…'
        name = t['name'].replace('|', '\\|')
        typ = t['type'].replace('|', '\\|')
        lines.append(f"| {t['id']} | {name} | {t['category'].replace('|','\\|')} | {t['priority']} | {typ} | {cur_val} | {t['kind']} |")

    OUT_MD.write_text('\n'.join(lines), encoding='utf-8')
    print(f"写入 {OUT_MD}")


if __name__ == '__main__':
    main()
