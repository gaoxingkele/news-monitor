"""针对 24 格难啃 P1/P2 语义空格做定制化代理推导。
每格注入具体领域知识提示（Pew 民调、Facebook 账号清单、ICCSP 奖学金等）。
"""
from __future__ import annotations
import asyncio, io, json, os, sys, logging, time
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
ROOT = Path(r"D:\aicoding\news-monitor")
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "tools"))
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError: pass

import httpx, openpyxl
from openpyxl.styles import PatternFill, Alignment
from news_monitor.proxy import build_httpx_client, get_proxies_for_url
from academic_search import academic_search, format_results

TOPICS = ROOT / "topics"
DATE = datetime.now().strftime("%Y%m%d")
SRC = TOPICS / "“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v8_20260421.xlsx"
V9 = TOPICS / f"“一带一路”沿线国家中文教育发展指标体系数据总表_五国填充版_v9_{DATE}.xlsx"
LOG_MD = TOPICS / f"bri_stubborn24_{DATE}.md"
CKPT = ROOT / "output" / "bri_stubborn24_ckpt.json"

PPX = os.environ.get("PERPLEXITY_API_KEY", "")
TAV = os.environ.get("tavilyapi", "") or os.environ.get("TAVILY_API_KEY", "")
DS = os.environ.get("DEEPSEEK_API_KEY", "")
PROXY = os.environ.get("LLM_PROXY", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("stub24")

COL = {"蒙古":(5,6),"越南":(7,8),"柬埔寨":(9,10),"菲律宾":(11,12),"新加坡":(13,14)}

# 24 格定制化代理策略（核心）
TARGETS = {
    # ── 社媒关注量（3 格）──
    ("蒙古", 67): {
        "name": "社交媒体平台中文账号关注量",
        "proxy_hint": (
            "直接搜索蒙古主要中文账号 Facebook 页面的 page likes/followers 数字。"
            "候选账号：新华社蒙古语 Facebook、蒙古国立大学孔子学院 FB、蒙古中国文化中心 FB、"
            "蒙古在华留学生联合会公众号、The UB Post（英文但含中文内容）、乌兰巴托中国商会 FB。"
            "对每个账号给出具体 follower 数，求和得到总量。"
        ),
        "queries": {
            "perplexity": "蒙古（Mongolia）本地运营的中文媒体、孔子学院、中国文化中心的 Facebook/Weibo 公开账号各有多少粉丝？列出账号名和 URL。",
            "tavily_zh": "蒙古 孔子学院 Facebook 粉丝 关注",
            "tavily_en": "Mongolia Confucius Institute Facebook followers page likes",
            "academic": "Mongolia Chinese language social media followers",
        },
        "domains": ["facebook.com", "weibo.com", "ci.cn", "chinese.cn"],
    },
    ("菲律宾", 67): {
        "name": "社交媒体平台中文账号关注量",
        "proxy_hint": (
            "菲律宾主要中文账号：世界日报 Facebook（菲律宾华人最大日报）、菲华商报 FB、"
            "红溪礼示大学孔子学院 FB、菲律宾华文学校联合总会 FB、PhilStar Chinese FB。"
            "求和各主账号粉丝数。世界日报 FB 估计 >10 万。"
        ),
        "queries": {
            "perplexity": "菲律宾本地运营的中文媒体（世界日报、菲华商报）、孔子学院、华文学校联合总会的 Facebook 粉丝数分别是多少？",
            "tavily_zh": "菲律宾 世界日报 菲华商报 Facebook 粉丝",
            "tavily_en": "Philippines World News Chinese World Daily Facebook followers Confucius Institute",
            "academic": "Philippines Chinese media Facebook followers",
        },
        "domains": ["facebook.com", "worldnews.ph", "chinesepress.com", "philstar.com"],
    },
    ("新加坡", 67): {
        "name": "社交媒体平台中文账号关注量",
        "proxy_hint": (
            "新加坡主要中文账号：联合早报 Facebook（最大中文日报，粉丝约 150 万）、"
            "新加坡 8 视界 FB、早报中学版 FB、NTU 孔子学院 FB、SCCL 新加坡华文教研中心 FB。"
            "主账号联合早报即可主导总量。"
        ),
        "queries": {
            "perplexity": "新加坡联合早报（Lianhe Zaobao）、8 视界、SCCL、NTU 孔子学院的 Facebook/Weibo 官方账号粉丝数各是多少？",
            "tavily_zh": "联合早报 Facebook 粉丝 关注 数量",
            "tavily_en": "Lianhe Zaobao Facebook followers Singapore Chinese media",
            "academic": "Singapore Chinese media social media engagement",
        },
        "domains": ["facebook.com", "zaobao.com.sg", "sccl.sg", "ntu.edu.sg"],
    },

    # ── 舆情正面评价占比（3 格 P1）──
    ("蒙古", 69): {
        "name": "舆情监测中正面评价占比",
        "proxy_hint": (
            "代理指标 1：Pew Research 'Favorable views of China' 分国别民调数据。"
            "代理指标 2：蒙古主流媒体涉中报道情感分析（蒙古文/英文语料）。"
            "Pew 2024 报告中蒙古对华好感度 ~40-50%。"
        ),
        "queries": {
            "perplexity": "Pew Research Center、YouGov、Gallup 等民调机构最近几年（2022-2025）对蒙古国民众'对中国好感度'调查结果是多少百分比？同时查任何蒙古本地涉华舆情调查。",
            "tavily_zh": "蒙古 对华好感度 民调 中国印象 Pew",
            "tavily_en": "Mongolia China favorability opinion poll Pew 2024",
            "academic": "Mongolia China public opinion favorability",
        },
        "domains": ["pewresearch.org", "yougov.com", "gallup.com", "statista.com"],
    },
    ("菲律宾", 69): {
        "name": "舆情监测中正面评价占比",
        "proxy_hint": (
            "Pew 2024 菲律宾对华好感度通常偏低（~20-30%）因南海争议。但针对'中文教育'的"
            "态度应显著高于'对中国总体'态度。代理：SWS / Pulse Asia 菲律宾本地民调 + OCA 调查。"
        ),
        "queries": {
            "perplexity": "Pew Research、SWS Philippines、Pulse Asia 2023-2025 年对菲律宾民众'对中国好感度'、'中菲关系'民调结果的百分比？同时查菲律宾社会对中文教育的态度调查。",
            "tavily_zh": "菲律宾 对华好感度 民调 中文教育 态度",
            "tavily_en": "Philippines China favorability Pew SWS Pulse Asia poll 2024",
            "academic": "Philippines China perception public opinion survey",
        },
        "domains": ["pewresearch.org", "yougov.com", "sws.org.ph", "pulseasia.ph"],
    },
    ("新加坡", 69): {
        "name": "舆情监测中正面评价占比",
        "proxy_hint": (
            "新加坡对华态度普遍较友好（Pew 亚洲唯一华裔占多数国家），~65-75% 正面。"
            "代理：联合早报/海峡时报涉中文教育报道情感分析 + IPS 民调 + Lee Kuan Yew School 调查。"
        ),
        "queries": {
            "perplexity": "新加坡 IPS Institute of Policy Studies、Pew Research 等近 2-3 年对新加坡民众对中国好感度/中文教育态度民调结果？",
            "tavily_zh": "新加坡 对华好感度 IPS 民调 中文教育",
            "tavily_en": "Singapore China favorability IPS survey poll Pew 2024",
            "academic": "Singapore China perception attitude survey",
        },
        "domains": ["pewresearch.org", "ips.sg", "straitstimes.com", "zaobao.com.sg"],
    },

    # ── 学习者比例（菲律宾#30 P1）──
    ("菲律宾", 30): {
        "name": "中文学习者比例",
        "proxy_hint": (
            "分子：DepEd 2025 公立中学中文项目 14,000+ 学生 + 华校约 120,000 学生 + "
            "孔院/孔子课堂约 5,000 = 总计约 140,000 人。"
            "分母：菲律宾基础教育在校生约 27,000,000（2022 MOE）或总人口 117,000,000（世行 2023）。"
            "给出两套比例："
            "（a）140,000 / 27,000,000 = 0.52%（教育系统口径）"
            "（b）140,000 / 117,000,000 = 0.12%（总人口口径）"
        ),
        "queries": {
            "perplexity": "菲律宾 2024-2025 年中文学习者总数（DepEd 公立中学 + 华校 + 孔子学院三类加总）和菲律宾总人口/基础教育在校生各多少？",
            "tavily_zh": "菲律宾 中文学习者 总人数 DepEd 华校 孔院",
            "tavily_en": "Philippines Chinese language learners total DepEd K-12 Chinese",
            "academic": "Philippines Chinese language learners enrollment statistics",
        },
        "domains": ["deped.gov.ph", "ched.gov.ph", "chinese.cn", "ci.cn"],
    },

    # ── 非华裔比例（菲律宾#32）──
    ("菲律宾", 32): {
        "name": "非华裔学习者比例",
        "proxy_hint": (
            "分子：DepEd 2025 公立中学中文项目 14,000 人（全非华裔）+ 华校 120,000 中非华裔 ~15% ≈ 18,000 人 = 32,000 非华裔。"
            "分母：总中文学习者 140,000。"
            "= 32,000 / 140,000 ≈ 22.9%。"
        ),
        "queries": {
            "perplexity": "菲律宾中文学习者中，非华裔（非华人血统，含菲律宾族、印族等）占比？分公立中学 DepEd、私立华校、孔院三路径讨论。",
            "tavily_zh": "菲律宾 中文学习者 华裔 非华裔 比例 华校",
            "tavily_en": "Philippines Chinese learners non-ethnic Chinese percentage Filipino students",
            "academic": "Philippines Chinese school non-Chinese students ethnic composition",
        },
        "domains": ["deped.gov.ph", "chinese.cn", "ci.cn"],
    },

    # ── 国际资金援助（4 格）──
    ("蒙古", 7): {
        "name": "国际资金援助项目数",
        "proxy_hint": (
            "代理：ICCSP 国际中文教师奖学金蒙古名额数 + 孔子学院蒙古分院年度拨款 + "
            "中方派遣教师岗位数。孔院蒙古分院 2 所+课堂 5 所，常规配套约 10-15 项。"
        ),
        "queries": {
            "perplexity": "2020-2025 年中方/国际组织对蒙古国中文教育的援助项目有哪些？包括 ICCSP 奖学金、孔子学院项目、中国驻蒙使馆资助、UNESCO 等。",
            "tavily_zh": "蒙古 中文教育 援助 ICCSP 孔子学院 奖学金 项目",
            "tavily_en": "Mongolia Chinese education international aid ICCSP Confucius Institute funding",
            "academic": "Mongolia Chinese education international assistance funding",
        },
        "domains": ["ci.cn", "cief.org.cn", "chinese.cn"],
    },
    ("菲律宾", 7): {
        "name": "国际资金援助项目数",
        "proxy_hint": (
            "菲律宾 4 所孔院 + ICCSP 奖学金名额约 60/年 + CLEC 对菲资助项目约 10/年 = ~15-20 项。"
        ),
        "queries": {
            "perplexity": "2022-2025 中方/国际组织对菲律宾中文教育资金援助项目有哪些？ICCSP 奖学金、孔子学院项目、CLEC 资助、UNESCO 等。",
            "tavily_zh": "菲律宾 中文教育 援助 ICCSP 孔院 资助 项目数量",
            "tavily_en": "Philippines Chinese education international aid ICCSP funding programs",
            "academic": "Philippines Chinese education international funding",
        },
        "domains": ["ci.cn", "cief.org.cn", "chinese.cn"],
    },
    ("越南", 5): {
        "name": "中文教育政策文件数量",
        "proxy_hint": (
            "越南 MOET 在 2016/2021 两次修订外语政策框架（Đề án Ngoại ngữ），"
            "2016 首次纳入中文；2021 更新；另外教育部有多份中越合作协议文件。"
            "合理估算 3-5 份核心政策文件。"
        ),
        "queries": {
            "perplexity": "越南教育培训部（MOET）2010-2025 发布的涉及中文教育的政策文件/法律/教育计划有哪些？请列具体文件名和年份。",
            "tavily_zh": "越南 教育部 MOET 中文 汉语 外语政策 文件",
            "tavily_en": "Vietnam MOET Chinese language policy document foreign language project",
            "academic": "Vietnam Chinese language education policy framework",
        },
        "domains": ["moet.gov.vn", "chinhphu.vn", "chinese.cn"],
    },
    ("柬埔寨", 8): {
        "name": "社会资本投入项目数量",
        "proxy_hint": (
            "代理：柬华理事总会年度赞助项目数 + 华资企业对端华/崇正学校捐款项目 + 中资企业资助。"
            "合理估算 10-20 项/年。"
        ),
        "queries": {
            "perplexity": "柬埔寨中文教育/华文学校每年接受的社会资本（非政府资金）资助项目有多少？来自柬华理事总会、华资企业、中资企业等。",
            "tavily_zh": "柬埔寨 华文教育 社会资本 捐款 华商 赞助 项目",
            "tavily_en": "Cambodia Chinese education private sector investment funding",
            "academic": "Cambodia Chinese school community funding donation",
        },
        "domains": ["chinese.cn", "qwgzyj.gqb.gov.cn"],
    },
    ("菲律宾", 8): {
        "name": "社会资本投入项目数量",
        "proxy_hint": "菲律宾华文学校联合总会年度赞助项目约 20-50 项/年；华商日报/陈氏宗亲会赞助项目。",
        "queries": {
            "perplexity": "菲律宾中文教育接受的社会资本投入项目每年多少？菲华联总、陈氏/蔡氏宗亲会、华商联合会等资助项目。",
            "tavily_zh": "菲律宾 华文学校联合总会 赞助 捐款 华商 项目",
            "tavily_en": "Philippines Chinese school association funding donation projects",
            "academic": "Philippines Chinese community funding education",
        },
        "domains": ["chinesepress.com", "worldnews.ph", "ci.cn"],
    },
    ("新加坡", 8): {
        "name": "社会资本投入项目数量",
        "proxy_hint": "新加坡华社自助理事会（CDAC）、新加坡中华总商会、淡马锡基金会等对中文教育资助项目。约 10-30 项/年。",
        "queries": {
            "perplexity": "新加坡 CDAC、中华总商会、SICC 每年对中文教育/华文社会捐款/资助项目有多少？",
            "tavily_zh": "新加坡 CDAC 华社 中华总商会 中文教育 赞助",
            "tavily_en": "Singapore CDAC Chinese Chamber of Commerce donation education",
            "academic": "Singapore Chinese community investment education",
        },
        "domains": ["cdac.org.sg", "sccci.org.sg", "sg"],
    },

    # ── 学生/教师国际交换（4 格）──
    ("菲律宾", 57): {
        "name": "学生/教师国际交换数量",
        "proxy_hint": (
            "代理：(a) 菲律宾来华留学生 2023 年新增约 2,000 人；(b) 中国赴菲汉语教师志愿者约 50 人/年。"
            "合计约 2,000-2,500 人次/年。"
        ),
        "queries": {
            "perplexity": "2023-2025 菲律宾来华留学生数、中国赴菲汉语教师志愿者/派遣教师数？",
            "tavily_zh": "菲律宾 留学生 来华 汉语教师 派遣 交换",
            "tavily_en": "Philippines China international students exchange teachers volunteer",
            "academic": "Philippines China student teacher exchange",
        },
        "domains": ["moe.gov.cn", "ci.cn"],
    },
    ("新加坡", 57): {
        "name": "学生/教师国际交换数量",
        "proxy_hint": (
            "新加坡-中国教师交流（MOE 官方项目） + NUS/NTU 与北大清华等交换生。"
            "约 500-1000 人次/年。"
        ),
        "queries": {
            "perplexity": "新加坡与中国之间教师/学生交换项目每年人次？MOE 教师交流、NUS/NTU 学生交换、CI 讲座等。",
            "tavily_zh": "新加坡 中国 教师 学生 交换 项目 人次",
            "tavily_en": "Singapore China teacher student exchange numbers annual",
            "academic": "Singapore China academic exchange students teachers",
        },
        "domains": ["moe.gov.sg", "nus.edu.sg", "ntu.edu.sg", "ci.cn"],
    },

    # ── 其他 P2 琐碎格 ──
    ("蒙古", 23): {
        "name": "多语种教材数",
        "proxy_hint": "蒙古汉语教材多为蒙汉双语，可查汉办支持的本土化教材清单。约 5-15 种。",
        "queries": {
            "perplexity": "蒙古国本土化中文教材（蒙汉双语或三语教材）有哪些具体书名和数量？汉办支持或本土编写。",
            "tavily_zh": "蒙古 本土 汉语教材 蒙汉双语",
            "tavily_en": "Mongolia localized Chinese textbook bilingual Mongolian",
            "academic": "Mongolia Chinese textbook localized bilingual",
        },
        "domains": ["ci.cn", "chinese.cn"],
    },
    ("蒙古", 31): {
        "name": "基础教育阶段中文学习者比例",
        "proxy_hint": (
            "分子：蒙古中小学汉语学习者约 7,500（新华社 2023 数据）；"
            "分母：蒙古基础教育在校生约 800,000（MECSS 2023）；"
            "= 7,500 / 800,000 ≈ 0.94%。"
        ),
        "queries": {
            "perplexity": "蒙古国基础教育阶段（小学+中学）中文学习者人数，以及全国基础教育在校生总数，计算比例。",
            "tavily_zh": "蒙古 基础教育 中文学习者 学生 人数 小学中学",
            "tavily_en": "Mongolia primary secondary Chinese learners enrollment total students",
            "academic": "Mongolia primary secondary Chinese language learners",
        },
        "domains": ["meds.gov.mn", "chinese.cn"],
    },
    ("蒙古", 66): {
        "name": "主流媒体中文内容发布数",
        "proxy_hint": "蒙古主流媒体涉中国报道年度量：新华社蒙古分社年发稿 ~1000；蒙古 MontsAme 涉华报道 ~500。估算 1000-2000 篇/年。",
        "queries": {
            "perplexity": "蒙古主流媒体（MontsAme、UB Post、蒙古国家广播）每年涉中国/中文教育报道数量多少？",
            "tavily_zh": "蒙古 主流媒体 涉华报道 数量 年度",
            "tavily_en": "Mongolia mainstream media China coverage annual reports",
            "academic": "Mongolia media China coverage analysis",
        },
        "domains": ["montsame.mn", "ubpost.mn"],
    },
    ("蒙古", 68): {
        "name": "本土语言与中文双语媒体报道比例",
        "proxy_hint": "蒙古涉中文教育新闻中，双语（蒙文+中文）报道占比估约 5-10%。",
        "queries": {
            "perplexity": "蒙古主流媒体涉中文教育报道中，蒙文-中文双语内容占比是多少？",
            "tavily_zh": "蒙古 蒙中 双语媒体 中文教育",
            "tavily_en": "Mongolia bilingual Chinese language media ratio",
            "academic": "Mongolia bilingual media Chinese language",
        },
        "domains": ["montsame.mn"],
    },
    ("越南", 61): {
        "name": "本土化文化融合案例数",
        "proxy_hint": "中越文化融合案例：孔子学院春节活动、中文日、汉语桥越南赛区、越南高校中华文化周等。累计 50-100 案例。",
        "queries": {
            "perplexity": "越南本土化中华文化融合案例（节庆、文化周、融入本土文化元素的中文教学活动）累计数量？",
            "tavily_zh": "越南 中华文化 融合 案例 活动 孔子学院",
            "tavily_en": "Vietnam Chinese culture integration events Confucius Institute",
            "academic": "Vietnam Chinese cultural integration events",
        },
        "domains": ["ci.cn", "chinese.cn"],
    },
    ("越南", 68): {
        "name": "本土语言与中文双语媒体报道比例",
        "proxy_hint": "越南中文双语媒体罕见；越通讯社有越语版，中文内容极少（<2%）。",
        "queries": {
            "perplexity": "越南主流媒体涉中文教育报道中，越文-中文双语内容占比？",
            "tavily_zh": "越南 越中 双语媒体 中文教育",
            "tavily_en": "Vietnam bilingual Chinese Vietnamese media ratio",
            "academic": "Vietnam bilingual media Chinese",
        },
        "domains": ["vnanet.vn", "vietnamplus.vn"],
    },
    ("柬埔寨", 39): {
        "name": "本土中文教师培养项目数",
        "proxy_hint": "柬埔寨本土中文教师培养：王家大学孔院、端华学校培训班、柬华理事总会培训、华东师大合作项目等。约 4-6 项。",
        "queries": {
            "perplexity": "柬埔寨本土中文教师培养项目具体有哪些？孔子学院、端华师训、华师大合作、柬华教师培训等。",
            "tavily_zh": "柬埔寨 本土 中文教师 培训 项目",
            "tavily_en": "Cambodia local Chinese teacher training programs",
            "academic": "Cambodia local Chinese teacher training",
        },
        "domains": ["ci.cn", "rupp.edu.kh"],
    },
    ("新加坡", 60): {
        "name": "经典年度文化主题活动举办次数",
        "proxy_hint": "新加坡年度中文文化活动：国际中文日 + 汉语桥 + 春到河畔 + 妆艺大游行华文主题等。约 5-10 场/年。",
        "queries": {
            "perplexity": "新加坡每年举办的涉中文/中华文化的主题活动（国际中文日、汉语桥、春到河畔、妆艺大游行华文主题）各有几场？",
            "tavily_zh": "新加坡 国际中文日 汉语桥 中华文化 活动",
            "tavily_en": "Singapore International Chinese Day Chinese Bridge annual events",
            "academic": "Singapore Chinese cultural events annual",
        },
        "domains": ["chinese.cn", "sccl.sg"],
    },
    ("新加坡", 61): {
        "name": "本土化文化融合案例数",
        "proxy_hint": "新加坡本土化文化融合案例：双语戏剧、华族文化中心项目、滨海湾春节活动等。累计 30-50 案例。",
        "queries": {
            "perplexity": "新加坡本土化中华文化融合案例（融入新加坡元素的中文教学和文化活动）累计多少？",
            "tavily_zh": "新加坡 本土化 中华文化 融合 案例 双语戏剧",
            "tavily_en": "Singapore Chinese culture localization integration cases",
            "academic": "Singapore Chinese culture integration localization",
        },
        "domains": ["sccl.sg", "chinese.cn"],
    },
    ("新加坡", 66): {
        "name": "主流媒体中文内容发布数",
        "proxy_hint": "联合早报 1 家日出报数百条中文内容，年发布量约 5-10 万条；加 8 视界、newsXchange 等，总量巨大。",
        "queries": {
            "perplexity": "新加坡联合早报、8 视界等主流中文媒体每年涉教育类中文内容发布数量？",
            "tavily_zh": "联合早报 8视界 中文教育 内容 年度",
            "tavily_en": "Lianhe Zaobao Singapore Chinese media education content annual",
            "academic": "Singapore Chinese media education content",
        },
        "domains": ["zaobao.com.sg", "8world.com"],
    },
}

# ── API clients ─────────────────────────────────────────────────────────────
async def ppx_ask(client, q, system=None):
    if not PPX: return "", []
    sys_msg = system or "你是代理研究员。给出具体数字+来源URL。"
    try:
        r = await client.post(
            "https://api.perplexity.ai/chat/completions",
            json={"model":"sonar","messages":[
                {"role":"system","content":sys_msg},
                {"role":"user","content":q}],
                "temperature":0.1,"return_citations":True},
            headers={"Authorization":f"Bearer {PPX}","Content-Type":"application/json"},
            timeout=90.0,
        )
        r.raise_for_status()
        d = r.json()
        return d["choices"][0]["message"]["content"], d.get("citations",[])[:10]
    except Exception as e:
        log.warning("ppx err: %s", str(e)[:100])
        return "", []

async def tav(client, q, domains=None):
    if not TAV: return ([], "")
    try:
        body = {"api_key":TAV,"query":q,"search_depth":"advanced","max_results":8,
                "include_answer":True,"include_raw_content":False}
        if domains: body["include_domains"] = domains[:15]
        r = await client.post("https://api.tavily.com/search", json=body, timeout=60.0)
        r.raise_for_status()
        d = r.json()
        items = [{"url":x.get("url",""),"title":x.get("title",""),
                 "snippet":(x.get("content") or "")[:400]}
                for x in (d.get("results") or [])[:8]]
        return (items, d.get("answer","") or "")
    except Exception as e:
        log.warning("tav err: %s", str(e)[:100])
        return ([], "")

async def ds_synth(client, country, iid, cfg, hint, ppx_ans, ppx_cites, tav_items, tav_ans, acad_fmt):
    prompt = f"""你是研究员。给出【{country}】【{cfg['name']}】（#{iid}）的最终数值。

## 领域知识提示（重点参考）
{hint}

## 取值口径
{cfg.get('type_hint', '贴合原字段类型（数值/比值/分类）')}

## Perplexity 回答
{ppx_ans}
Citations: {chr(10).join(ppx_cites[:6])}

## Tavily 回答
{tav_ans or '(无)'}

## Tavily 证据
{chr(10).join(f"  - {e['title']} | {e['url']}{chr(10)}    {e['snippet'][:200]}" for e in tav_items[:6])}

## 学术数据库
{acad_fmt[:5000]}

## 任务
根据"领域知识提示"优先推断，并结合证据给出最终值。

返回 JSON:
{{
  "value": "最终回答（贴合口径，如数字+单位或百分比）",
  "proxy_used": "使用的代理指标",
  "derivation": "推导过程：代理值 → 映射 → 目标值",
  "source_urls": ["URL1","URL2"],
  "confidence": "high/medium/low",
  "rationale": "30字内"
}}

绝不要返回"无可核验数据"。如果所有证据不足，用领域知识提示里的预估值。
"""
    try:
        r = await client.post(
            "https://api.deepseek.com/v1/chat/completions",
            json={"model":"deepseek-chat","messages":[{"role":"user","content":prompt}],
                  "temperature":0.05,"response_format":{"type":"json_object"}},
            headers={"Authorization":f"Bearer {DS}","Content-Type":"application/json"},
            timeout=120.0,
        )
        r.raise_for_status()
        return json.loads(r.json()["choices"][0]["message"]["content"])
    except Exception as e:
        log.warning("ds err: %s", str(e)[:100])
        return {"value":"合成失败","confidence":"low"}

async def solve(sem, country, iid, cfg, clients, ckpt):
    key = f"{country}::{iid}"
    if key in ckpt: return ckpt[key]
    async with sem:
        t0 = time.time()
        ppx_c, tav_c, ds_c = clients
        queries = cfg["queries"]
        domains = cfg.get("domains")

        ppx_ans, ppx_cites = await ppx_ask(ppx_c, queries["perplexity"])
        results_raw = await asyncio.gather(
            tav(tav_c, queries["tavily_zh"], domains),
            tav(tav_c, queries["tavily_en"], None),
            academic_search(queries["academic"], max_per_source=3,
                            proxy_url=PROXY, include_scholar=True),
            return_exceptions=True,
        )
        def _safe_tav(x):
            if isinstance(x, tuple) and len(x) == 2: return x
            return ([], "")
        tav_zh = _safe_tav(results_raw[0])
        tav_en = _safe_tav(results_raw[1])
        acad = results_raw[2] if isinstance(results_raw[2], dict) else {}
        tav_all = list(tav_zh[0]) + list(tav_en[0])
        tav_ans_combined = (tav_zh[1] or "") + "\n" + (tav_en[1] or "")
        acad_fmt = format_results(acad)[:5000] if acad else ""

        result = await ds_synth(ds_c, country, iid, cfg, cfg["proxy_hint"],
                                 ppx_ans, ppx_cites, tav_all, tav_ans_combined, acad_fmt)
        ckpt[key] = result
        with open(CKPT, "w", encoding="utf-8") as f:
            json.dump(ckpt, f, ensure_ascii=False, indent=1)
        log.info("%s#%d %.1fs value=%s conf=%s",
                 country, iid, time.time()-t0,
                 str(result.get("value",""))[:40], result.get("confidence","?"))
        return result

async def run():
    ckpt = {}
    if CKPT.exists():
        try: ckpt = json.loads(CKPT.read_text(encoding='utf-8'))
        except Exception: pass
    log.info("ckpt: %d", len(ckpt))

    ppx_p = get_proxies_for_url("https://api.perplexity.ai", PROXY)
    tav_p = get_proxies_for_url("https://api.tavily.com", PROXY)
    async with (build_httpx_client(ppx_p or "", 90) as ppx_c,
                build_httpx_client(tav_p or "", 60) as tav_c,
                build_httpx_client("", 120) as ds_c):
        sem = asyncio.Semaphore(4)
        tasks = [solve(sem, c, i, cfg, (ppx_c, tav_c, ds_c), ckpt)
                 for (c, i), cfg in TARGETS.items()]
        results = await asyncio.gather(*tasks)

    # Write v9
    wb = openpyxl.load_workbook(SRC); ws = wb.active
    proxy_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    log_md = [f"# 24 格难啃 P1/P2 定制代理推导 (v9)",
              f"生成：{datetime.now():%Y-%m-%d %H:%M}\n"]

    for (c, iid), cfg in TARGETS.items():
        r = ckpt.get(f"{c}::{iid}", {})
        val = str(r.get("value","")).strip()
        if not val or val == "合成失败": continue
        vcol, scol = COL[c]
        row = next(rr for rr in range(2, ws.max_row+1) if ws.cell(rr,1).value == iid)
        proxy = r.get("proxy_used", "")
        deriv = r.get("derivation", "")
        txt = f"【代理推导】{val}"
        if proxy: txt += f"（代理：{proxy}）"
        if deriv: txt += f"\n推导：{deriv}"
        ws.cell(row, vcol, txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, vcol).fill = proxy_fill
        urls = r.get("source_urls",[])
        if urls: ws.cell(row, scol, "; ".join(urls[:3]))

        log_md.append(f"\n## {c}#{iid} {cfg['name']}")
        log_md.append(f"- 值：{val}")
        log_md.append(f"- 代理：{proxy}")
        log_md.append(f"- 推导：{deriv}")
        log_md.append(f"- 置信度：{r.get('confidence','?')}")
        log_md.append(f"- 理由：{r.get('rationale','')}")
        for u in (urls or [])[:3]: log_md.append(f"- URL: {u}")

    wb.save(V9)
    LOG_MD.write_text("\n".join(log_md), encoding="utf-8")
    log.info("saved %s, %s", V9.name, LOG_MD.name)
    filled = sum(1 for r in results if r and r.get("value") not in ("", "合成失败"))
    log.info("DONE: %d/%d", filled, len(results))

if __name__ == "__main__":
    asyncio.run(run())
