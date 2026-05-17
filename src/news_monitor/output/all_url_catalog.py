"""Maintain a cumulative Excel catalog of article source domains.

The workbook stores unique first/second/third-level hostnames and keeps
category counts for China / Taiwan / US coverage. It also tracks processed
article-level keys in a hidden sheet so reruns do not double count.
"""
from __future__ import annotations

import hashlib
import logging
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook

from news_monitor.models import NewsArticle

logger = logging.getLogger("news_monitor.output.all_url_catalog")

_VISIBLE_SHEETS: dict[int, str] = {
    1: "一级域名",
    2: "二级域名",
    3: "三级域名",
}
_TRACKING_SHEET = "_processed_articles"
_HEADERS = [
    "域名",
    "美国篇数",
    "台湾篇数",
    "中国篇数",
    "其他篇数",
    "总篇数",
    "最近国家",
    "最近更新",
]
_CATEGORY_COLUMNS = {
    "us": 2,
    "taiwan": 3,
    "china": 4,
    "other": 5,
}
_COMMON_SLD = {
    "ac",
    "co",
    "com",
    "edu",
    "gov",
    "mil",
    "net",
    "org",
}


def update_all_url_excel(
    articles: list[NewsArticle],
    *,
    country_code: str,
    country_label: str,
    output_path: str = "output/allURL.xlsx",
) -> dict[str, int]:
    """Update the cumulative Excel catalog with a batch of articles."""
    if not articles:
        return {"new_domains": 0, "new_article_entries": 0}

    wb = _load_or_create_workbook(output_path)
    tracking = wb[_TRACKING_SHEET]
    processed = _load_processed_keys(tracking)
    row_indexes = {level: _build_sheet_index(wb[name]) for level, name in _VISIBLE_SHEETS.items()}

    new_domains = 0
    new_entries = 0
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for article in articles:
        category = _normalize_category(article.category)
        article_key = _article_key(article)
        for level, domain in _extract_domain_levels(article.source_url).items():
            if not domain:
                continue
            processed_key = f"{article_key}|L{level}|{domain}"
            if processed_key in processed:
                continue

            sheet = wb[_VISIBLE_SHEETS[level]]
            row_map = row_indexes[level]
            row = row_map.get(domain)
            if row is None:
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=domain)
                for col in range(2, 7):
                    sheet.cell(row=row, column=col, value=0)
                row_map[domain] = row
                new_domains += 1

            category_col = _CATEGORY_COLUMNS.get(category, _CATEGORY_COLUMNS["other"])
            sheet.cell(row=row, column=category_col, value=_int_cell(sheet, row, category_col) + 1)
            sheet.cell(row=row, column=6, value=_int_cell(sheet, row, 6) + 1)
            sheet.cell(row=row, column=7, value=country_label or country_code.upper())
            sheet.cell(row=row, column=8, value=now_text)

            tracking.append(
                [
                    processed_key,
                    article_key,
                    level,
                    domain,
                    category,
                    country_code,
                    article.source_url,
                    now_text,
                ]
            )
            processed.add(processed_key)
            new_entries += 1

    _autosize_visible_sheets(wb)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(
        "Updated allURL Excel: path=%s new_domains=%d new_entries=%d",
        output_path,
        new_domains,
        new_entries,
    )
    return {"new_domains": new_domains, "new_article_entries": new_entries}


def _load_or_create_workbook(output_path: str):
    path = Path(output_path)
    if path.exists():
        wb = load_workbook(path)
        for level, name in _VISIBLE_SHEETS.items():
            if name not in wb.sheetnames:
                ws = wb.create_sheet(title=name)
                _init_visible_sheet(ws)
        if _TRACKING_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(title=_TRACKING_SHEET)
            _init_tracking_sheet(ws)
            ws.sheet_state = "hidden"
        return wb

    wb = Workbook()
    ws = wb.active
    ws.title = _VISIBLE_SHEETS[1]
    _init_visible_sheet(ws)
    for level in (2, 3):
        extra = wb.create_sheet(title=_VISIBLE_SHEETS[level])
        _init_visible_sheet(extra)
    tracking = wb.create_sheet(title=_TRACKING_SHEET)
    _init_tracking_sheet(tracking)
    tracking.sheet_state = "hidden"
    return wb


def _init_visible_sheet(ws) -> None:
    ws.append(_HEADERS)
    ws.freeze_panes = "A2"


def _init_tracking_sheet(ws) -> None:
    ws.append(["processed_key", "article_key", "level", "domain", "category", "country", "source_url", "updated_at"])


def _load_processed_keys(ws) -> set[str]:
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        if key:
            seen.add(str(key))
    return seen


def _build_sheet_index(ws) -> dict[str, int]:
    out: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        domain = ws.cell(row=row, column=1).value
        if domain:
            out[str(domain).strip().lower()] = row
    return out


def _autosize_visible_sheets(wb) -> None:
    for name in _VISIBLE_SHEETS.values():
        ws = wb[name]
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 22


def _int_cell(ws, row: int, col: int) -> int:
    value = ws.cell(row=row, column=col).value
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _article_key(article: NewsArticle) -> str:
    raw = (article.fingerprint or "").strip()
    if raw:
        return raw
    url = (article.source_url or "").strip().lower()
    if not url:
        url = (article.title or "").strip().lower()
    return hashlib.sha256(url.encode("utf-8")).hexdigest()


def _normalize_category(category: str) -> str:
    value = (category or "").strip().lower()
    if value in {"us", "taiwan", "china"}:
        return value
    return "other"


def _extract_domain_levels(url: str) -> dict[int, str]:
    host = _hostname_from_url(url)
    if not host:
        return {}

    labels = [part for part in host.split(".") if part]
    if len(labels) < 2:
        return {1: host}

    suffix_len = _public_suffix_length(labels)
    root_size = min(len(labels), suffix_len + 1)
    levels: dict[int, str] = {}
    for level in (1, 2, 3):
        size = root_size + (level - 1)
        if len(labels) >= size:
            levels[level] = ".".join(labels[-size:])
    return levels


def _hostname_from_url(url: str) -> str:
    if not url:
        return ""
    raw = url.strip()
    if "://" not in raw:
        raw = f"https://{raw}"
    try:
        parsed = urlparse(raw)
    except ValueError:
        return ""
    return (parsed.hostname or "").strip().lower().rstrip(".")


def _public_suffix_length(labels: list[str]) -> int:
    if len(labels) < 2:
        return 1
    if len(labels[-1]) == 2 and len(labels) >= 3 and labels[-2] in _COMMON_SLD:
        return 2
    return 1
